"""
RF Rotator
Copyright 2013 Travis Fagerness
v2.0 update by Mike Collins
"""
import sys, os, random,csv,time
from PyQt4.QtCore import *
from PyQt4.QtGui import *

import multiprocessing,logging

#TODO change for actual 3d plotting
import matplotlib
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt4agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm

import openpyxl as pyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image

# from openpyxl.styles import *
# from openpyxl.drawing.image import *

from SignalHound import *
from worker import *
from setup import *
from specan import *
from arcus import *
#from openpyxl import *


from Calibrator import Calibrator

import numpy as np
import math
import time
import datetime

from pip._vendor.requests.packages.chardet.latin1prober import FREQ_CAT_NUM

from SignalHound.bb_api_h import BB_TIME_GATE


##
#===============================================================================
# adjust matplotlib display settings
#===============================================================================
matplotlib.rcParams.update({'font.size': 8})

version = "2.0"
year = "2017"
author = "Travis Fagerness v2.0 update by Mike Collins"
website = "http://www.nextgenrf.com"
email = "mike.collins@nextgenrf.com"
#General TODOs===============================================================================
# TODO: add save plot functionality to save current tabs plot
# 

#===============================================================================


#set colors for plots
XCOLOR="#00BB00"
YCOLOR="#f48642"
ZCOLOR="#0000FF"

class AppForm(QMainWindow):#create main application
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
           
        #function and Variable Ddeclaration   
        self.threads=[]         #create empty list for threads
        self.legend=""          #create empy list for legend
        self.rotationAxis='Z'   #set default rotation axis for data collection
        
        self.csvLegend=['Z (Raw)','Z (Cal)','X (Raw)','X (Cal)','Y (Raw)','Y (Cal)']
        #=======================================================================
        # setup data collection variables
        #=======================================================================
        self.data=np.array([1,2,3]) #holds raw data
        
        self.zRawData=[]        # holds raw z axis data
        self.xRawData=[]        # holds raw x axis data
        self.yRawData=[]        # holds raw y axis data
        
        self.zCalData=[]        # holds calibrated z axis data
        self.xCalData=[]        # holds calibrated x axis data
        self.yCalData=[]        # holds calibrated y axis data
        
        self.angles=[]          # holds angle data
        
        self.color=ZCOLOR          # setup plot color for easy diferentiation
        self.pltColor=ZCOLOR
        
        self.data_available=False
        self.deviceIsConnected=False
        
        #create calibrator object
        self.cal=Calibrator()
        self.cal.set_mainForm(self)       #set calibrators pointer to self
        
        #=======================================================================
        # setup EMC testing tab
        #=======================================================================
        self.emc_regs='FCC'                                     #select regulation set for testing
        self.emc_class='A'                                      #select class of emc testing
        
        self.emc_testComplete=False                             #hold wheather an EMC test has been run or not
        
        #=======================================================================
        # setup 3D Rendering Variables
        #=======================================================================
        self.render3D=False                                     #is false until application has rendered in 3D 
        
        
        #==================================================
        #setup main window
        #==================================================
        self.setWindowTitle('Rotation RX Strength Plotter')
        self.create_menu()                                      #create menu Qwidget
        
        self.create_tabs()                                      #create tab object to hold application tabs(data collection, calibration, 3d rendering)
        
        self.create_dataCollectionTab(self.tab_dataCollection)  #create data collection tabs
        
        self.create_emcTab(self.tab_emc)
        
        #calibrator object creates it's own tab
        self.cal.create_calibrationTab(self.tab_calibration)
        
        self.create_3dTab()                                     #create 3D rendering tab
        
        
        self.create_status_bar()                                #create status bar at bottom of app
        self.textbox.setText('1 2 3 4')
        
        
        
        #==================================================
        #create worker object
        #==================================================
        self.worker=Worker()
        self.worker.set_cal(self.cal)                   #give worker access to calibrator
        self.manual_mode=False
        
        #set threading to run worker at same time as this object
        self.threads.append(self.worker)
        
        #worker setup
        self.worker.status_msg.connect(self.status_text.setText)
        self.worker.data_pair.connect(self.on_data_ready)
        self.worker.dev_found.connect(self.device_found)
        self.worker.worker_sleep.connect(self.worker_asleep)
        self.worker.set_cal(self.cal)                   #pass the calibrator to the worker
        self.worker.start()
          
        #=======================================================================
        # create setup dialog box object
        #=======================================================================
        self.setup = Setup(self,self.worker,self.cal)   #create setup object for worker object
        
        self.worker.set_setup(self.setup)               #pass the setup params to the worker
        
        #=======================================================================
        # setup worker and setup access for calibrator
        #=======================================================================
        self.cal.set_setup(self.setup)
        self.cal.set_worker(self.worker)

        #TODO: fix mpl = multiprocessing.log_to_stderr(logging.CRITICAL)#
    
    def worker_asleep(self):#worker wating for command
        #=======================================================================
        #          Name:    worker_asleep
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function enables/disables the appropriate
        #                    pushbuttons when the worker object is awake/asleep
        #
        #=======================================================================
        'enable/disable GUI buttons for when worker is asleep'
        
        #if the worker is asleep (not paused) the rotation table should be at home
        if self.deviceIsConnected:
            self.b_start.setEnabled(not self.manual_mode)
            self.b_manual.setEnabled(True)
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.cal.b_specan.setEnabled(True)
            self.rb_axisSelZ.setEnabled(True)
            self.rb_axisSelX.setEnabled(True)
            self.rb_axisSelY.setEnabled(True)
            
            #display specan type in calibration tab
            self.cal.gui_specan.setText(self.worker.specan.device)
            
        else:
            self.cal.b_specan.setEnabled(False)
            self.b_start.setEnabled(False)
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_manual.setEnabled(False)
            self.rb_axisSelZ.setEnabled(False)
            self.rb_axisSelX.setEnabled(False)
            self.rb_axisSelY.setEnabled(False)
            
            #display specan type in calibration tab
            self.cal.gui_specan.setText("--Spectrum analyzer not detected--")
        
    def device_found(self,devices=[False,'Not Found','Not Found']):
        #=======================================================================
        #          Name:    device_found
        #
        #    Parameters:    devices
        #
        #        Return:    None
        #
        #   Description:    this function sets if the spectrum analyzer AND turntable are found
        #
        #=======================================================================
        'set if specan AND turntable are found'
        self.deviceIsConnected=devices[0]
    
    def save_report(self):#create .xlsx report file
        #=======================================================================
        #          Name:    save_report
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function saves the collected data and the test info
        #                    is .xlsx format
        #        
        #                    this functions uses the openpyxl library Heavily
        #
        #=======================================================================
        'save current plot data as .csv'

        file_choices = "Excel Workbook ( *.xlsx)"
        path = unicode(QFileDialog.getSaveFileName(self, 
                        'Save', '', 
                        file_choices))
        
        #========================================================================
        # make data arrays the same size for export to csv
        #========================================================================
        self.fill_dataArray()
       
            #===================================================================
            # save Report to .xlsx
            #===================================================================
        if path:
            #===================================================================
            # create styles for automatic reporting
            #===================================================================
            
            #data style
            style_data = NamedStyle(name="style_data")
            thinbd = Side(style='thin', color="000000")
            thickbd = Side(style='medium', color="000000")
            style_data.border = Border(left=thinbd, right=thinbd)
            style_data.alignment=Alignment(horizontal="right")
            
            #pass style
            style_pass = NamedStyle(name="style_pass")
            style_pass.font = Font(bold=True, size=14,color="009B00")
            style_pass.border = Border(left=thinbd, right=thinbd)
            style_pass.alignment=Alignment(horizontal="right")
            
            #fail style
            style_fail = NamedStyle(name="style_fail")
            style_fail.font = Font(bold=True, size=14,color="FF0000")
            style_fail.border = Border(left=thinbd, right=thinbd)
            style_fail.alignment=Alignment(horizontal="right")
            
            #top header style
            style_headerTop = NamedStyle(name="style_headerTop")
            style_headerTop.font = Font(bold=False, size=12)
            style_headerTop.border = Border(left=thinbd, top=thinbd, right=thinbd)
            style_headerTop.alignment=Alignment(horizontal="center")
            style_headerTop.fill=PatternFill("solid", fgColor="DDDDDD")
            
            #left header style
            style_headerLeft = NamedStyle(name="style_headerLeft")
            style_headerLeft.font = Font(bold=False, size=12)
            style_headerLeft.border = Border(left=thinbd,right=thinbd)
            style_headerLeft.alignment=Alignment(horizontal="left")
            style_headerLeft.fill=PatternFill("solid", fgColor="DDDDDD")
            
            #Title style
            style_title = NamedStyle(name="style_title")
            style_title.font = Font(bold=True, size=14,color="FFFFFF")
            style_title.alignment=Alignment(horizontal="center",vertical="center")
            style_title.fill=PatternFill("solid", fgColor="555555")
            
            #===================================================================
            # initialize workbook to save as .xlsx
            #===================================================================
            
            wb = pyxl.Workbook()

#===============================================================================
# DATA SHEET
#===============================================================================
            #create new worksheet in first position
            ws = wb.create_sheet("Overview & Data", 0) # 
            
            #setup variable locations for data for easy formatting during design
            DATA_HEIGHT =10
            SETUP_HEIGHT = 10
            #===================================================================
            # Create informations cells
            #===================================================================
            
            ws.merge_cells('A1:D1')
            ws.row_dimensions[1].height = 50
            ws['A1']= 'Gain Testing Report'
            ws['A1'].style=style_title
            ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
            
            # add NGRFlogo
            img = Image('images/ngrf.png')
            ws.add_image(img, 'F1')
            
            #create date cells
            ws['A2']= 'Date:'
            ws['A2'].style=style_headerLeft
            ws["B2"]= datetime.date.today()
            ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
            
            #Create Customer Cells
            ws['A3']= 'Customer:'
            ws['A3'].style=style_headerLeft
            ws["B3"]= self.cal.cal_customer
            
            #Create tester name Cells
            ws['A4']= 'Tested By:'
            ws['A4'].style=style_headerLeft
            ws["B4"]= self.cal.cal_tester
            
            #Create comment Cells
            ws['A5']= 'Comments:'
            ws['A5'].style=style_headerLeft
            ws["B5"]= self.cal.cal_comments
            
            #Create DUT label Cells
            ws['C2']= 'DUT Label:'
            ws['C2'].style=style_headerLeft
            ws["D2"]= self.cal.cal_dutLabel
            
            #Create DUT serial number Cells
            ws['C3']= 'DUT Serial Number:'
            ws['C3'].style=style_headerLeft
            ws["D3"]= self.cal.cal_dutSN
            
            #===================================================================
            # Write data and angles to xlsx file
            #===================================================================
            
            ws.column_dimensions['A'].width = 20
            ws['A'+str(DATA_HEIGHT)]= "Angle (degrees)"
            ws['A'+str(DATA_HEIGHT)].style=style_headerTop
            i=DATA_HEIGHT+1
            for angle in self.angles:
                ws['A'+str(i)] = angle
                ws['A'+str(i)].number_format = '0.00E+00'
                ws['A'+str(i)].style=style_data
                i=i+1
            
            ws.column_dimensions['B'].width = 20
            ws['B7']= self.csvLegend[0]
            ws['B7'].style=style_headerTop
            i=11
            for zraw in self.zRawData:
                ws['B'+str(i)] = zraw
                ws['B'+str(i)].style=style_data
                i=i+1  
            
            ws.column_dimensions['C'].width = 20    
            ws['C7']= self.csvLegend[1]
            ws['C7'].style=style_headerTop
            i=11
            for zcal in self.zCalData:
                ws['C'+str(i)] = zcal
                ws['C'+str(i)].style=style_data
                i=i+1 
            
            ws.column_dimensions['D'].width = 20
            ws['D7']= self.csvLegend[2]
            ws['D7'].style=style_headerTop
            i=11
            for xraw in self.xRawData:
                ws['D'+str(i)] = xraw
                ws['D'+str(i)].style=style_data
                i=i+1
            
            ws.column_dimensions['E'].width = 20    
            ws['E7']= self.csvLegend[3]
            ws['E7'].style=style_headerTop
            i=11
            for xcal in self.xCalData:
                ws['E'+str(i)] = xcal
                ws['E'+str(i)].style=style_data
                i=i+1     
                
            ws.column_dimensions['F'].width = 20
            ws['F7']= self.csvLegend[4]
            ws['F7'].style=style_headerTop
            i=11
            for yraw in self.yRawData:
                ws['F'+str(i)] = yraw
                ws['F'+str(i)].style=style_data
                i=i+1 
            
            ws.column_dimensions['G'].width = 20
            ws['G7']= self.csvLegend[5]
            ws['G7'].style=style_headerTop
            i=11
            for ycal in self.yCalData:
                ws['G'+str(i)] = ycal
                ws['G'+str(i)].style=style_data
                i=i+1 
            
            #===================================================================
            # Write Max/averages and headers
            #===================================================================
            ws.merge_cells('A6:G6')
            ws['A6']="Test Data"
            ws['A6'].style=style_title
            
            #extra cell to look pretty
            ws['A7']=""
            ws['A7'].style=style_title
            
            #create max value cells
            ws['A8']= 'Max Gain (dBi):'
            ws['A8'].style=style_headerLeft

            #create average value cells
            ws['A9']= 'Average Gain (dBi):'
            ws['A9'].style=style_headerLeft
            
            ws['B8'] = "=MAX(B11:B111)"
            ws['B8'].style=style_data
            ws['B9'] = "=AVERAGE(B11:B111)"
            ws['B9'].style=style_data
            #insert blank cells
            ws['B10']='-'
            ws['B10'].style=style_headerTop
            
            ws['C8'] = "=MAX(C11:C111)"
            ws['C8'].style=style_data
            ws['C9'] = "=AVERAGE(C11:C111)"
            ws['C9'].style=style_data
            #insert blank cells
            ws['C10']='-'
            ws['C10'].style=style_headerTop
            
            ws['D8'] = "=MAX(D11:D111)"
            ws['D8'].style=style_data
            ws['D9'] = "=AVERAGE(D11:D111)"
            ws['D9'].style=style_data
            #insert blank cells
            ws['D10']='-'
            ws['D10'].style=style_headerTop
            
            ws['E8'] = "=MAX(E11:E111)"
            ws['E8'].style=style_data
            ws['E9'] = "=AVERAGE(E11:E111)"
            ws['E9'].style=style_data
            #insert blank cells
            ws['E10']='-'
            ws['E10'].style=style_headerTop
            
            ws['F8'] = "=MAX(F11:F111)"
            ws['F8'].style=style_data
            ws['F9'] = "=AVERAGE(F11:F111)"
            ws['F9'].style=style_data
            #insert blank cells
            ws['F10']='-'
            ws['F10'].style=style_headerTop
            
            ws['G8'] = "=MAX(G11:G111)"
            ws['G8'].style=style_data
            ws['G9'] = "=AVERAGE(G11:G111)"
            ws['G9'].style=style_data
            #insert blank cells
            ws['G10']='-'
            ws['G10'].style=style_headerTop
            
            #===================================================================
            # Create setup / Calibration data area
            #===================================================================
            
            #title
            ws.merge_cells('I6:J6')
            ws['I6']="Test Setup"
            ws['I6'].style=style_title
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 20
            
            #Frequency
            ws['I7']="Test Frequency (Hz):"
            ws['I7'].style=style_headerLeft
            ws['J7']=float(self.cal.cal_freq)
            ws['J7'].style=style_data
            
            #Frequency span
            ws['I8']="Frequency Span (Hz):"
            ws['I8'].style=style_headerLeft
            ws['J8']=float(self.cal.cal_span)
            ws['J8'].style=style_data
            
            #distance
            ws['I9']="Testing Distance (m):"
            ws['I9'].style=style_headerLeft
            ws['J9']=float(self.cal.cal_dist)
            ws['J9'].style=style_data
            
            #distance
            ws['I10']="Sweep Time (s):"
            ws['I10'].style=style_headerLeft
            ws['J10']=float(self.cal.cal_sc_sweepTime)
            ws['J10'].style=style_data
            
            #calibration title
            ws.merge_cells('I11:J11')
            ws['I11']="Calibration"
            ws['I11'].style=style_title
            
            #Input power
            ws['I12']="Input Power (dBm):"
            ws['I12'].style=style_headerLeft
            ws['J12']=float(self.cal.cal_inputPwr)
            ws['J12'].style=style_data
            
            #amplifer gain
            ws['I13']="PreAmplifier Gain (dB):"
            ws['I13'].style=style_headerLeft
            ws['J13']=float(self.cal.cal_ampGain)
            ws['J13'].style=style_data
            
            #tx cable loss
            ws['I14']="Tx Cable Loss (dB):"
            ws['I14'].style=style_headerLeft
            ws['J14']=float(self.cal.cal_txCableLoss)
            ws['J14'].style=style_data
            
            #DUT Gain
            ws['I15']="DUT Gain (dBi):"
            ws['I15'].style=style_headerLeft
            ws['J15']=float(self.cal.cal_txGain)
            ws['J15'].style=style_data
            
            #fspl
            ws['I16']="FSPL (dB):"
            ws['I16'].style=style_headerLeft
            ws['J16']=float(self.cal.cal_fspl)
            ws['J16'].style=style_data
            
            #rx antenna gain
            ws['I17']="Calibrated Antenna Gain (dBi):"
            ws['I17'].style=style_headerLeft
            ws['J17']=float(self.cal.cal_rxGain)
            ws['J17'].style=style_data
            
            #Rx cable Loss
            ws['I18']="Rx Cable Loss (dB):"
            ws['I18'].style=style_headerLeft
            ws['J18']=float(self.cal.cal_rxCableLoss)
            ws['J18'].style=style_data
            
            #additional elements
            it=0;
            for i in self.cal.addGainLoss:
                ws['I'+str(19+it)]=str(i)+"(dB):"
                ws['I'+str(19+it)].style=style_headerLeft
                ws['J'+str(19+it)]=float(self.cal.addGainLoss[i])
                ws['J'+str(19+it)].style=style_data
                it=it+1
                
            #Total Cal value
            ws['I'+str(19+it)]="Total (dB):"
            ws['I'+str(19+it)].style=style_headerLeft
            ws['J'+str(19+it)]=float(-self.cal.calibrate_data(0))
            ws['J'+str(19+it)].style=style_data
            
            
#===============================================================================
# data collection Plot Sheet
#===============================================================================

            #create new worksheet in 2nd position
            ws = wb.create_sheet("Data Plots", 1) # 
            
            ws.merge_cells('A1:D1')
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.row_dimensions[1].height = 50
            ws['A1']= 'Data Collection Plots'
            ws['A1'].style=style_title
            ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
            
            #create date cells
            ws['A2']= 'Date:'
            ws['A2'].style=style_headerLeft
            ws["B2"]= datetime.date.today()
            ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
            
            #Create DUT label Cells
            ws['A3']= 'DUT Label:'
            ws['A3'].style=style_headerLeft
            ws["B3"]= self.cal.cal_dutLabel
            ws['B3'].style=style_data
            
            #Create DUT serial number Cells
            ws['A4']= 'DUT Serial Number:'
            ws['A4'].style=style_headerLeft
            ws["B4"]= self.cal.cal_dutSN
            ws['B4'].style=style_data
            
            # add NGRFlogo
            img = Image('images/ngrf.png')
            ws.add_image(img, 'F1')
            

            self.canvas.print_figure("temp_fig1.png", dpi=100,facecolor=self.figEmc.get_facecolor())
            img = Image('temp_fig1.png')
            ws.add_image(img, 'A5')
            
#===============================================================================
# EMC Compliance Plot Sheet
#===============================================================================
#create new worksheet in 3rd position if EMC test has been run
            if(self.emc_testComplete):
                ws = wb.create_sheet("EMC Compliance", 2) 
                ws.merge_cells('A1:E1')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.row_dimensions[1].height = 50
                
                ws['A1']= 'EMC Compliance Test Results'
                ws['A1'].style=style_title
                ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
                
                # add NGRFlogo
                img = Image('images/ngrf.png')
                ws.add_image(img, 'D3')
                
                #===============================================================
                # add test information
                #===============================================================
                #create date cells
                ws['A2']= 'Date:'
                ws['A2'].style=style_headerLeft
                ws["B2"]= datetime.date.today()
                ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
                
                #Create DUT label Cells
                ws['A3']= 'DUT Label:'
                ws['A3'].style=style_headerLeft
                ws["B3"]= self.cal.cal_dutLabel
                ws['B3'].style=style_data
                
                #Create DUT serial number Cells
                ws['A4']= 'DUT Serial Number:'
                ws['A4'].style=style_headerLeft
                ws["B4"]= self.cal.cal_dutSN
                ws['B4'].style=style_data
                
                #regs
                ws['A5']="Regulatory Convention:"
                ws['A5'].style=style_headerLeft
                ws['B5']=self.emc_regs
                ws['B5'].style=style_data
                
                #class
                ws['A6']="Class:"
                ws['A6'].style=style_headerLeft
                ws['B6']=self.emc_class
                ws['B6'].style=style_data
                
                #limit
                ws['A7']="Electric Field Limit (dBuV/m):"
                ws['A7'].style=style_headerLeft
                ws['B7']=float(self.get_emcTestLimit(self.cal.cal_freq))
                ws['B7'].style=style_data
                
                #Test Status
                ws['A8']="Test Status:"
                ws['A8'].style=style_headerLeft
                ws['B8']=self.run_emcTest();
                if ws['B8'].value=="Pass":
                    ws['B8'].style=style_pass
                else:
                    ws['B8'].style=style_fail
                
                
                
                #Test Status
                ws['A9']="Error Margin (dB):"
                ws['A9'].style=style_headerLeft
                ws['B9']=float(self.e_emc_margin.text());
                ws['B9'].style=style_data
                

                self.emcCanvas.print_figure("temp_fig2.png", dpi=100, facecolor=self.figEmc.get_facecolor())
                img = Image('temp_fig2.png')
                ws.add_image(img, 'A10')
            
#===============================================================================
# 3D radiation Pattern Sheett
#===============================================================================
#create new worksheet in 3rd position if EMC test has been run
            if(self.render3D):
                
                #if emc test is run put in 4th position else 3rd
                if(self.emc_testComplete):
                    ws = wb.create_sheet("3D Radiation Pattern", 3) 
                else:
                    ws = wb.create_sheet("3D Radiation Pattern", 2) 
                    
                ws.merge_cells('A1:E1')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.row_dimensions[1].height = 50
                
                ws['A1']= '3D Radiation Pattern'
                ws['A1'].style=style_title
                ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
                
                #create date cells
                ws['A2']= 'Date:'
                ws['A2'].style=style_headerLeft
                ws["B2"]= datetime.date.today()
                ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
                
                #Create DUT label Cells
                ws['A3']= 'DUT Label:'
                ws['A3'].style=style_headerLeft
                ws["B3"]= self.cal.cal_dutLabel
                ws['B3'].style=style_data
                
                #Create DUT serial number Cells
                ws['A4']= 'DUT Serial Number:'
                ws['A4'].style=style_headerLeft
                ws["B4"]= self.cal.cal_dutSN
                ws['B4'].style=style_data
                
                # add NGRFlogo
                img = Image('images/ngrf.png')
                ws.add_image(img, 'G1')
                
                #add 3D radiation pattern plot
                self.canvas3d.print_figure("temp_fig3.png", dpi=100, facecolor=self.figEmc.get_facecolor())
                img = Image('temp_fig3.png')
                ws.add_image(img, 'A5')
            #===================================================================
            # save .xlsx file
            #===================================================================
            
            wb.save(path)
                  
    def open_report(self):#open previous test fro .xlsx file
        #=======================================================================
        #          Name:    open_report
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    opens a previous test from .xlsx file
        #
        #=======================================================================
        'open .xlsx file of previous test'
        
        
        from openpyxl import load_workbook#needed to open a workbook
        
        
        #=======================================================================
        # open file
        #=======================================================================
        file_choices = "Excel Workbook ( *.xlsx)"
        
        #open file 
        path = unicode(QFileDialog.getOpenFileName(self, 
                        'Open', '', 
                        file_choices))
        if path:
            wb2 = load_workbook(path)   #load openpyxl workbook
            print wb2.get_sheet_names() #show sheet names
            names=wb2.get_sheet_names() #get name of data sheet
            ws=wb2[str(names[0])];      #set active worksheet to data sheet
            
            #===================================================================
            # load data from .xlsx file
            #===================================================================
            startingVal=11
            for i in range(0,101):
                self.angles[i] = float(ws['A'+str(i+startingVal)].value)
                
            for i in range(0,101):
                self.zRawData[i] = float(ws['B'+str(i+startingVal)].value)
            
            for i in range(0,101):
                self.zCalData[i] = float(ws['C'+str(i+startingVal)].value)
            for i in range(0,101):
                self.xRawData[i] = float(ws['D'+str(i+startingVal)].value)
            
            for i in range(0,101):
                self.xCalData[i] = float(ws['E'+str(i+startingVal)].value)
            
            for i in range(0,101):
                self.yRawData[i] = float(ws['F'+str(i+startingVal)].value)
            
            for i in range(0,101):
                self.yCalData[i] = float(ws['G'+str(i+startingVal)].value)
            
            #print opened data
            print self.angles
            print self.zRawData
            print self.zCalData
            print self.xRawData
            print self.xCalData
            print self.yRawData
            print self.yCalData
            
            self.statusBar().showMessage('Opened file %s' % path, 2000)

            #===================================================================
            # draw data plots with opened data
            #===================================================================
            self.legend = ws['C7'].value        
            self.rb_axisSelZ.click()            #set draw axis to Z
            self.data=np.array(self.zCalData)   #add z axis data to draw array
            self.draw_dataPlots()               #draw data collection plots
            
            self.legend = ws['E7'].value
            self.rb_axisSelX.click()
            self.data=np.array(self.xCalData)            
            self.draw_dataPlots()
            
            self.legend = ws['G7'].value
            self.rb_axisSelY.click()
            self.data=np.array(self.yCalData) 
            self.draw_dataPlots()
            
            #reset data array
            self.data=[]
            
    def save_plot(self):
        #=======================================================================
        #          Name:    save_plot
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    saves the data collection plots as a .png file
        #
        #=======================================================================
        'Saves plot as .png'
        file_choices = "PNG *.png"
        
        path = unicode(QFileDialog.getSaveFileName(self, 
                        'Save file', '', 
                        file_choices))
        if path:
            self.canvas.print_figure(path, dpi=self.dpi)
            self.statusBar().showMessage('Saved to %s' % path, 2000)
    
    def on_about(self):#display program information
        #=======================================================================
        #          Name:    on_about
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function displays applicaiton version info when 
        #                    the user selects "about" from the menu
        #
        #=======================================================================
        'calls about data'
        
        msg = "NGRF Rotator\r\n"\
                + "Version: " + version + "\r\n"\
                + "Author: " + author + "\r\n"\
                + "Contact: " + email + "\r\n"\
                + "Copyright " + year + "\r\n"\
                + website
        QMessageBox.about(self, "About", msg.strip())
    
    def click_manualTarget(self, event):#sets turntable target and begins test
        #=======================================================================
        #
        #          Name:    click_manualTarget
        #
        #    Parameters:    event(mouse click event)
        #
        #        Return:    None
        #
        #   Description:    this function will begin a manual test when the user
        #                    clicks a point of the data collection plot in manual mode
        #
        #=======================================================================
        """
        Uses the button_press_event to begin manual mode test
        """
        if self.manual_mode:
            print event.xdata
            print event.ydata
            worker_data=[event.xdata*180/3.14]
            self.worker.do_work(self.worker.Functions.goto_location,worker_data)

    def click_setup(self):#activates setup dialog
        #=======================================================================
        #
        #          Name:    click_setup
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function executes the setup dialog box when the user clicks the 
        #                    "setup" pushbutton
        #
        #=======================================================================
        'initiates setup dialog'
        
        self.setup.exec_()
        #=======================================================================
        # update line edit boxes in setup dialog
        #=======================================================================
        self.setup.e_span.setText(str(self.cal.cal_span/1e6))  
        self.setup.e_cfreq.setText(str(self.cal.cal_freq/1e6))
        self.setup.e_sweep.setText(str(self.cal.cal_sc_sweepTime*1e3))
   
    def on_data_ready(self,new_data):#sends raw data to data lists and starts drawing plots
        #=======================================================================
        #
        #          Name:    on_data_ready
        #
        #    Parameters:    new_data
        #
        #        Return:    None
        #
        #   Description:    this function adds new data to data collection arrays
        #                    when new data is available from the spectrum analyzer
        #                    this functions also redraws the selected data collection plot
        #                    when new data become available
        #
        #=======================================================================
        'adds new data to dat arrays, calls to redraw plot'
        #=======================================================================
        # append new data to appropriate array
        #=======================================================================
        if (self.rotationAxis=='Z'):
            self.zRawData.append(new_data[1])
        elif(self.rotationAxis=='X'):
            self.xRawData.append(new_data[1])
        elif(self.rotationAxis=='Y'):
            self.yRawData.append(new_data[1])
            
        #===================================================================
        # create arrays for drawing plot
        #===================================================================
        
        self.angles.append(360-new_data[0])                     #subtract turntable angle from 360 to get correct angle for plot
        
        self.data.append(self.cal.calibrate_data(new_data[1]))  #XXX:calibrate data and append it to drawing array
        self.progress.setValue(new_data[0])
        
        if (self.rotationAxis=='Z'):
            self.zCalData.append(self.cal.calibrate_data(new_data[1]))
        elif(self.rotationAxis=='X'):
            self.xCalData.append(self.cal.calibrate_data(new_data[1]))
        elif(self.rotationAxis=='Y'):
            self.yCalData.append(self.cal.calibrate_data(new_data[1]))
        
        self.draw_dataPlots()#draw new data to graph
            
    def click_start(self):#begin test
        #=======================================================================
        #
        #          Name:    click_start
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function begins the test when the user clicks
        #                    the "start" push button
        #
        #=======================================================================
        'begin test'
        
        #=======================================================================
        # enable/disabkle gui buttons during test
        #=======================================================================
        self.b_pause.setEnabled(True)
        self.b_stop.setEnabled(True)
        self.b_start.setEnabled(False)
        self.rb_axisSelZ.setEnabled(False)
        self.rb_axisSelX.setEnabled(False)
        self.rb_axisSelY.setEnabled(False)
        self.cal.b_specan.setEnabled(False)
        
        #=======================================================================
        # apply settings to specan for test
        #=======================================================================
        
        self.cal.apply_specanSettings()
        
        #=======================================================================
        # get name of plotfordisplay
        #=======================================================================
        text, ok = QInputDialog.getText(self, 'Name of data', 
            'Enter a data name:')
        if ok:
            self.legend=str(text)
        else:
            self.click_stop()
            return
        #=======================================================================
        # clear arrays that will store axis data
        #=======================================================================
        self.data=[]
        self.angles=[]
        self.worker.do_work(self.worker.Functions.rotate)

        if (self.rotationAxis=='Z'):
            self.zRawData=[]
            self.zCalData=[]

            #set legend for .csv output
            del self.csvLegend[0]
            self.csvLegend.insert(0, str(text)+" (Raw)")
            del self.csvLegend[1]
            self.csvLegend.insert(1, str(text)+" (Calibrated)")
                
        elif(self.rotationAxis=='X'):
            self.xRawData=[]
            self.xCalData=[]
            
            #set legend for .csv output
            del self.csvLegend[2]
            self.csvLegend.insert(2, str(text)+" (Raw)")
            del self.csvLegend[3]
            self.csvLegend.insert(3,str(text)+" (Calibrated)")
            
        elif(self.rotationAxis=='Y'):
            self.yRawData=[]
            self.yCalData=[]
            
            #set legend for .csv output
            del self.csvLegend[4]
            self.csvLegend.insert(4, str(text)+" (Raw)")
            del self.csvLegend[5]
            self.csvLegend.insert(5,str(text)+" (Calibrated)")
        
    def click_stop(self):#abort current test
        #=======================================================================
        #          Name:    click_stop
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    this function manages the functionality when the 
        #                    user clicks the stop buttons, it disable/enable the appropriate
        #                    pushbuttons and change the state of the worker object.
        #                    the test cannot be started again from the same point
        #
        #=======================================================================
        'abort current test'
        
        self.b_pause.setEnabled(False)
        self.b_stop.setEnabled(False)
        self.b_start.setEnabled(True)
        self.cal.b_specan.setEnabled(True)
        self.b_manual.setEnabled(True)
        self.rb_axisSelZ.setEnabled(True)
        self.rb_axisSelX.setEnabled(True)
        self.rb_axisSelY.setEnabled(True)
        self.worker.cancel_work=True
        
    def click_pause(self):#pause mid-test without reseting data
        #=======================================================================
        #          Name:    click_pause
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This function manages the functionality when the user
        #                    clicks the pause push button, it will stop the turn-table mid test
        #                    and change the state of the worker object
        #
        #=======================================================================
        'Pause mid-test'
        self.b_stop.setEnabled(not self.b_pause.isChecked())            
        self.worker.pause_work(self.b_pause.isChecked())
    
    def click_manual(self):#activates manual mode
        #=======================================================================
        #          Name:    click_manual
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function puts the application in and out of manual mode
        #                    and set disables/enables the appropriate push buttons
        #
        #=======================================================================
        'switch to manual mode'
        self.manual_mode=self.b_manual.isChecked()
        if self.manual_mode:
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_start.setEnabled(False)
            self.rb_axisSelZ.setEnabled(False)
            self.rb_axisSelX.setEnabled(False)
            self.rb_axisSelY.setEnabled(False)
        else:
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_start.setEnabled(True)
            self.rb_axisSelZ.setEnabled(True)
            self.rb_axisSelX.setEnabled(True)
            self.rb_axisSelY.setEnabled(True)
            self.b_manual.setEnabled(True)
    
    def click_clear(self):#clears data from active axis list and plot
        #=======================================================================
        #          Name:    click_clear
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function clears the data from the selected axis plot
        #                    when the user clicks the clear button
        #
        #=======================================================================
        'clears data arrays and resest plot data'
        
        self.data=[]    #reset raw data list
        self.angles=[]  #reset angles list
        
        #clear data arrays
        if (self.rotationAxis=='Z'):
            self.zRawData=[]
        elif(self.rotationAxis=='X'):
            self.xRawData=[]
        elif(self.rotationAxis=='Y'):
            self.yRawData=[]
            
        self.legend=""
        self.axes.clear()
        self.axes.grid(self.grid_cb.isChecked())
        self.axes.set_title(self.rotationAxis+'-axis',color=self.color)
        self.canvas.draw() 
            
    def update_plot_settings(self):
        #=======================================================================
        #          Name:    update_plot_settings
        #
        #    Parameters:    None
        #
        #        Return:    None    
        #
        #   Description:    this function redraws the main data collection plot 
        #                    with or without the grid lines depending on user selection
        #
        #=======================================================================
        'update grid setting for Plot'
        self.axes.grid(self.grid_cb.isChecked())
        self.canvas.draw()
    
    def click_axisSelect(self):#select rotation axis for test
        #=======================================================================
        #          Name:    click_axisSelect
        #    
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function manages the functionality when the user
        #                    clicks one of the axis select radio buttons
        #
        #=======================================================================
        'select rotation axis for test'
        
        self.pltColor='000000'
        
        if (self.rb_axisSelX.isChecked()):
            self.rotationAxis='X'
            self.axes=self.x_axis
            self.pltColor=XCOLOR
            self.color=XCOLOR
            
        elif(self.rb_axisSelY.isChecked()):
            self.rotationAxis='Y'
            self.axes=self.y_axis
            self.pltColor=YCOLOR
            self.color=YCOLOR
        else:
            self.rotationAxis='Z'
            self.axes=self.z_axis
            self.pltColor=ZCOLOR
            self.color=ZCOLOR
            
        #change curAxis label and format text
        self.curAxis.setText('<span style=" font-size:14pt; font-weight:600; color:'+self.pltColor+';">Current Rotation Axis: ' + str(self.rotationAxis)+'</span>')
        
        print "Current Rotation Axis: " + self.rotationAxis
        
    def draw_dataPlots(self):#draw plots
        #=======================================================================
        #          Name:    draw_dataPlots
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This functions draws the main data collection plots on 
        #                    the data collection tab
        #
        #=======================================================================
        """ Redraws the figure
        """
        # clear the axes and redraw the plot anew
        self.axes.clear()        
        self.axes.grid(self.grid_cb.isChecked())
        self.axes.set_title(self.legend,fontsize=14,fontweight=300)
        
        r = np.array(self.data)#[:,1]
        theta = np.array(self.angles) * np.pi / 180
        self.axes.plot(theta, r, lw=2,color=self.color)
        
        #set up grid for plot
        gridmin=10*round(np.amin(r)/10)
        if gridmin>np.amin(r):
            gridmin = gridmin-10
        gridmax=10*round(np.amax(r)/10)
        if gridmax < np.amax(r):
                gridmax=gridmax+10
        self.axes.set_ylim(gridmin,gridmax)
        self.axes.set_yticks(np.arange(gridmin,gridmax,(gridmax-gridmin)/5))
        
        #create legend for plot
        leg = self.axes.legend([self.legend], loc=(-.1,-.2))
        
        leg.draggable(True)
        self.canvas.draw()

    def create_tabs(self):#create tab architecture for application
        #=======================================================================
        #          Name:    create_tabs
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function creates theQTabWidget that holds the 
        #                    tabs of the application, it also populates with the application's tabs 
        #
        #=======================================================================

        #create tab widget to hold tabs
        #self.t_bar=QTabBar(shape=QTabBar.TriangularEast)
        self.tabs=QTabWidget()
        
        #create data collection tab
        
        self.tab_dataCollection= QWidget()
        self.tabs.addTab(self.tab_dataCollection,"Data Collection")
        
        #create calibration tab
        self.tab_calibration=QWidget()
        self.tabs.addTab(self.tab_calibration,"Calibration")
        
        #create emc testing tab
        self.tab_emc=QWidget()
        self.tabs.addTab(self.tab_emc,"EMC")
        
        #create 3d imaging tab
        self.tab_3D=QWidget()
        self.tabs.addTab(self.tab_3D,"3D Rendering")
        
        self.setCentralWidget(self.tabs)
        
    def create_dataCollectionTab(self,tab):#create data collection tab as well as main window
        #=======================================================================
        #          Name:    create_dataCollectionTab
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function creates the form and user interface of the data
        #                    collection tab
        #
        #=======================================================================
        
        self.main_frame = QWidget()
        
        #==========================================================================
        #create Label for current axis
        #===========================================================================
        
        self.curAxis=QLabel('<span style=" font-size:14pt; font-weight:600;color:'+ZCOLOR+'">Current Rotation Axis: Z</span>')
        self.curAxis.setAlignment(Qt.AlignLeft)
        
        #=======================================================================
        # create figure and canvas for data collection plots
        #=======================================================================
        self.dpi = 100
        self.fig = Figure(dpi=self.dpi)
        self.fig.set_facecolor('#DDDDDD')
        self.canvas = FigureCanvas(self.fig)
        

        self.canvas.setParent(self.tab_dataCollection)
        
        #=======================================================================
        # create subplots for data plots
        #=======================================================================
        self.z_axis = self.fig.add_subplot(131,polar=True)
        self.z_axis.set_title('X-Y Plane (Rotation on Z-Axis)',color=ZCOLOR,fontsize=12,fontweight=300)
        self.x_axis = self.fig.add_subplot(132,polar=True)
        self.x_axis.set_title('Y-Z Plane (Rotation on X-Axis)',color=XCOLOR,fontsize=12,fontweight=300)
        self.y_axis = self.fig.add_subplot(133,polar=True)
        self.y_axis.set_title('X-Z Plane (Rotation on Y-Axis)',color=YCOLOR,fontsize=12,fontweight=300)
        
        self.axes=self.z_axis#set current axis to axes variable

        #=======================================================================
        # create buttons and GUI controls
        #=======================================================================
        
        # Bind the 'button_press_event' event for clicking on one of the bars
        self.canvas.mpl_connect('button_press_event', self.click_manualTarget)
        
        # Create the navigation toolbar, tied to the canvas
        self.mpl_toolbar = NavigationToolbar(self.canvas, self.tab_dataCollection)
        
        # Other GUI controls
        self.b_setup = QPushButton("&Setup/Find Devices")
        self.connect(self.b_setup, SIGNAL('clicked()'), self.click_setup)
        self.b_setup.setToolTip("Setup tools for test")
        
        
        self.b_manual= QPushButton("&Manual Mode",enabled=False,checkable=True)
        self.b_manual.setEnabled(False)
        self.connect(self.b_manual, SIGNAL('clicked()'), self.click_manual)
        self.b_manual.setToolTip("Move table to specific point while continuously performing test")
        
        self.b_start= QPushButton("&Rotate Start")
        self.b_start.setEnabled(False)
        self.connect(self.b_start, SIGNAL('clicked()'), self.click_start)
        self.b_start.setToolTip("Begin Test")
        
        
        self.b_stop= QPushButton("Stop/&Home",enabled=False)
        self.connect(self.b_stop, SIGNAL('clicked()'), self.click_stop)
        self.b_stop.setToolTip("Abort test and return to home position")
        
        self.b_pause= QPushButton("&Pause",enabled=False,checkable=True)
        self.connect(self.b_pause, SIGNAL('clicked()'), self.click_pause)
        self.b_pause.setToolTip("Pause current test")
        
        self.b_reset= QPushButton("&Clear",enabled=True)
        self.connect(self.b_reset, SIGNAL('clicked()'), self.click_clear)
        self.b_reset.setToolTip("Clear data plot from active rotation axis")
        
        self.textbox = QLineEdit()
        self.textbox.setMinimumWidth(200)
        self.connect(self.textbox, SIGNAL('editingFinished ()'), self.draw_dataPlots)
        
        self.draw_button = QPushButton("&Draw")
        self.connect(self.draw_button, SIGNAL('clicked()'), self.draw_dataPlots)
        
        
        self.grid_cb = QCheckBox("Show &Grid",checked=True)
        self.connect(self.grid_cb, SIGNAL('stateChanged(int)'), self.update_plot_settings)
        self.b_reset.setToolTip("Show grid on active axis?")
        
        #====================================================================================
        #Create rotation axis selection controls
        #=================================================================================
    
        axisVbox=QVBoxLayout()
        axisVbox.addWidget(QLabel("Select Axis"))
        axisHbox=QHBoxLayout()
        
        self.rb_axisSelZ=QRadioButton('Z')      #create axis select radio buttons
        self.rb_axisSelZ.click()                #set Z axis to default axis select radio button
        self.rb_axisSelX=QRadioButton('X')      #create axis select radio buttons
        self.rb_axisSelY=QRadioButton('Y')      #create axis select radio buttons
        axisHbox.addWidget(self.rb_axisSelZ)
        axisHbox.addWidget(self.rb_axisSelX)
        axisHbox.addWidget(self.rb_axisSelY)
        axisVbox.addLayout(axisHbox)
        
        #connect buttons
        self.connect(self.rb_axisSelZ, SIGNAL('clicked()'), self.click_axisSelect)
        self.connect(self.rb_axisSelX, SIGNAL('clicked()'), self.click_axisSelect)
        self.connect(self.rb_axisSelY, SIGNAL('clicked()'), self.click_axisSelect)
        
        #set radio button tool tips
        self.rb_axisSelZ.setToolTip("Switch to Z axis")
        self.rb_axisSelX.setToolTip("Switch to Y axis")
        self.rb_axisSelY.setToolTip("Switch to X axis")
        
        progess_label = QLabel("Rotation Progress:")
        self.progress = QProgressBar()
        self.progress.setAlignment = Qt.Horizontal
        self.progress.setMaximum(360)
        self.progress.setMinimum(0)
        
        
        #===============================================================================
        # Layout with box sizers
        # ==============================================================================
        hbox = QHBoxLayout()
        
        #=======================================================================
        # create button bar
        #=======================================================================
        hbox.addLayout(axisVbox)
        for w in [  self.b_setup,self.b_manual, self.b_start,self.b_stop,self.b_pause,self.b_reset, self.grid_cb,
                    progess_label, self.progress]:
            hbox.addWidget(w)
            hbox.setAlignment(w, Qt.AlignVCenter)
            
        
        vbox = QVBoxLayout()                #create layout    
        vbox.addWidget(self.curAxis)        #add current rotation axis display label    
        vbox.addWidget(self.canvas,10)      #add graph area to display
        vbox.addWidget(self.mpl_toolbar)    #add matplotlib toolbar to display
        vbox.addLayout(hbox)                #add control buttons to display
        
        #add layout to tab
        self.tab_dataCollection.setLayout(vbox)
        
    def create_3dTab(self):#create 3d rendering tab
        #=======================================================================
        #          Name:    create_3dTab
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    this function create the form and user interface of the 
        #                   3d plotting tab
        #
        #=======================================================================
        
        #==========================================================================
        #create Label for current axis V2.0
        #===========================================================================
        
        self.l_3d=QLabel()
        self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:blue;">Render 3D Interpolation of Radiation Pattern</span>')
        self.l_3d.setAlignment(Qt.AlignLeft)
        
        #=======================================================================
        # create figure and canvas for 3d rendering
        #=======================================================================
        self.fig3d = Figure(figsize=(6.0, 6.0), dpi=self.dpi)
        self.canvas3d = FigureCanvas(self.fig3d)
        self.fig3d.set_facecolor('#8E8E8E')
        #=======================================================================
        # self.canvas.setParent(self.main_frame)
        #=======================================================================
        self.canvas3d.setParent(self.tab_3D)
    
        #=======================================================================
        # create "Render" button
        #=======================================================================
        self.b_render= QPushButton("&Render")
        self.b_render.setEnabled(True)
        self.connect(self.b_render, SIGNAL('clicked()'), self.draw_3dPlot)
        self.b_render.setToolTip("render collected data in 3D, NOTE: will only work when all 3 axes have data")
        
        #=======================================================================
        # create subplots
        #=======================================================================
        self.plt3dx = self.fig3d.add_subplot(132,projection='3d', aspect='equal')
        self.plt3dy = self.fig3d.add_subplot(133,projection='3d', aspect='equal')
        self.plt3dz = self.fig3d.add_subplot(131,projection='3d', aspect='equal')
        
        #===================================================================
        # create colobars to show field strength
        #===================================================================
        m = matplotlib.cm.ScalarMappable(cmap=matplotlib.cm.jet)
        m.set_array(self.data)
        
        self.cbarx = self.fig3d.colorbar(m,ax=self.plt3dx, orientation="horizontal")
        self.cbarx.set_label('Power (dBm)',)
        
        self.cbary = self.fig3d.colorbar(m,ax=self.plt3dy, orientation="horizontal")
        self.cbary.set_label('Power (dBm)',)
        
        self.cbarz = self.fig3d.colorbar(m,ax=self.plt3dz, orientation="horizontal")
        self.cbarz.set_label('Power (dBm)',)
        
        self.canvas3d.draw()
        
        hbox3d = QHBoxLayout()
        
        #=======================================================================
        # create button bar
        #=======================================================================
        hbox3d.addStretch()
        for w in [self.b_render]:
            hbox3d.addWidget(w)
            hbox3d.setAlignment(w, Qt.AlignVCenter)
        hbox3d.addStretch()
            
        vbox3d = QVBoxLayout()              #create layout 
        vbox3d.addWidget(self.l_3d)     
        vbox3d.addWidget(self.canvas3d,10)  #add graph area to display
        vbox3d.addWidget(self.mpl_toolbar)  #add matplotlib toolbar to display
        vbox3d.addLayout(hbox3d)            #add control buttons to display
        
        self.tab_3D.setLayout(vbox3d)
        
    def draw_3dPlot(self):#Draws 3D interpolation of data
        #=======================================================================
        #          Name:    draw_3dPlot
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    This function draws the 3d plot of the collected data
        #
        #=======================================================================
        'Draws 3D interpolation of data'
        
        self.b_render.setEnabled(False)             #disable button while rendering
        
        self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; black;">Rendering..........</span>')
        
        #initialize variables for interpolation and drawing
        theta, phi = np.linspace(0, 2 * np.pi, 100), np.linspace(0, np.pi, 50)

        THETA, PHI = np.meshgrid(theta, phi)
        
        R = np.ones((50,100)) 
        
        #=======================================================================
        #check if data arrays have any test data in them
        #if they are empty the app will not render them
        #=======================================================================
        
        zhasData=False
        xhasData=False
        yhasData=False
        
        for i in self.zCalData:
            if(i!=0):                   #set draw to true if array has data
                zhasData=True
        for i in self.xCalData:
            if(i!=0):                   #set draw to true if array has data
                xhasData=True
        for i in self.yCalData:
            if(i!=0):                   #set draw to true if array has data
                yhasData=True  
         
    #======================================================================
    # create 3Dimage from 3 axes projections
    #======================================================================
        if(zhasData and xhasData and yhasData):#only draw if all axes have data
            for th in range(100):
                for rh in range(50):
                        
                    if th<25:
                        if rh<25:
                            
                            xzInterp = self.interp(rh,0,self.xCalData[75+rh],24,self.zCalData[100-th])
                             
                            yzInterp = self.interp(rh,0,self.yCalData[75+rh],24,self.zCalData[100-th])
                             
                            xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
                            
                        else:
                            xzInterp = self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[rh-25])
                            
                            yzInterp = self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[rh-25])
                            
                            xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
                     
                    elif th<50:
                        if rh<25:
                             
                            xzInterp=self.interp(rh,0,self.xCalData[75-rh],24,self.zCalData[100-th])
                             
                            yzInterp=self.interp(rh,0,self.yCalData[75+rh],24,self.zCalData[100-th])
                             
                            xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
                        else:
                             
                            xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[75-rh])
                             
                            yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[rh-25])
                             
                            xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
                      
                    elif th<75:
                        if rh<25:
                             
                            xzInterp=self.interp(rh,0,self.xCalData[75-rh],24,self.zCalData[100-th])
                             
                            yzInterp=self.interp(rh,0,self.yCalData[75-rh],24,self.zCalData[100-th])
                             
                            xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
                             
                        else:
                            xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[75-rh])
                             
                            yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[75-rh])
                             
                            xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
                       
                    else:
                        if rh<25:
                            xzInterp=self.interp(rh,0,self.xCalData[75+rh],24,self.zCalData[100-th])
                            
                            yzInterp=self.interp(rh,0,self.yCalData[75-rh],24,self.zCalData[100-th])
                            
                            xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
                            
                        else:
                            
                            xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[rh-25])
                            
                            yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[75-rh])
                            
                            xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
                    #Radius is equal to this interpolation         
                    R[rh,th]=xyInterp

            #create unchanged copy of Radius array for colorbars
            cbR=R.copy()
            
            #add lowest value to all Radius array elements so all values will be >=0
            R+=abs(np.amin(R))
            
            #change array values from spherical to cartesian and multiply by radius
            X = R * np.sin(PHI) * np.cos(THETA)
            Y = R * np.sin(PHI) * np.sin(THETA)
            Z = R * np.cos(PHI)
            
            #set color of plot to change based on value of radius
            strengthColor = matplotlib.cm.jet((R)/np.amax(R))
            
            #===================================================================
            # draw 3D plots
            #===================================================================
            self.plt3dx.cla()
            self.plt3dy.cla()
            self.plt3dz.cla()
            
            # X Axis
            self.plt3dx.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dx.view_init(elev=0, azim=270)    

            # Y axis
            self.plt3dy.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dy.view_init(elev=0, azim=0)   
               
            # Z Axis
            self.plt3dz.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dz.view_init(elev=90, azim=270)     
            
            #Ensure 3D plots are square so images is not distorted
            self.axisEqual3D(self.plt3dx)
            self.axisEqual3D(self.plt3dy)
            self.axisEqual3D(self.plt3dz)
            
            self.plt3dx.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            self.plt3dx.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            self.plt3dx.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            #set 3d plot titles
            self.plt3dx.set_title('Y-Z Plane (Rotation on X-Axis)',y=1, size=12)    
            self.plt3dy.set_title('X-Z Plane (Rotation on Y-Axis)',y=1, size=12)
            self.plt3dz.set_title('X-Y Plane (Rotation on Z-Axis)',y=1, size=12)
            
            #set plot axis labels
            self.plt3dx.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dx.set_ylabel("X")
            self.plt3dx.set_zlabel("Z")
            
            self.plt3dy.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dy.set_ylabel("X")
            self.plt3dy.set_zlabel("Z")
            
            self.plt3dz.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dz.set_ylabel("X")
            self.plt3dz.set_zlabel("Z")
            
            #remove ticks
            self.plt3dx.set_xticks([], [])
            self.plt3dx.set_yticks([], [])
            self.plt3dx.set_zticks([], [])
            
            self.plt3dy.set_xticks([], [])
            self.plt3dy.set_yticks([], [])
            self.plt3dy.set_zticks([], [])
            
            self.plt3dz.set_xticks([], [])
            self.plt3dz.set_yticks([], [])
            self.plt3dz.set_zticks([], [])
            
            #===================================================================
            # update colobars
            #===================================================================
            
            #set color values for colorbar
            m = cm.ScalarMappable(cmap=matplotlib.cm.jet)
            m.set_array(cbR)
            
            #x colorbar
            self.cbarx.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbarx.draw_all()

            #y colorbar
            self.cbary.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbary.draw_all()
            
            #z colorbar
            self.cbarz.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbarz.draw_all()
            
            #set tight layout and draw plots
            self.fig3d.tight_layout()
            self.canvas3d.draw()
            
            #===================================================================
            # warning/ information label text
            #===================================================================
            self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:blue;">3D Interpolation of Radiation Pattern</span>')
        else:
            self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:red;">Unable to plot! One or more data arrays are empty!</span>')
            
        
        self.b_render.setEnabled(True)      #enable button after rendering
        
        self.render3D=True                  #set render 3d to true so 3D plot will be added to test report
    
    def interp(self,x,x1,y1,x3,y3): 
        #=======================================================================
        #
        #          Name:    interp
        #
        #    Parameters:    (float) x, x1, y1, x3, y3
        #
        #        Return:    (float) retval
        #
        #   Description:    returns the linear interpolation at point x between points (x1,y1) and (x3,y3)
        #
        #=======================================================================
        retval=(((x-x1)*(y3-y1))/(x3-x1))+y1
        return retval
    
    def axisEqual3D(self,ax):
        #=======================================================================
        #
        #          Name:    axisEqual3D
        #
        #    Parameters:    (subplot of matplotlib figure) ax
        #
        #        Return:    None
        #
        #   Description:    THis function sets the axes equal in a 3D matplotlib plot
        #                    so the image wont be distorted
        #
        #=======================================================================
        extents = np.array([getattr(ax, 'get_{}lim'.format(dim))() for dim in 'xyz'])
        sz = extents[:,1] - extents[:,0]
        centers = np.mean(extents, axis=1)
        maxsize = max(abs(sz))
        r = maxsize/2
        for ctr, dim in zip(centers, 'xyz'):
            getattr(ax, 'set_{}lim'.format(dim))(ctr - r, ctr + r)
            
            
    def create_emcTab(self,tab):#create EMC testing tab
        #=======================================================================
        #          Name:    create_emcTab
        #
        #    Parameters:    (pointer to tab object)tab
        #
        #        Return:    None
        #
        #   Description:    this function creates the form and user interface of the EMC testing Tab
        #
        #=======================================================================
        'create EMC pre compliance testing tab'
        
        #set tab style
        tab.setStyleSheet(self.cal.create_styleSheet('dataTab'))
        
        #=======================================================================
        # create test result plot
        #=======================================================================

        self.fill_dataArray()#populate data arrays
            
        self.figEmc = Figure()
        self.emcCanvas = FigureCanvas(self.figEmc)
        
        self.emcCanvas.setParent(self.tab_emc)
        self.emcPlot=self.figEmc.add_subplot(111)
        self.emcPlot.set_title('EMC Compliance Testing\nField Strength vs. Angle',fontsize=14,fontweight=200)
        self.figEmc.tight_layout()
        
        #=======================================================================
        # create warning label groupbox
        #=======================================================================
        
        #create distance warning message area
        self.emc_distWarning=QLabel('<span style="  color:lime; font-size:11pt; font-weight:400;">--Distance Good-- Ready to run test</span>')
        
        #create frequency warning message area
        self.emc_freqWarning=QLabel('<span style="  color:lime; font-size:11pt; font-weight:400;">--Frequency Good-- Ready to run test</span>')
        
        #create Far Feild warning message area
        self.emc_farFieldWarning=QLabel('<span style="  color:lime; font-size:11pt; font-weight:400;">--Far-Field Good-- Ready to run test</span>')
        #=======================================================================
        # Create "Run Test" pushbutton
        #=======================================================================
        self.b_run_test= QPushButton("&Run Test")
        self.b_run_test.setEnabled(True)
        self.connect(self.b_run_test, SIGNAL('clicked()'), self.run_emcTest)
        self.b_run_test.setToolTip("Run EMC pre-compliance test on collected data")
        
        #=======================================================================
        # create Regulation selection area
        #=======================================================================
        
        #select regulations (FCC/CISPR)
        self.regs=QGroupBox("Select Compliance Testing Regulations")
        regVbox=QVBoxLayout()
        regVbox.setAlignment(Qt.AlignCenter)
        rHbox=QHBoxLayout()
        
        self.r_fcc=QRadioButton('FCC')
        self.r_cispr=QRadioButton('CISPR')
        
        rHbox.addWidget(self.r_fcc)
        rHbox.addWidget(self.r_cispr)
        self.connect(self.r_fcc, SIGNAL('clicked()'), self.set_emcRegulations)
        self.connect(self.r_cispr, SIGNAL('clicked()'), self.set_emcRegulations)
        self.r_fcc.setToolTip("Select radiation regulations")
        
        self.r_cispr.setToolTip("Select radiation regulations")
        regVbox.addLayout(rHbox)
        self.regs.setLayout(regVbox)
        
        #select class (A/B)
        classbox=QGroupBox("Select DUT Class")
        cVbox=QVBoxLayout()
        cVbox.setAlignment(Qt.AlignCenter)
        cHbox=QHBoxLayout()
        
        self.r_classA=QRadioButton('Class A')
        self.r_classB=QRadioButton('Class B')
        
        cHbox.addWidget(self.r_classA)
        cHbox.addWidget(self.r_classB)
        
        self.connect(self.r_classA, SIGNAL('clicked()'), self.set_emcRegulations)
        self.connect(self.r_classB, SIGNAL('clicked()'), self.set_emcRegulations)
        
        self.r_classA.setToolTip("Select radiation regulations")
        self.r_classB.setToolTip("Select radiation regulations")
        cVbox.addLayout(cHbox)
        classbox.setLayout(cVbox)
        #=======================================================================
        # create Test settings groupbox
        #=======================================================================
        lfbox=QFormLayout()
        lfbox.setAlignment(Qt.AlignCenter)
        #Left form
        
        #=======================================================================
        # create test information groupbox
        #=======================================================================
        infoBox=QGroupBox("Test Information")
        infoBoxLayout=QFormLayout()
        infoBox.setLayout(infoBoxLayout)
        self.emc_gui_freq=QLabel("")
        infoBoxLayout.addRow(QLabel("Test Frequency (MHz):"),self.emc_gui_freq)
        self.emc_gui_dist=QLabel("")
        infoBoxLayout.addRow(QLabel("Test Distance (m):"),self.emc_gui_dist)
        self.emc_gui_farField=QLabel("")
        infoBoxLayout.addRow(QLabel("Far Field (m):"),self.emc_gui_farField)
        self.emc_gui_limit=QLabel("")
        infoBoxLayout.addRow(QLabel("Electrical Field Strength Limit (dBuV/m):"),self.emc_gui_limit)
        
        lfbox.addRow(infoBox)
        
        #=======================================================================
        # create error margin groupbox
        #=======================================================================
        errorMarginBox=QGroupBox("Enter Error Margin")
        errorMarginBoxLayout=QFormLayout()
        errorMarginBox.setLayout(errorMarginBoxLayout)
        self.e_emc_margin  =QLineEdit('0')
        errorMarginBoxLayout.addRow(QLabel("Error Margin (dB)"),self.e_emc_margin)
        lfbox.addRow(errorMarginBox);
        
        #=======================================================================
        # create far field groupbox
        #=======================================================================
        farFieldBox=QGroupBox("Far Field Check")
        farFieldBoxLayout=QFormLayout()
        farFieldBox.setLayout(farFieldBoxLayout)
        
        self.e_emc_diameter  =QLineEdit('0')
        farFieldBoxLayout.addRow(QLabel("Max Diameter of DUT (m)"),self.e_emc_diameter)
        self.e_emc_diameter.returnPressed.connect(self.set_emcRegulations)
        
        b_checkFarfield = QPushButton("Check")
        b_checkFarfield.clicked.connect(self.set_emcRegulations)
        b_checkFarfield.setToolTip("check far field to ensure testing is valid")
        farFieldBoxLayout.addRow(QLabel("Check Farfield"),b_checkFarfield)
        
        lfbox.addRow(farFieldBox)
        #add regulations to form layout
        lfbox.addRow(self.regs)
        #add class selection to form layout
        lfbox.addRow(classbox)

        #=======================================================================
        # create test results Group box
        #=======================================================================
        #format style of test results
        self.emc_testResults=QLabel('<span style="  color:yellow; font-size:16pt; font-weight:400;">-----No Test Data-----</span>')
        self.emc_testResults.setAlignment(Qt.AlignCenter)
        
        #setup result layout
        resultsBox=QGroupBox("Test Results")
        resultsBox.setStyleSheet(self.cal.create_styleSheet('EMC2'))#apply styling
        resultsBoxLayout=QVBoxLayout()
        resultsBox.setLayout(resultsBoxLayout)
        resultsBoxLayout.addWidget(self.emc_testResults)
        
        #=======================================================================
        # set some defaults for EMC tab 
        #=======================================================================
        #set default radio buttons select to default
        self.r_classB.click()       #set button to default
        self.r_fcc.click()          #set button to default
        
        #set regulations to default so warnings display correct message
        self.set_emcRegulations()

        #=======================================================================
        # Create Layout for EMC Testing
        #=======================================================================
        
        vbox=QVBoxLayout()
        vbox.setAlignment(Qt.AlignCenter)
        vbox.setAlignment(Qt.AlignTop)
        
        #create warning groupbox
        warningBox=QGroupBox("Warnings")
        warningBox.setStyleSheet(self.cal.create_styleSheet('EMC2'))#apply styling
        warningBoxLayout=QVBoxLayout()
        warningBox.setLayout(warningBoxLayout)
        warningBoxLayout.addWidget(self.emc_distWarning)
        warningBoxLayout.addWidget(self.emc_farFieldWarning)
        warningBoxLayout.addWidget(self.emc_freqWarning)
        vbox.addWidget(warningBox)

        #create horizontal box to hold setup gui and plot
        hbox=QHBoxLayout()
        
        #create settings group box
        settingBox=QGroupBox("Test Settings")
        settingBox.setLayout(lfbox)
        settingBox.setStyleSheet(self.cal.create_styleSheet('setup'))#apply styling
        
        #create vertical box to hold settings
        settingsVBox=QVBoxLayout();
        
        #populate settings vertical box
        settingsVBox.addWidget(settingBox)
        settingsVBox.addWidget(self.b_run_test)
        
        #populate horizontal box
        hbox.addLayout(settingsVBox)
        hbox.addStretch()
        hbox.addWidget(self.emcCanvas)
        
        #add horizontal to tab layout
        vbox.addLayout(hbox)
        
        #add test results to tab layout
        vbox.addWidget(resultsBox)
        
        #set tab layout
        self.tab_emc.setLayout(vbox)
        
    def get_emcTestLimit(self,target):#return the max field strength in uV/m type='fcc' or 'CISPR' target = target frequency in Hz
        
        #=======================================================================
        #          Name:    get_emcTestLimit
        #
        #    Parameters:    Target(frequency in Hz)
        #
        #        Return:    (float) value of FCC or CISPR limit for target frequency
        #
        #   Description:    given the test frequency and regulation type
        #                   this function returns the maximum legal EMC in dBuV/m
        #
        #=======================================================================
        
        retval=0
        
        if (self.emc_regs=='FCC'):#return FCC values
            if self.emc_class=='A':
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @3 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 49.5
                    elif target<216e6:      #88-216MHz
                        retval = 54
                    elif target<906e6:      #216-906MHz
                        retval = 56.5
                    else:                   #906MHz-40GHz
                        retval = 60
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @10 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 40
                    elif target<216e6:      #88-216MHz
                        retval = 43.5
                    elif target<906e6:      #216-906MHz
                        retval = 46
                    else:                   #906MHz-40GHz
                        retval = 54
                if(self.cal.cal_dist==30):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @30 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 29.5
                    elif target<216e6:      #88-216MHz
                        retval = 34
                    elif target<906e6:      #216-906MHz
                        retval = 36.5
                    else:                   #906MHz-40GHz
                        retval = 40
            else:
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # FCC Class B Values in dBuV/m @3 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 40
                    elif target<216e6:      #88-216MHz
                        retval = 43.5
                    elif target<906e6:      #88-906MHz
                        retval = 46
                    else:                   #906MHz-40GHz
                        retval = 54
                        
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # FCC Class B Values in dBuV/m @10 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 29.5
                    elif target<216e6:      #88-216MHz
                        retval = 33
                    elif target<906e6:      #216-906MHz
                        retval = 35.5
                    else:                   #906MHz-40GHz
                        retval = 43.5
        else:#return CISPR values
            if self.emc_class=='A':
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # CISPR Class A Max Values in dBuV/m @ 10 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 40
                    else:                   #>230MHz
                        retval = 47
                if(self.cal.cal_dist==30):
                    #===============================================================
                    # CISPR Class A Max Values in dBuV/m @ 30 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 30
                    else:                   #>230MHz
                        retval = 37
            else:
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # CISPR Class B Max Values in dBuV/m @ 10 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 30
                    else:                   #>230MHz
                        retval = 37
            
        print 'retval '+ str(retval)
        
        return retval
    
    def run_emcTest(self):#run EMC Test
        #=======================================================================
        #    Name:            run_emcTest
        #
        #    Parameters:      None
        #
        #    Return:          None  
        #
        #    Description:     tests all data arrays and shows results in graph and 
        #                     in results label
        #
        #=======================================================================
        print 'Running EMC TEST\n  -VALUES-\nTarget: ' +str(self.cal.cal_freq)+'\nUpper Margin: '+str(float(self.e_emc_margin.text()))
        print "running test"
        
        #set margin from user input
        margin = float(self.e_emc_margin.text())
        
        self.fill_dataArray()               #fill data array for plotting
        
        self.b_run_test.setEnabled(False)   #disable run test button while testing
        
        #clear the axes and redraw the plot anew
        self.emcPlot.clear()
        testVal=(self.get_emcTestLimit(self.cal.cal_freq))
        
        #=======================================================================
        # setup plot data
        #=======================================================================
        a=np.array(self.angles)*np.pi/180
        
        #create temporary arrays to hold field strengths
        ztemp=[]
        xtemp=[]
        ytemp=[]
        
        #these variables prevent data from plotting if array is only zeros
        zdraw=False
        xdraw=False
        ydraw=False
        
        #hasData will tell the test if there is any data in data arrays
        hasData=False
        
        #build temporary arrays of field strength
        for i in self.zCalData:
            ztemp.append(self.get_fieldStrength_dBuVm(i))
            if(i!=0):                                       #set draw to true if array has data
                zdraw=True
                hasData=True
        for i in self.xCalData:
            xtemp.append(self.get_fieldStrength_dBuVm(i))
            if(i!=0):                                       #set draw to true if array has data
                xdraw=True
                hasData=True
        for i in self.yCalData:
            ytemp.append(self.get_fieldStrength_dBuVm(i)) 
            if(i!=0):                                       #set draw to true if array has data
                ydraw=True
                hasData=True  
        
        #create numpy arrays for plotting
        z=np.array(ztemp)      
        x=np.array(xtemp)
        y=np.array(ytemp)
        
        #reset tep arrays to hold data + error margin values 
        ztemp=[]
        xtemp=[]
        ytemp=[]
         
         
        for i in self.zCalData:
            ztemp.append(self.get_fieldStrength_dBuVm(i+margin))
        for i in self.xCalData:
            xtemp.append(self.get_fieldStrength_dBuVm(i+margin))
        for i in self.yCalData:
            ytemp.append(self.get_fieldStrength_dBuVm(i+margin))
        
        #create numpy arrays for plotting
        zPlusMargin=np.array(ztemp)      
        xPlusMargin=np.array(xtemp)
        yPlusMargin=np.array(ytemp)
        
        #delete temporary arrays
        del ztemp 
        del xtemp
        del ytemp
        
        #create zero array
        zeros=np.zeros_like(a)
        
        #=======================================================================
        # plot results
        #=======================================================================
        
        #plot limit
        self.emcPlot.plot(a,zeros+testVal,lw=1,color='r',ls='--',label=self.emc_regs + ' Class ' + self.emc_class + "  @"+str(self.cal.cal_dist)+"m")
        
        #z-axis data
        if(zdraw):
            self.emcPlot.plot(a,z,lw=1,color=ZCOLOR,label="Z-axis")
            if(margin!=0):
                self.emcPlot.plot(a,zPlusMargin,lw=1,color=ZCOLOR,label="Z-axis + Margin", ls=':')  
            
        #x-axis data     
        if(xdraw):
            self.emcPlot.plot(a,x,lw=1,color=XCOLOR,label="X-axis")
            if(margin!=0):
                self.emcPlot.plot(a,xPlusMargin,lw=1,color=XCOLOR,label="X-axis + Margin" ,ls=':')
            
        #y-axis data
        if(ydraw):
            self.emcPlot.plot(a,y,lw=1,color=YCOLOR,label="Y-axis")
            if(margin!=0):
                self.emcPlot.plot(a,yPlusMargin,lw=1,color=YCOLOR,label="Y-axis + Margin" ,ls=':')
        
        #configure result plot
        self.emcPlot.set_xlabel("Angle (radians)")
        self.emcPlot.set_ylabel("Field Strength (dBuV/m)")
        ymin,ymax=self.emcPlot.get_ylim()        
        self.emcPlot.set_ylim([ymin-10,ymax+10])
        self.figEmc.subplots_adjust(wspace=.1,bottom=.2)
        self.emcPlot.set_title('EMC Compliance Testing\nField Strength vs. Angle',fontsize=14,fontweight=200)
        self.emcPlot.set_xlim(0,2*np.pi)
        self.emcPlot.legend(fontsize=8,loc="best")
        self.figEmc.tight_layout()
        
        #draw plot
        self.emcCanvas.draw()
        self.emc_testComplete=True #set testComplete to true because test is run
        
        #===================================================================
        # Run test
        #===================================================================
        if(zdraw):#skip if no data in array
            for i in self.zCalData:
                print self.get_fieldStrength_dBuVm(i)
                if self.get_fieldStrength_dBuVm(i)+float(self.e_emc_margin.text()) > testVal:
                    print 'EMC Test complete--FAIL'
                    self.b_run_test.setEnabled(True)    #enable run test button after test
                    self.emc_testResults.setText('<span style="  color:red; font-size:16pt; font-weight:400;">-----Test Failed-----</span>')
                    return 'Fail' 
        if(xdraw):#skip if no data in array
            for i in self.xCalData:
                if self.get_fieldStrength_dBuVm(i)+float(self.e_emc_margin.text()) > testVal:
                    print 'EMC Test complete--FAIL'
                    self.b_run_test.setEnabled(True)    #enable run test button after test
                    self.emc_testResults.setText('<span style="  color:red; font-size:16pt; font-weight:400;">-----Test Failed-----</span>')
                    return 'Fail'
        if(ydraw):#skip if no data in array
            for i in self.yCalData:
                if self.get_fieldStrength_dBuVm(i)+float(self.e_emc_margin.text()) > testVal:
                    print 'EMC Test complete--FAIL---'
                    self.b_run_test.setEnabled(True)    #enable run test button after test
                    self.emc_testResults.setText('<span style="  color:red; font-size:16pt; font-weight:400;">-----Test Failed-----</span>')
                    return 'Fail' 
            
        if(hasData==True):   
            print 'EMC Test complete--PASS'
            self.emc_testResults.setText('<span style="  color:lime; font-size:16pt; font-weight:400;">-----Test Passed-----</span>')
            self.b_run_test.setEnabled(True)            #enable run test button after test
            return 'Pass'
        else:
            print 'EMC Test Will Not Run!--INSUFFICIENT DATA'
            self.emc_testResults.setText('<span style="  color:yellow; font-size:16pt; font-weight:400;">-----No Test Data-----</span>')
            self.b_run_test.setEnabled(True)            #enable run test button after test
            self.emc_testComplete=False                 #reset testComplete to false because no data is present
            return 'Fail'
            

        self.b_run_test.setEnabled(True)                #enable run test button after test
  
    def get_farField(self):
        #=======================================================================
        #
        #          Name:    get_farField
        #
        #    Parameters:    None
        #
        #        Return:    (float)farField
        #
        #   Description:    this function returns the far filed of the antenna given the test frequency
        #
        #=======================================================================
        DUTDiameter=float(self.e_emc_diameter.text())
        waveLength = 2.997e8/self.cal.cal_freq
        
        farField=(2*DUTDiameter**2)/waveLength

        print "far-field: ", farField, " m"
        
        #set text in emc test information groupbox
        self.emc_gui_farField.setText(str(farField))
        
        return farField
  
    def get_fieldStrength_dBuVm(self,value):             #Take Recieved power and distance, and convert to Field strength
        'Takes Recieved power and distance, and convert to Field strength'
        #=======================================================================
        #          Name:    get_fieldStrength_dBuVm
        #
        #    Parameters:    (float)value 
        #
        #        Return:    (float)fieldStrength
        #
        #   Description:    this function takes the collected data in dBm and returns the 
        #                   electrical field strength in dBuV/m
        #
        #=======================================================================
        
        EiRP=(10**((float(value)-30)/10))                       #convert dBm to W (EiRP)
        fieldStrength=(math.sqrt(30*EiRP))/self.cal.cal_dist    #calculate field strength from distance and EiRP
        fieldStrength=20*math.log10(fieldStrength/1e-6)         #convert V/m to dBuV/m
        
        return fieldStrength
    
    def get_fieldStrength_uVm(self,value):#Take Recieved power and distance, and convert to Field strength
        'Takes Recieved power and distance, and convert to Field strength'
        #=======================================================================
        #          Name:    get_fieldStrength_uVm
        #
        #    Parameters:    (float)value 
        #
        #        Return:    (float)fieldStrength
        #
        #   Description:    this function takes the collected data in dBm and returns the 
        #                   electrical field strength in uV/m
        #
        #=======================================================================
        
        EiRP=(10**((float(value)-30)/10))                       #convert dBm to W (EiRP)
        fieldStrength=(math.sqrt(30*EiRP))/self.cal.cal_dist    #calculate field strength at distance and EiRP    
        fieldStrength=fieldStrength*1e6                         #convert from V/m to uV/m
        
        return fieldStrength
        
    def set_emcRegulations(self):#sets which set of EMC Regulations should be tested
        #=======================================================================
        #          Name:    set_emcRegulations
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This Function sets the EMC regulations for testign
        #                   Based on user input
        #
        #                   This function also sets the warnings based on the user input and testing 
        #                   parameters
        #
        #=======================================================================
        'sets which set of EMC Regulations should be tested'
        
        dist=[3,10,30]#holds required testing distance
        
        #initialize min and max testing frequencies
        minFreq=30e6
        maxFreq=3e9
        
        if self.r_fcc.isChecked():
        #=======================================================================
        # set fcc testing settings
        #=======================================================================
            self.emc_regs='FCC'
            
            #set min and max testing frequencies
            minFreq=30e6                    #30 MHz
            maxFreq=40e9                    #40 GHz
            
            #set desting distance
            if self.r_classA.isChecked():
                self.emc_class='A'
                dist=[3,10,30]                     #emc testing distance should be 10 meters
            else:
                self.emc_class='B'
                dist=[3,10,10]                     #emc testing distance should be 3 meters
                
        elif self.r_cispr.isChecked():
        #=======================================================================
        # set cispr testing settings
        #=======================================================================
            self.emc_regs='CISPR'
            
            #set min and max testing frequencies
            minFreq=30e6
            maxFreq=1e9
            
            #set desting distance
            if self.r_classA.isChecked():
                self.emc_class='A'
                dist=[10,30,30]                     #emc testing distance should be 30 meters
            else:
                self.emc_class='B'
                dist=[10,10,10]                     #emc testing distance should be 10 meters
        
        #update gain limit in test info groupbox
        self.emc_gui_limit.setText(str(self.get_emcTestLimit(self.cal.cal_freq))) 
        
        #=======================================================================
        # set distance warning label text
        #=======================================================================
        if (self.cal.cal_dist==dist[0] or self.cal.cal_dist==dist[1] or self.cal.cal_dist==dist[2]):
            self.emc_distWarning.setText('<span style="  color:lime; font-size:11pt; font-weight:400;">--Testing Distance ('+ str(self.cal.cal_dist) +' m) Good-- Ready to run test</span>')
        else:
            self.emc_distWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Testing distance set to '+str(self.cal.cal_dist)+' m, ' + self.emc_regs+ ' Class '+ self.emc_class+' Testing Requires '+str(dist[0])+', '+str(dist[1])+' or '+str(dist[2])+' m. </span>')
        
        #=======================================================================
        # set frequency warning label text
        #=======================================================================
        if self.cal.cal_freq<minFreq or self.cal.cal_freq> maxFreq:
            if self.cal.cal_freq<minFreq:
                self.emc_freqWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Testing frequency set to '+str(float(self.cal.cal_freq)/1e6)+' MHz, ' + self.emc_regs+ ' Class '+ self.emc_class+' testing lower frequency limit is '+str(float(minFreq)/1e6)+' MHz. </span>')
            else:
                if(self.emc_regs=='CISPR'):
                    self.emc_freqWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Testing frequency set to '+str(float(self.cal.cal_freq)/1e6)+' MHz, ' + self.emc_regs+ ' Class '+ self.emc_class+' testing upper frequency limit is '+str(float(maxFreq)/1e6)+' MHz. Run test using FCC Regulations.</span>')
                else:
                    self.emc_freqWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Testing frequency set to '+str(float(self.cal.cal_freq)/1e6)+' MHz, ' + self.emc_regs+ ' Class '+ self.emc_class+' testing upper frequency limit is '+str(float(maxFreq)/1e6)+' MHz. </span>')
        else:
            self.emc_freqWarning.setText('<span style="  color:lime; font-size:11pt; font-weight:400;">--Testing Frequency ('+ str(self.cal.cal_freq/1e6) +' MHz) Good-- Ready to run test</span>')
            
        #=======================================================================
        # set far-field warning label text
        #=======================================================================
        ffield=self.get_farField()
        if (self.cal.cal_dist<ffield):
            self.emc_farFieldWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Far-Field for '+ str(self.cal.cal_freq/1e6) +' MHz calulated to '+str(ffield)+' m, '+ ' Testing distance currently set to '+str(self.cal.cal_dist)+' m. </span>')
        else:    
            self.emc_farFieldWarning.setText('<span style="  color:lime; font-size:11pt; font-weight:400;">--Far-Field ('+ str(ffield) +' m) Good-- Ready to run test</span>')
        
    def create_status_bar(self):#create status bar at bottom of aplication
        #=======================================================================
        #          Name:    create_status_bar
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    creates status bar at bottom of application
        #
        #=======================================================================
        self.status_text = QLabel("Click Setup to find instruments.")
        self.statusBar().addWidget(self.status_text, 1)
        
    def create_menu(self):#create menu at top of application 
        #=======================================================================
        #          Name:    create_menu
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    This function creates the menu at the top of the application
        #
        #=======================================================================    
        self.file_menu = self.menuBar().addMenu("&File")
        
        open_file_action = self.create_action("&Open Report",
            shortcut="Ctrl+O", slot=self.open_report, 
            tip="Load a CSV file, first row is Title, first column is deg, subsequent columns mag")
        
        save_csv_action = self.create_action("&Save Report",
            shortcut="Ctrl+S", slot=self.save_report, 
            tip="Save a CSV file, first row is Title, first column is deg, subsequent columns mag")
        
        save_file_action = self.create_action("Save plot",
            slot=self.save_plot, 
            tip="Save the plot")
        quit_action = self.create_action("&Quit", slot=self.close, 
            shortcut="Ctrl+Q", tip="Close the application")
        
        self.add_actions(self.file_menu, 
            (open_file_action,save_csv_action,save_file_action, None, quit_action))
        
        self.help_menu = self.menuBar().addMenu("&Help")
        about_action = self.create_action("&About", 
            shortcut='F1', slot=self.on_about, 
            tip='About')
        
        self.add_actions(self.help_menu, (about_action,))

    def fill_dataArray(self):#fill data arrays so they are all the same size
        #=======================================================================
        #          Name:    fill_dataArray
        #    
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this functions fills the data arrays with zeros which is needed for 
        #                    saving reports and plotting
        #
        #=======================================================================
       

        short=101-(len(self.zRawData))
        if short>0:
            for i in range(0,short):
                self.zRawData.append(0)
                 
        short=101-(len(self.xRawData))
        if short>0:
            for i in range(0,short):
                self.xRawData.append(0)
                 
        short=101-(len(self.yRawData))
        if short>0:
            for i in range(0,short):
                self.yRawData.append(0)
                 
        short=101-(len(self.zCalData))
        if short>0:
            for i in range(0,short):
                self.zCalData.append(0)
                 
        short=101-(len(self.xCalData))
        if short>0:
            for i in range(0,short):
                    self.xCalData.append(0)
                 
        short=101-(len(self.yCalData))
        if short>0:
            for i in range(0,short):
                self.yCalData.append(0)
         
        short=101-(len(self.angles))
        if short>0:
            for i in range(0,short):
                self.angles.append((i*3.6))
        
    def add_actions(self, target, actions):#do something..apparently 
        #=======================================================================
        #          Name:    add_actions
        #
        #    Parameters:    target , actions
        #
        #        Return:    None
        #
        #   Description:    'Unknown'
        #
        #=======================================================================
        for action in actions:
            if action is None:
                target.addSeparator()
            else:
                target.addAction(action)

    def create_action(  self, text, slot=None, shortcut=None, 
                        icon=None, tip=None, checkable=False, 
                        signal="triggered()"):
        #=======================================================================
        #          Name:    create_action
        #
        #    Parameters:    text , slot , shortcut, icon, tip , checkable, signal
        #
        #        Return:    action
        #
        #   Description:    'Unknown'
        #
        #=======================================================================
        action = QAction(text, self)
        if icon is not None:
            action.setIcon(QIcon(":/%s.png" % icon))
        if shortcut is not None:
            action.setShortcut(shortcut)
        if tip is not None:
            action.setToolTip(tip)
            action.setStatusTip(tip)
        if slot is not None:
            self.connect(action, SIGNAL(signal), slot)
        if checkable:
            action.setCheckable(True)
        return action

def main():
    #=======================================================================
    #          Name:    main
    #
    #    Parameters:    None
    #
    #        Return:    None
    #
    #   Description:    this function creates and executes the main application
    #
    #=======================================================================
    app = QApplication(sys.argv)                        #create Qapplication (pyqt)
    app.setStyle(QStyleFactory.create("plastique"))     #change style for readability
    form = AppForm()                                    #create Qmainwindow subclass(pyqt)
    form.resize(800,600)
    form.move(100,10)                                    #move app to upper left corner of display
    #form.showMaximized()
    form.show()                                         #Make application visible (PYQT)
    app.exec_()                                         #enter main loop of QApplication class (pyqt)


if __name__ == "__main__":
    main()
