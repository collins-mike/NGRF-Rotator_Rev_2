"""
RF Rotator
Copyright 2013 Travis Fagerness
v2.0 update by Mike Collins
"""
#from mayavi import mlab
import sys, os, random,time
from PyQt4.QtCore import *
from PyQt4.QtGui import *

import multiprocessing,logging

import datetime

import matplotlib
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt4agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm
from matplotlib import ticker

import openpyxl as pyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

from SignalHound import *
from worker import Worker
from setup import *
from specan import *
from arcus import Arcus
from TestContainer import TestContainer




import numpy
from numpy import minimum
from test.test_binop import isint
# from tvtk.plugins.scene.ui.actions import XMinusView
# from pygments.lexers.jvm import JasminLexer
# from pygments.lexers._scilab_builtins import new_data
numpy.set_printoptions(threshold=numpy.nan)
from Calibrator import Calibrator

import numpy as np
import math
import time
import datetime

from pip._vendor.requests.packages.chardet.latin1prober import FREQ_CAT_NUM

from SignalHound.bb_api_h import BB_TIME_GATE



#===============================================================================
# adjust matplotlib display settings
#===============================================================================
matplotlib.rcParams.update({'font.size': 8})

version = "2.0"
year = "2017"
author = "Travis Fagerness v2.0 update by Mike Collins"
website = "http://www.nextgenrf.com"
email = "mike.collins@nextgenrf.com"

#set colors for plots
XCOLOR="#00BB00"
YCOLOR="#FF0000"
ZCOLOR="#0000FF"

#testing type
PATTERN=0
EMC=1

#antenna polarity constants
POL_H=0
POL_V=1


class AppForm(QMainWindow):#create main application
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
           
        #function and Variable Ddeclaration   
        self.threads=[]         #create empty list for threads
        self.legend=""          #create empy list for legend
        self.rotationAxis='Z'   #set default rotation axis for data collection
        
        
        self.customScaling= False
        #=======================================================================
        # setup data collection variables
        #=======================================================================
                                
        self.TEST_Z=TestContainer(self)  #create tests to hold individual data
        self.TEST_X=TestContainer(self)
        self.TEST_Y=TestContainer(self)
        
        self.currentTest=self.TEST_Z
        
        self.data=np.array([1,2,3]) #holds raw data
#         self.data=[]
        
#         self.zRawData=[]        # holds raw z axis data
#         self.xRawData=[]        # holds raw x axis data
#         self.yRawData=[]        # holds raw y axis data
        
#         self.zCalData=[]        # holds calibrated z axis data
#         self.xCalData=[]        # holds calibrated x axis data
#         self.yCalData=[]        # holds calibrated y axis data
        
        self.angles=[]          # holds angle data
        
        self.color=YCOLOR          # setup plot color for easy diferentiation
        self.pltColor=YCOLOR
        
        self.data_available=False
        self.deviceIsConnected=False

        #create calibrator object
        self.cal=Calibrator()
        self.cal.set_mainForm(self)       #set calibrators pointer to self
        
        #=======================================================================
        # setup EMC testing tab
        #=======================================================================
        self.emc_regs='FCC'                     #select regulation set for testing
        self.emc_class='A'                      #select class of emc testing
        
        self.emc_testComplete=False             #hold wheather an EMC test has been run or not
        self.emc_polarity=POL_V             #set polarity of rx antenna
        self.calArray=[]                        #calArray holds a list of values to be added to the collected data for calibration
        self.calFreqArray=[]                    #calFreq array holds the list of frequencies that the calArray values relate to
        self.EMC_Vertical_Results=[]            #holds a list of vertical polarity errors
        self.EMC_Horizontal_Results=[]          #holds a list of horizontal polarity errors
        self.EMC_VerticalFailureNum=0           #holds the number of failures in the vertical polarity
        self.EMC_HorizontalFailureNum=0         #holds the number of failures in the horizontal polarity
        self.emcTestRunning=False               #will be true whenever the a EMC test is being performed
        self.test_type = PATTERN                #used in the worker to change its behavior based on test type
        self.emc_resuolutionCount=0             #holds the number of data points to be collected in an EMC test
        self.emc_failureLimitReached= False     #will be true when a data point collects more than 2000 errors
        self.emc_testStatus="Fail"              #holds the result status of an EMC test at a data point
        self.emc_reportRow = 2                  #holds row in which to put data in report
        
        self.EMC_DATASET=[]                     #holds faliure data for testing report
        #=======================================================================
        # setup 3D Rendering Variables
        #=======================================================================
        self.render3D=False                                     #is false until application has rendered in 3D 
        
        
        #==================================================
        #setup main window
        #==================================================
        self.setWindowTitle('Rotation RX Strength Plotter')
        self.create_menu()                                      #create menu Qwidget
        
        self.create_tabs()                                      #create tab object to hold application tabs(data collection, calibration, 3d rendering)
        
        self.create_dataCollectionTab(self.tab_dataCollection)  #create data collection tabs
        
        self.create_emcTab(self.tab_emc)
        
        #calibrator object creates it's own tab
        self.cal.create_calibrationTab(self.tab_calibration)
        
        self.create_3dTab()                                     #create 3D rendering tab
        
        
        self.create_status_bar()                                #create status bar at bottom of app
        
        
        
        #==================================================
        #create worker object
        #==================================================
        self.worker=Worker()
        self.worker.set_cal(self.cal)                   #give worker access to calibrator
        self.manual_mode=False
        self.worker.set_mainForm(self)
        #set threading to run worker at same time as this object
        self.threads.append(self.worker)
        
        #worker setup
        self.worker.status_msg.connect(self.status_text.setText)
        self.worker.data_pair.connect(self.on_data_ready)
        self.worker.dev_found.connect(self.device_found)
        self.worker.worker_sleep.connect(self.worker_asleep)
        self.worker.bad_data.connect(self.show_badData)
        self.worker.run_emcTest.connect(self.run_emcTest)
        self.worker.set_cal(self.cal)                   #pass the calibrator to the worker
        self.worker.start()
          
        #=======================================================================
        # create setup dialog box object
        #=======================================================================
        self.setup = Setup(self,self.worker,self.cal)   #create setup object for worker object
        
        self.worker.set_setup(self.setup)               #pass the setup params to the worker
        
        #=======================================================================
        # setup worker and setup access for calibrator
        #=======================================================================
        self.cal.set_setup(self.setup)
        self.cal.set_worker(self.worker)

        
        
        #TODO: fix mpl = multiprocessing.log_to_stderr(logging.CRITICAL)#
        
    def create_dataCollectionTab(self,tab):#create data collection tab as well as main window
        #=======================================================================
        #          Name:    create_dataCollectionTab
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function creates the form and user interface of the data
        #                    collection tab
        #
        #=======================================================================
        
        self.main_frame = QWidget()
        
        #==========================================================================
        #create Label for current axis
        #===========================================================================
        
        self.curAxis=QLabel('<span style=" font-size:14pt; font-weight:600;color:'+YCOLOR+'">Current Rotation Axis: Z</span>')
        self.curAxis.setAlignment(Qt.AlignLeft)
        
        #=======================================================================
        # create figure and canvas for data collection plots
        #=======================================================================
        self.dpi = 100
        self.fig = Figure(dpi=self.dpi)
        self.fig.set_facecolor('#DDDDDD')
        self.canvas = FigureCanvas(self.fig)
#         self.fig.suptitle('DUT: Serial Num: RX Orientation:', fontsize=14, fontweight='bold')
#         self.fig.text(0.95, .1, "Customer: \nTester: \nDate", verticalalignment='bottom', horizontalalignment='right', fontsize=12)
        
        self.canvas.setParent(self.tab_dataCollection)
        
        #=======================================================================
        # create subplots for data plots
        #=======================================================================
        
        self.TEST_Z.setSubplot(self.fig.add_subplot(131,polar=True))
        self.TEST_Z.setTitle('X-Y Plane\n(Rotation on Z-Axis)')
        self.z_axis = self.TEST_Z.getSubplot();
        self.z_axis.set_title('X-Y Plane\n(Rotation on Z-Axis)',y=1.08,fontsize=10,fontweight=300)
        self.z_axis.set_facecolor('white')
        
        self.TEST_X.setSubplot(self.fig.add_subplot(132,polar=True))
        self.TEST_X.setTitle('Y-Z Plane\n(Rotation on X-Axis)')
        self.x_axis = self.TEST_X.getSubplot();
#         self.x_axis = self.fig.add_subplot(132,polar=True)
        self.x_axis.set_title('Y-Z Plane\n(Rotation on X-Axis)',y=1.08,fontsize=10,fontweight=300)
        self.x_axis.set_facecolor('grey')
        
        self.TEST_Y.setSubplot(self.fig.add_subplot(133,polar=True))
        self.TEST_Y.setTitle('X-Z Plane\n(Rotation on Y-Axis)')
        self.y_axis = self.TEST_Y.getSubplot();
#         self.y_axis = self.fig.add_subplot(133,polar=True)
        self.y_axis.set_title('X-Z Plane\n(Rotation on Y-Axis)',y=1.08,fontsize=10,fontweight=300)
        self.y_axis.set_facecolor('grey')
        
        self.axes=self.z_axis#set current axis to axes variable

        #=======================================================================
        # create buttons and GUI controls
        #=======================================================================
        
        # Bind the 'button_press_event' event for clicking on one of the bars
        self.canvas.mpl_connect('button_press_event', self.click_manualTarget)
        
        # Create the navigation toolbar, tied to the canvas
        self.mpl_toolbar = NavigationToolbar(self.canvas, self.tab_dataCollection)
        
        # Other GUI controls
        self.b_setup = QPushButton("&Setup/Find Devices")
        self.connect(self.b_setup, SIGNAL('clicked()'), self.click_setup)
        self.b_setup.setToolTip("Setup tools for test")
        
        
        self.b_manual= QPushButton("&Manual Mode",enabled=False,checkable=True)
        self.b_manual.setEnabled(False)
        self.connect(self.b_manual, SIGNAL('clicked()'), self.click_manual)
        self.b_manual.setToolTip("Move table to specific point while continuously performing test")
        
        self.b_start= QPushButton("&Rotate Start")
        self.b_start.setEnabled(False)
        self.connect(self.b_start, SIGNAL('clicked()'), self.click_start)
        self.b_start.setToolTip("Begin Test")
        
        
        self.b_stop= QPushButton("Stop/&Home",enabled=False)
        self.connect(self.b_stop, SIGNAL('clicked()'), self.click_stop)
        self.b_stop.setToolTip("Abort test and return to home position")
        
        self.b_pause= QPushButton("&Pause",enabled=False,checkable=True)
        self.connect(self.b_pause, SIGNAL('clicked()'), self.click_pause)
        self.b_pause.setToolTip("Pause current test")
        
        self.b_reset= QPushButton("&Clear",enabled=True)
        self.connect(self.b_reset, SIGNAL('clicked()'), self.click_clear)
        self.b_reset.setToolTip("Clear data plot from active rotation axis")

        self.grid_cb = QCheckBox("Show &Grid",checked=True)
        self.connect(self.grid_cb, SIGNAL('stateChanged(int)'), self.update_plot_settings)
        self.b_reset.setToolTip("Show grid on active axis?")
        
        #====================================================================================
        #Create rotation axis selection controls
        #=================================================================================
    
        axisVbox=QVBoxLayout()
        axisVbox.addWidget(QLabel("Select Axis"))
        axisHbox=QHBoxLayout()
        
        self.rb_axisSelZ=QRadioButton('Z')      #create axis select radio buttons
        self.rb_axisSelZ.click()                #set Z axis to default axis select radio button
        self.rb_axisSelX=QRadioButton('X')      #create axis select radio buttons
        self.rb_axisSelY=QRadioButton('Y')      #create axis select radio buttons
        axisHbox.addWidget(self.rb_axisSelZ)
        axisHbox.addWidget(self.rb_axisSelX)
        axisHbox.addWidget(self.rb_axisSelY)
        axisVbox.addLayout(axisHbox)
        
        #connect buttons
        self.connect(self.rb_axisSelZ, SIGNAL('clicked()'), self.click_axisSelect)
        self.connect(self.rb_axisSelX, SIGNAL('clicked()'), self.click_axisSelect)
        self.connect(self.rb_axisSelY, SIGNAL('clicked()'), self.click_axisSelect)
        
        #set radio button tool tips
        self.rb_axisSelZ.setToolTip("Switch to Z axis")
        self.rb_axisSelX.setToolTip("Switch to Y axis")
        self.rb_axisSelY.setToolTip("Switch to X axis")
        
        progess_label = QLabel("Rotation Progress:")
        self.progress = QProgressBar()
        self.progress.setAlignment = Qt.Horizontal
        self.progress.setMaximum(360)
        self.progress.setMinimum(0)
        
        
        #===============================================================================
        # Layout with box sizers
        # ==============================================================================
        hbox = QHBoxLayout()
        
        #=======================================================================
        # create button bar
        #=======================================================================
        hbox.addLayout(axisVbox)
        for w in [  self.b_setup,self.b_manual, self.b_start,self.b_stop,self.b_pause,self.b_reset, self.grid_cb,
                    progess_label, self.progress]:
            hbox.addWidget(w)
            hbox.setAlignment(w, Qt.AlignVCenter)
            
        
        vbox = QVBoxLayout()                #create layout    
        
        ###################################
#         self.b_testbutton= QPushButton("&TEST",enabled=True)
#         self.b_testbutton.clicked.connect(lambda: self.show_badData(3))
#         
#         vbox.addWidget(self.b_testbutton)
        ######################################################################
        tophbox=QHBoxLayout()
        tophbox.addWidget(self.curAxis)
        
        #=======================================================================
        # create scaling controls
        #=======================================================================
        scaleForm=QFormLayout()
        self.e_scaleMax=QLineEdit()
        self.e_scaleMin=QLineEdit()
        self.btn_applyScaling=QPushButton("Apply")
        scaleForm.addRow(QLabel("scale Max"), self.e_scaleMax)
        scaleForm.addRow(QLabel("scale Min"), self.e_scaleMin)
        scaleForm.addRow("", self.btn_applyScaling)
        self.connect(self.btn_applyScaling, SIGNAL('clicked()'), self.click_applyScaling)
        tophbox.addStretch()
        tophbox.addLayout(scaleForm)
        tophbox.addStretch()
        
        #=======================================================================
        # place all vertical components of data collection tab
        #=======================================================================
        
        vbox.addLayout(tophbox)       #add current rotation axis display label    
        vbox.addWidget(self.canvas,10)      #add graph area to display
        vbox.addWidget(self.mpl_toolbar)    #add matplotlib toolbar to display
        vbox.addLayout(hbox)                #add control buttons to display
        
        #add layout to tab
        self.tab_dataCollection.setLayout(vbox)
        
    def worker_asleep(self):#worker wating for command
        #=======================================================================
        #          Name:    worker_asleep
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function enables/disables the appropriate
        #                    pushbuttons when the worker object is awake/asleep
        #
        #=======================================================================
        'enable/disable GUI buttons for when worker is asleep'
        
        #if the worker is asleep (not paused) the rotation table should be at home
        if self.deviceIsConnected:
            self.b_start.setEnabled(not self.manual_mode)
            self.cal.testConfigEnable(True)
            self.b_manual.setEnabled(True)
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_setup.setEnabled(True)
            self.cal.b_specan.setEnabled(True)
            self.rb_axisSelZ.setEnabled(True)
            self.rb_axisSelX.setEnabled(True)
            self.rb_axisSelY.setEnabled(True)
            self.b_run_test.setEnabled(True)
            self.b_emcPause.setEnabled(False)
            self.b_abort.setEnabled(False)
            
            
            #display specan type in calibration tab
            self.cal.gui_specan.setText(self.worker.specan.device)
            
        else:
            self.cal.b_specan.setEnabled(False)
            self.b_start.setEnabled(False)
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_manual.setEnabled(False)
#             self.rb_axisSelZ.setEnabled(False)
#             self.rb_axisSelX.setEnabled(False)
#             self.rb_axisSelY.setEnabled(False)
            self.b_run_test.setEnabled(False)
            self.b_emcPause.setEnabled(False)
            self.b_abort.setEnabled(False)
            
            #display specan type in calibration tab
            self.cal.gui_specan.setText("--Spectrum analyzer not detected--")
    def show_badData(self,difference):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setInformativeText("The first and last values in the collected data array have a difference of "+ str(difference) +" dB. The recommended maximum difference is 2 dB or less.\n\nClick \"Retry\" to run test again.\n\nClick \"Average\" to set the first and last data point values to the average of those two values")
        
        msg.setWindowTitle("Bad Data Warning")
        msg.setText("The beginning and end values of the collected data do not match. These values represent the same data point, and should hold the same value.")
        

        msg.setStandardButtons(QMessageBox.Retry | QMessageBox.Ignore)
        msg.addButton("Average",QMessageBox.ApplyRole)
        msg.buttonClicked.connect(self.msgbtn)
        
        retval = msg.exec_()
        print "value of pressed message box button:", retval
        
        self.rb_axisSelZ.setEnabled(True)
        self.rb_axisSelX.setEnabled(True)
        self.rb_axisSelY.setEnabled(True)
        
        self.b_setup.setEnabled(True)
    
        self.cal.testConfigEnable(True)
        
    def show_errorDialog(self,title,text,info):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(text)
        msg.setInformativeText(info)
        msg.setWindowTitle(title)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
        
    def msgbtn(self,i):
        print "Button pressed is:",i.text()
        if i.text()== "Retry":
            self.click_start()
            
        elif i.text()== "Ignore":
            pass
        
        elif i.text()== "Average":
            if (self.rotationAxis=='Z'):
                test=self.TEST_Z
                
            elif(self.rotationAxis=='X'):
                test=self.TEST_X
                
            elif(self.rotationAxis=='Y'):
                test=self.TEST_Y
                
            test.dataArrayRaw[0] = (test.dataArrayRaw[0]+test.dataArrayRaw[len(test.dataArrayRaw)-1])/2
            test.dataArrayRaw[len(test.dataArrayRaw)-1] = test.dataArrayRaw[0]
            
            test.dataArrayCal[0] = (test.dataArrayCal[0]+test.dataArrayCal[len(test.dataArrayCal)-1])/2
            test.dataArrayCal[len(test.dataArrayCal)-1] = test.dataArrayCal[0]
            
            self.data=np.array(test.dataArrayCal)   
            self.draw_dataPlots()  
            
          
    def device_found(self,devices=[False,'Not Found','Not Found']):
        #=======================================================================
        #          Name:    device_found
        #
        #    Parameters:    devices
        #
        #        Return:    None
        #
        #   Description:    this function sets if the spectrum analyzer AND turntable are found
        #
        #=======================================================================
        'set if specan AND turntable are found'
        self.deviceIsConnected=devices[0]
    
    def save_report(self):#create .xlsx report file
        #=======================================================================
        #          Name:    save_report
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function saves the collected data and the test info
        #                    is .xlsx format
        #        
        #                    this functions uses the openpyxl library Heavily
        #
        #=======================================================================
        'save current plot data as .csv'

        file_choices = "Excel Workbook ( *.xlsx)"
        path = unicode(QFileDialog.getSaveFileName(self, 
                        'Save', '', 
                        file_choices))
        
        #========================================================================
        # make data arrays the same size for export to csv
        #========================================================================
        #self.fill_dataArray()
       
            #===================================================================
            # save Report to .xlsx
            #===================================================================
        if path:
            #===================================================================
            # create styles for automatic reporting
            #===================================================================
            
            #data style
            style_data = NamedStyle(name="style_data")
            thinbd = Side(style='thin', color="000000")
            thickbd = Side(style='medium', color="000000")
            style_data.border = Border(left=thinbd, right=thinbd)
            style_data.alignment=Alignment(horizontal="right")
            
            #pass style
            style_pass = NamedStyle(name="style_pass")
            style_pass.font = Font(bold=True, size=14,color="009B00")
            style_pass.border = Border(left=thinbd, right=thinbd)
            style_pass.alignment=Alignment(horizontal="right")
            
            #fail style
            style_fail = NamedStyle(name="style_fail")
            style_fail.font = Font(bold=True, size=14,color="FF0000")
            style_fail.border = Border(left=thinbd, right=thinbd)
            style_fail.alignment=Alignment(horizontal="right")
            
            #top header style
            style_headerTop = NamedStyle(name="style_headerTop")
            style_headerTop.font = Font(bold=False, size=12)
            style_headerTop.border = Border(left=thinbd, top=thinbd, right=thinbd)
            style_headerTop.alignment=Alignment(horizontal="center")
            style_headerTop.fill=PatternFill("solid", fgColor="DDDDDD")
            
            #left header style
            style_headerLeft = NamedStyle(name="style_headerLeft")
            style_headerLeft.font = Font(bold=False, size=12)
            style_headerLeft.border = Border(left=thinbd,right=thinbd)
            style_headerLeft.alignment=Alignment(horizontal="left")
            style_headerLeft.fill=PatternFill("solid", fgColor="DDDDDD")
            
            #Title style
            style_title = NamedStyle(name="style_title")
            style_title.font = Font(bold=True, size=14,color="FFFFFF")
            style_title.alignment=Alignment(horizontal="center",vertical="center")
            style_title.fill=PatternFill("solid", fgColor="555555")
            
            #===================================================================
            # initialize workbook to save as .xlsx
            #===================================================================
            
            wb = pyxl.Workbook()

#===============================================================================
# DATA SHEET
#===============================================================================
            #create new worksheet in first position
            ws = wb.create_sheet("Overview & Data", 0) # 
            
            #setup variable locations for data for easy formatting during design
            DATA_HEIGHT =10
            SETUP_HEIGHT = 10
            #===================================================================
            # Create informations cells
            #===================================================================
            
            ws.merge_cells('A1:D1')
            ws.row_dimensions[1].height = 50
            ws['A1']= 'Radiation Pattern Testing'
            ws['A1'].style=style_title
            ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
            
            # add NGRFlogo
            img = Image('images/ngrf.png')
            ws.add_image(img, 'F1')
            
            #create date cells
            ws['A2']= 'Date:'
            ws['A2'].style=style_headerLeft
            ws["B2"]= datetime.date.today()
            ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
            
            #Create Customer Cells
            ws['A3']= 'Customer:'
            ws['A3'].style=style_headerLeft
            ws["B3"]= self.cal.cal_customer
            
            #Create tester name Cells
            ws['A4']= 'Tested By:'
            ws['A4'].style=style_headerLeft
            ws["B4"]= self.cal.cal_tester
            
            #Create comment Cells
            ws['A5']= 'Comments:'
            ws['A5'].style=style_headerLeft
            ws["B5"]= self.cal.cal_comments
            
            #Create DUT label Cells
            ws['C2']= 'DUT Label:'
            ws['C2'].style=style_headerLeft
            ws["D2"]= self.cal.cal_dutLabel
            
            #Create DUT serial number Cells
            ws['C3']= 'DUT Serial Number:'
            ws['C3'].style=style_headerLeft
            ws["D3"]= self.cal.cal_dutSN
            
            #===================================================================
            # Write data and angles to xlsx file
            #===================================================================
            
            ws.column_dimensions['A'].width = 20
            ws['A'+str(DATA_HEIGHT)]= "Angle (degrees)"
            ws['A'+str(DATA_HEIGHT)].style=style_headerTop
            i=DATA_HEIGHT+1
            for angle in self.angles:
                ws['A'+str(i)] = angle
                ws['A'+str(i)].number_format = '0.00E+00'
                ws['A'+str(i)].style=style_data
                i=i+1
            
            ws.column_dimensions['B'].width = 20
            ws['B7']= self.TEST_Z.getTitle()
            ws['B7'].style=style_headerTop
            i=11
            for zraw in self.TEST_Z.dataArrayRaw:
                ws['B'+str(i)] = zraw
                ws['B'+str(i)].style=style_data
                i=i+1  
            
            ws.column_dimensions['C'].width = 20    
            ws['C7']= self.TEST_Z.getTitle()
            ws['C7'].style=style_headerTop
            i=11
            for zcal in self.TEST_Z.dataArrayCal:
                ws['C'+str(i)] = zcal
                ws['C'+str(i)].style=style_data
                i=i+1 
            
            ws.column_dimensions['D'].width = 20
            ws['D7']= self.TEST_X.getTitle()
            ws['D7'].style=style_headerTop
            i=11
            for xraw in self.TEST_X.dataArrayRaw:
                ws['D'+str(i)] = xraw
                ws['D'+str(i)].style=style_data
                i=i+1
            
            ws.column_dimensions['E'].width = 20    
            ws['E7']= self.TEST_X.getTitle()
            ws['E7'].style=style_headerTop
            i=11
            for xcal in self.TEST_X.dataArrayCal:
                ws['E'+str(i)] = xcal
                ws['E'+str(i)].style=style_data
                i=i+1     
                
            ws.column_dimensions['F'].width = 20
            ws['F7']= self.TEST_Y.getTitle()
            ws['F7'].style=style_headerTop
            i=11
            for yraw in self.TEST_Y.dataArrayRaw:
                ws['F'+str(i)] = yraw
                ws['F'+str(i)].style=style_data
                i=i+1 
            
            ws.column_dimensions['G'].width = 20
            ws['G7']= self.TEST_Y.getTitle()
            ws['G7'].style=style_headerTop
            i=11
            for ycal in self.TEST_Y.dataArrayCal:
                ws['G'+str(i)] = ycal
                ws['G'+str(i)].style=style_data
                i=i+1 
            
            #===================================================================
            # Write Max/averages and headers
            #===================================================================
            ws.merge_cells('A6:G6')
            ws['A6']="Test Data"
            ws['A6'].style=style_title
            
            #extra cell to look pretty
            ws['A7']=""
            ws['A7'].style=style_title
            
            #create max value cells
            ws['A8']= 'Max Gain (dBi):'
            ws['A8'].style=style_headerLeft

            #create average value cells
            ws['A9']= 'Average Gain (dBi):'
            ws['A9'].style=style_headerLeft
            
            ws['B8'] = "=MAX(B11:B111)"
            ws['B8'].style=style_data
            ws['B9'] = "=AVERAGE(B11:B111)"
            ws['B9'].style=style_data
            #insert blank cells
            ws['B10']='(Z) RAW'
            ws['B10'].style=style_headerTop
            
            ws['C8'] = "=MAX(C11:C111)"
            ws['C8'].style=style_data
            ws['C9'] = "=AVERAGE(C11:C111)"
            ws['C9'].style=style_data
            #insert blank cells
            ws['C10']='(Z) CALIBRATED'
            ws['C10'].style=style_headerTop
            
            ws['D8'] = "=MAX(D11:D111)"
            ws['D8'].style=style_data
            ws['D9'] = "=AVERAGE(D11:D111)"
            ws['D9'].style=style_data
            #insert blank cells
            ws['D10']='(X) RAW'
            ws['D10'].style=style_headerTop
            
            ws['E8'] = "=MAX(E11:E111)"
            ws['E8'].style=style_data
            ws['E9'] = "=AVERAGE(E11:E111)"
            ws['E9'].style=style_data
            #insert blank cells
            ws['E10']='(X) CALIBRATED'
            ws['E10'].style=style_headerTop
            
            ws['F8'] = "=MAX(F11:F111)"
            ws['F8'].style=style_data
            ws['F9'] = "=AVERAGE(F11:F111)"
            ws['F9'].style=style_data
            #insert blank cells
            ws['F10']='(Y) RAW'
            ws['F10'].style=style_headerTop
            
            ws['G8'] = "=MAX(G11:G111)"
            ws['G8'].style=style_data
            ws['G9'] = "=AVERAGE(G11:G111)"
            ws['G9'].style=style_data
            #insert blank cells
            ws['G10']='(Y) CALIBRATED'
            ws['G10'].style=style_headerTop
            
            #===================================================================
            # Create setup / Calibration data area
            #===================================================================
            
            #title
            ws.merge_cells('I6:J6')
            ws['I6']="Test Setup"
            ws['I6'].style=style_title
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 20
            
            #Frequency
            ws['I7']="Test Frequency (Hz):"
            ws['I7'].style=style_headerLeft
            ws['J7']=float(self.cal.cal_freq)
            ws['J7'].style=style_data
            
            #Frequency span
            ws['I8']="Frequency Span (Hz):"
            ws['I8'].style=style_headerLeft
            ws['J8']=float(self.cal.cal_span)
            ws['J8'].style=style_data
            
            #distance
            ws['I9']="Testing Distance (m):"
            ws['I9'].style=style_headerLeft
            ws['J9']=float(self.cal.cal_dist)
            ws['J9'].style=style_data
            
            #distance
            ws['I10']="Sweep Time (s):"
            ws['I10'].style=style_headerLeft
            ws['J10']=float(self.cal.cal_sc_sweepTime)
            ws['J10'].style=style_data
            
            #calibration title
            ws.merge_cells('I11:J11')
            ws['I11']="Calibration"
            ws['I11'].style=style_title
            
            #Input power
            ws['I12']="Input Power (dBm):"
            ws['I12'].style=style_headerLeft
            ws['J12']=float(self.cal.cal_inputPwr)
            ws['J12'].style=style_data
            
            #amplifer gain
            ws['I13']="PreAmplifier Gain (dB):"
            ws['I13'].style=style_headerLeft
            ws['J13']=float(self.cal.cal_ampGain)
            ws['J13'].style=style_data
            
            #tx cable loss
            ws['I14']="Tx Cable Loss (dB):"
            ws['I14'].style=style_headerLeft
            ws['J14']=float(self.cal.cal_txCableLoss)
            ws['J14'].style=style_data
            
            #DUT Gain
            ws['I15']="DUT Gain (dBi):"
            ws['I15'].style=style_headerLeft
            ws['J15']=float(self.cal.cal_txGain)
            ws['J15'].style=style_data
            
            #fspl
            ws['I16']="FSPL (dB):"
            ws['I16'].style=style_headerLeft
            ws['J16']=float(self.cal.cal_fspl)
            ws['J16'].style=style_data
            
            #rx antenna gain
            ws['I17']="Calibrated Antenna Gain (dBi):"
            ws['I17'].style=style_headerLeft
            ws['J17']=float(self.cal.cal_rxGain)
            ws['J17'].style=style_data
            
            #Rx cable Loss
            ws['I18']="Rx Cable Loss (dB):"
            ws['I18'].style=style_headerLeft
            ws['J18']=float(self.cal.cal_rxCableLoss)
            ws['J18'].style=style_data
            
            #additional elements
            it=0;
            for i in self.cal.addGainLoss:
                ws['I'+str(19+it)]=str(i)+"(dB):"
                ws['I'+str(19+it)].style=style_headerLeft
                ws['J'+str(19+it)]=float(self.cal.addGainLoss[i])
                ws['J'+str(19+it)].style=style_data
                it=it+1
                
            #Total Cal value
            ws['I'+str(19+it)]="Total (dB):"
            ws['I'+str(19+it)].style=style_headerLeft
            ws['J'+str(19+it)]=float(-self.cal.calibrate_data(0))
            ws['J'+str(19+it)].style=style_data
            
            
#===============================================================================
# data collection Plot Sheet
#===============================================================================

            #create new worksheet in 2nd position
            ws = wb.create_sheet("Data Plots", 1) # 
            
            ws.merge_cells('A1:D1')
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.row_dimensions[1].height = 50
            ws['A1']= 'Data Collection Plots'
            ws['A1'].style=style_title
            ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
            
            #create date cells
            ws['A2']= 'Date:'
            ws['A2'].style=style_headerLeft
            ws["B2"]= datetime.date.today()
            ws["B2"].alignment=Alignment(horizontal="right")#correct alignment
            
            #Create DUT label Cells
            ws['A3']= 'DUT Label:'
            ws['A3'].style=style_headerLeft
            ws["B3"]= self.cal.cal_dutLabel
            ws['B3'].style=style_data
            
            #Create DUT serial number Cells
            ws['A4']= 'DUT Serial Number:'
            ws['A4'].style=style_headerLeft
            ws["B4"]= self.cal.cal_dutSN
            ws['B4'].style=style_data
            
            # add NGRFlogo
            img = Image('images/ngrf.png')
            ws.add_image(img, 'F1')
            
            self.x_axis.set_facecolor('white')
            self.y_axis.set_facecolor('white')
            self.z_axis.set_facecolor('white')
            
            self.canvas.print_figure("temp_fig1.png", dpi=100,facecolor=self.figEmc.get_facecolor())
            img = Image('temp_fig1.png')
            ws.add_image(img, 'A5')
            
#===============================================================================
# EMC Compliance Plot Sheet
#===============================================================================
#create new worksheet in 3rd position if EMC test has been run
            if(self.emc_testComplete):
                ws = wb.create_sheet("EMC Pre-Compliance", 2) 
                ws.merge_cells('A1:E1')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 20
                ws.row_dimensions[1].height = 50
                ws.row_dimensions[12].height = 50
                
                ws['A1']= 'EMC Pre-Compliance Test Results'
                ws['A1'].style=style_title
                ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
                
                # add NGRFlogo
                img = Image('images/ngrf.png')
                ws.add_image(img, 'G1')
                
                #===============================================================
                # add test information
                #===============================================================
                #create date cells
                ws['A2']= 'Date:'
                ws['A2'].style=style_headerLeft
                ws["B2"]= datetime.date.today()
                ws["B2"].alignment=Alignment(horizontal="right")#correct alignment
                
                #Create DUT label Cells
                ws['A3']= 'DUT Label:'
                ws['A3'].style=style_headerLeft
                ws["B3"]= self.cal.cal_dutLabel
                ws['B3'].style=style_data
                
                #Create DUT serial number Cells
                ws['A4']= 'DUT Serial Number:'
                ws['A4'].style=style_headerLeft
                ws["B4"]= self.cal.cal_dutSN
                ws['B4'].style=style_data
                
                #regs
                ws['A5']="Regulatory Convention:"
                ws['A5'].style=style_headerLeft
                ws['B5']=self.emc_regs
                ws['B5'].style=style_data
                
                #class
                ws['A6']="Class:"
                ws['A6'].style=style_headerLeft
                ws['B6']=self.emc_class
                ws['B6'].style=style_data
                
                #Test Distance
                ws['A7']="Testing Distance (m):"
                ws['A7'].style=style_headerLeft
                ws['B7']=self.cal.cal_dist;
                ws['B7'].style=style_data
                
                #error margin
                ws['A8']="Error Margin (dBuV/m):"
                ws['A8'].style=style_headerLeft
                ws['B8']=float(self.e_emc_margin.text())
                ws['B8'].style=style_data
                
                #detector type
                ws['A9']="Detector Type:"
                ws['A9'].style=style_headerLeft
                if(self.emc_regs=="FCC"):
                    ws['B9']="+Peak if <1GHz, Average if >1GHz"
                else:
                    ws['B9']="+Peak"
                ws['B9'].style=style_data


                #Test Status
                ws['C2']="Test Status:"
                ws['C2'].style=style_headerLeft
                ws['D2']=self.emc_testStatus;
                if self.emc_testStatus=="Pass":
                    ws['D2'].style=style_pass
                else:
                    ws['D2'].style=style_fail
                
                
                #Create vertical failure number Cells
                ws['C3']= 'Vertical Failure Number'
                ws['C3'].style=style_headerLeft
                ws["D3"]= self.EMC_VerticalFailureNum
                ws['D3'].style=style_data
                
                #Create horizontal failure number Cells
                ws['C4']= 'Horizontal Failure Number'
                ws['C4'].style=style_headerLeft
                ws["D4"]= self.EMC_HorizontalFailureNum
                ws['D4'].style=style_data
                
                
                #Create Total Failures
                ws['C5']= 'Total Failures'
                ws['C5'].style=style_headerLeft
                ws["D5"]= self.EMC_HorizontalFailureNum + self.EMC_VerticalFailureNum
                ws['D5'].style=style_data
                
                #Calibrated antenna
                ws['C6']= 'Calibrated Antenna'
                ws['C6'].style=style_headerLeft
                ws["D6"]= str(self.cal.dia_rx.cb_antennaSel.currentText())
                ws['D6'].style=style_data
                
                #Calibrated cable
                ws['C7']= 'Calibrated Cable'
                ws['C7'].style=style_headerLeft
                ws["D7"]= str(self.cal.dia_rxCable.cb_cableSel.currentText())
                ws['D7'].style=style_data
                
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 10
                ws.column_dimensions['H'].width = 20
                
                
                
                ws['A11']='Failure Points'
                ws['A11'].style=style_title
                
                row=12#start data at row 10
                
                ws['A'+str(row)]='Angle (Degrees)'
                ws['A'+str(row)].style=style_headerTop
                ws['A'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['B'+str(row)]='Antenna Polarity'
                ws['B'+str(row)].style=style_headerTop
                ws['B'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['C'+str(row)]='Frequency (Hz)'
                ws['C'+str(row)].style=style_headerTop
                ws['C'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['D'+str(row)]='Raw Data (dBm)'
                ws['D'+str(row)].style=style_headerTop
                ws['D'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['E'+str(row)]='Calibration (dB)'
                ws['E'+str(row)].style=style_headerTop
                ws['E'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['F'+str(row)]='Field strength (dBuV/m)'
                ws['F'+str(row)].style=style_headerTop
                ws['F'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['G'+str(row)]='Field strength Limit (dBuV/m)'
                ws['G'+str(row)].style=style_headerTop
                ws['G'+str(row)].alignment=Alignment(wrap_text=True)
                
                ws['H'+str(row)]='Failure Margin (dBuV/m)'
                ws['H'+str(row)].style=style_headerTop
                ws['H'+str(row)].alignment=Alignment(wrap_text=True)
                
                
                row+=1
                vtop3=[]
                vtop=0
                vtopVal=0
                if (self.EMC_VerticalFailureNum>0 ):
                    for i in range(3):
                        vtop=0
                        vtopVal=-1000
                        for x in range(self.EMC_VerticalFailureNum):
                            testval=self.EMC_Vertical_Results[x][4]-self.EMC_Vertical_Results[x][5]
                            inList=False
                            for j in vtop3:
                                if int(x)==int(j):
                                    inList=True
                        
                            if ((testval>vtopVal) and (inList==False)):
                                vtopVal=testval
                                vtop=x
                                
                        vtop3.append(vtop)       
                        

                    for i in vtop3:
                        ws['A'+str(row)]=self.EMC_Vertical_Results[i][0]
                        ws['B'+str(row)]='Vertical'
                        ws['C'+str(row)]=self.EMC_Vertical_Results[i][1]
                        ws['D'+str(row)]=self.EMC_Vertical_Results[i][2]
                        ws['E'+str(row)]=self.EMC_Vertical_Results[i][3]
                        ws['F'+str(row)]=self.EMC_Vertical_Results[i][4]
                        ws['G'+str(row)]=self.EMC_Vertical_Results[i][5]
                        ws['H'+str(row)]=("=(F"+str(row)+"-G"+str(row)+")")
                        row+=1 
                        
                htop3=[]
                htop=0 
                htopVal=0       
                if ( self.EMC_HorizontalFailureNum>0):
                    for i in range(3):
                        htop=0 
                        htopVal=-1000
                        for x in range(self.EMC_HorizontalFailureNum):
                            testval=self.EMC_Horizontal_Results[x][4]-self.EMC_Horizontal_Results[x][5]
                            
                            inList=False
                            for j in htop3:
                                if int(x)==int(j):
                                    inList=True
                            if ((testval>htopVal) and (inList==False)):
                                htopVal=testval
                                htop=x
                                
                        htop3.append(htop) 
                    
                    for i in htop3:
                        ws['A'+str(row)]=self.EMC_Horizontal_Results[i][0]
                        ws['B'+str(row)]='Horizontal'
                        ws['C'+str(row)]=self.EMC_Horizontal_Results[i][1]
                        ws['D'+str(row)]=self.EMC_Horizontal_Results[i][2]
                        ws['E'+str(row)]=self.EMC_Horizontal_Results[i][3]
                        ws['F'+str(row)]=self.EMC_Horizontal_Results[i][4]
                        ws['G'+str(row)]=self.EMC_Horizontal_Results[i][5]
                        ws['H'+str(row)]=("=(F"+str(row)+"-G"+str(row)+")")
                        row+=1  
                
                
                
                imgV = Image('temp_emcVFig.png')
                imgH = Image('temp_emcHFig.png')
                
                ws.add_image(imgV, 'A20')
                ws.add_image(imgH, 'A51')
            
#===============================================================================
# 3D radiation Pattern Sheett
#===============================================================================
#create new worksheet in 3rd position if EMC test has been run
            if(self.render3D):
                
                #if emc test is run put in 4th position else 3rd
                if(self.emc_testComplete):
                    ws = wb.create_sheet("3D Radiation Pattern", 3) 
                else:
                    ws = wb.create_sheet("3D Radiation Pattern", 2) 
                    
                ws.merge_cells('A1:E1')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.row_dimensions[1].height = 50
                
                ws['A1']= '3D Radiation Pattern'
                ws['A1'].style=style_title
                ws['A1'].font=Font(bold=False, size=40, color="FFFFFF")
                
                #create date cells
                ws['A2']= 'Date:'
                ws['A2'].style=style_headerLeft
                ws["B2"]= datetime.date.today()
                ws["B2"].alignment=Alignment(horizontal="left")#correct alignment
                
                #Create DUT label Cells
                ws['A3']= 'DUT Label:'
                ws['A3'].style=style_headerLeft
                ws["B3"]= self.cal.cal_dutLabel
                ws['B3'].style=style_data
                
                #Create DUT serial number Cells
                ws['A4']= 'DUT Serial Number:'
                ws['A4'].style=style_headerLeft
                ws["B4"]= self.cal.cal_dutSN
                ws['B4'].style=style_data
                
                # add NGRFlogo
                img = Image('images/ngrf.png')
                ws.add_image(img, 'G1')
                
                #add 3D radiation pattern plot
                self.canvas3d.print_figure("temp_fig3.png", dpi=100, facecolor=self.figEmc.get_facecolor())
                img = Image('temp_fig3.png')
                ws.add_image(img, 'A5')
            #===================================================================
            # save .xlsx file
            #===================================================================
            try:
                wb.save(path)
            except: 
                print "Error"
                self.show_errorDialog("File Save Error!", "Unable to save report!", "Ensure report is not open in another program.")
                  
    def click_open(self):
        if(self.TEST_Z.getHoldsData() or self.TEST_X.getHoldsData() or self.TEST_Y.getHoldsData()):
            
            open_msg = "Opening a report will cause any unsaved data to be lost.\n\nDo you want to continue?"
            reply = QMessageBox.warning(self, 'Data Overwrite Warning', open_msg, QMessageBox.Yes, QMessageBox.No)
        
            if reply == QMessageBox.Yes:
                self.open_report()
        else:
            self.open_report()
        
    def open_report(self):#open previous test fro .xlsx file
        #=======================================================================
        #          Name:    open_report
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    opens a previous test from .xlsx file
        #
        #=======================================================================
        'open .xlsx file of previous test'
        
        
        #=======================================================================
        # open file
        #=======================================================================
        file_choices = "Excel Workbook ( *.xlsx)"
        
        #open file 
        path = unicode(QFileDialog.getOpenFileName(self, 'Open', '', file_choices))
        if path:
            wb2 = load_workbook(path)   #load openpyxl workbook
            print wb2.get_sheet_names() #show sheet names
            names=wb2.get_sheet_names() #get name of data sheet
            ws=wb2[str(names[0])];      #set active worksheet to data sheet
            
            
            #===================================================================
            # get test info from file
            #===================================================================
            
            self.cal.cal_customer=str(ws['B3'].value)
            self.cal.e_cal_customer.setText(self.cal.cal_customer)
            
            self.cal.cal_tester=str(ws['B4'].value)
            self.cal.e_cal_tester.setText(self.cal.cal_tester)
            
            self.cal.cal_dutLabel=str(ws['D2'].value)
            self.cal.e_cal_dutLabel.setText(self.cal.cal_dutLabel)
            
            self.cal.cal_dutSN=str(ws['D3'].value)
            self.cal.e_cal_dutSN.setText(self.cal.cal_dutSN)
            
            self.cal.apply_testInfo()  
            
            #===================================================================
            # get test setup from file
            #===================================================================
            
            self.cal.cal_dist=int(ws['J9'].value)
            self.cal.e_cal_dist.setText(str(self.cal.cal_dist))
            
            self.cal.cal_freq=int(ws['J7'].value)
            self.cal.e_cal_freq.setText(str(self.cal.cal_freq/1e6))
            
            self.cal.cal_span=int(ws['J8'].value)
            self.cal.e_cal_span.setText(str(self.cal.cal_span/1e6))
            
            self.cal.cal_sc_sweepTime=int(ws['J10'].value)
            self.cal.e_cal_sc_sweepTime.setText(str(self.cal.cal_sc_sweepTime*1000))
            
            self.cal.apply_testConfig()
            #===================================================================
            # load data from .xlsx file
            #===================================================================
            
            #get number of angle collected
            startingVal=11
            cnt=0;
            while(ws['A'+str(cnt+startingVal)].value != None):
                cnt+=1;
            
            self.TEST_Z.clearAllData()
            self.TEST_X.clearAllData()
            self.TEST_Y.clearAllData()
            
            del self.angles          # holds angle data
             
            self.angles=[]          # holds angle data
            
            #default drawing to false so empty arrays will not be drawn
            drawz=False;
            drawx=False;
            drawy=False;
            
            
            
            for i in range(0,cnt):
                if(ws['A'+str(i+startingVal)].value!=None):
                    self.angles.append(float(ws['A'+str(i+startingVal)].value))
                    self.TEST_X.angleArray.append(float(ws['A'+str(i+startingVal)].value))
                    self.TEST_Y.angleArray.append(float(ws['A'+str(i+startingVal)].value))
                    self.TEST_Z.angleArray.append(float(ws['A'+str(i+startingVal)].value))
                else:
                    self.angles.append(0)
                    self.TEST_X.angleArray.append(0)
                    self.TEST_Y.angleArray.append(0)
                    self.TEST_Z.angleArray.append(0)
            
            
            for i in range(0,cnt):
                if(ws['B'+str(i+startingVal)].value!=None):
#                     self.zRawData.append(float(ws['B'+str(i+startingVal)].value))
                    self.TEST_Z.appendToRawData(float(ws['B'+str(i+startingVal)].value))
                else:
#                     self.zRawData.append(0)
                    self.TEST_Z.appendToRawData(0)
                    
            for i in range(0,cnt):
                if(ws['C'+str(i+startingVal)].value!=None):
#                     self.zCalData.append(float(ws['C'+str(i+startingVal)].value))
                    self.TEST_Z.appendToCalData(float(ws['C'+str(i+startingVal)].value))
#                     if self.zCalData[i]!=0 and self.TEST_Z.dataArrayCal[i]!=0:
                    if self.TEST_Z.dataArrayCal[i]!=0:    
                        drawz=True;
                else:
#                     self.zCalData.append(0)
                    self.TEST_Z.appendToCalData(0)
   
            for i in range(0,cnt):
                if(ws['D'+str(i+startingVal)].value!=None):
#                     self.xRawData.append(float(ws['D'+str(i+startingVal)].value))
                    self.TEST_X.appendToRawData(float(ws['D'+str(i+startingVal)].value))
                    
                else: 
#                     self.xRawData.append(0)   
                    self.TEST_X.appendToRawData(0) 
                    
            for i in range(0,cnt):
                if(ws['E'+str(i+startingVal)].value!=None):
#                     self.xCalData.append(float(ws['E'+str(i+startingVal)].value))
                    self.TEST_X.appendToCalData(float(ws['E'+str(i+startingVal)].value))
#                     if self.xCalData[i]!=0 and self.TEST_X.dataArrayCal[i]!=0:
                    if self.TEST_X.dataArrayCal[i]!=0:
                        drawx=True;
                else:
#                     self.xCalData.append(0) 
                    self.TEST_X.appendToCalData(0)             
                
            for i in range(0,cnt):
                if(ws['F'+str(i+startingVal)].value!=None):
#                     self.yRawData.append(float(ws['F'+str(i+startingVal)].value))
                    self.TEST_Y.appendToRawData(float(ws['F'+str(i+startingVal)].value))
                else:
#                     self.yRawData.append(0)
                    self.TEST_Y.appendToRawData(0) 
                    
            for i in range(0,cnt):
                if(ws['G'+str(i+startingVal)].value!=None):
#                     self.yCalData.append(float(ws['G'+str(i+startingVal)].value))
                    self.TEST_Y.appendToCalData(float(ws['G'+str(i+startingVal)].value))
#                     if self.yCalData[i]!=0 and self.TEST_Y.dataArrayCal[i]!=0:
                    if self.TEST_Y.dataArrayCal[i]!=0:
                        drawy=True;
                else:
#                     self.yCalData.append(0)    
                    self.TEST_Y.appendToCalData(0)   
                    
            #print opened data
            print self.angles
            print self.TEST_Z.dataArrayRaw
            print self.TEST_Z.dataArrayCal
            print self.TEST_X.dataArrayRaw
            print self.TEST_X.dataArrayCal
            print self.TEST_Y.dataArrayRaw
            print self.TEST_Y.dataArrayCal
            
            self.statusBar().showMessage('Opened file %s' % path, 2000)
            
            
            #===================================================================
            # draw data plots with opened data
            #===================================================================    
            
            if drawz:
                self.TEST_Z.setTitle(ws['C7'].value)
                self.TEST_Z.setHoldsData(True)
            else:
                self.TEST_Z.clearAllData()
                self.TEST_Z.setHoldsData(False)
            self.rb_axisSelZ.click()  
            
            
            if drawx:
                self.TEST_X.setTitle(ws['E7'].value)
                self.TEST_X.setHoldsData(True)
            else:
                self.TEST_X.clearAllData()
                self.TEST_X.setHoldsData(False)
            self.rb_axisSelX.click()
            
            if drawy:
                self.TEST_Y.setTitle(ws['G7'].value)
                self.TEST_Y.setHoldsData(True)
            else:
                self.TEST_Y.clearAllData()
                self.TEST_Y.setHoldsData(False)
            self.rb_axisSelY.click()    
            
    def save_plot(self):
        #=======================================================================
        #          Name:    save_plot
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    saves the data collection plots as a .png file
        #
        #=======================================================================
        if self.tabs.currentIndex()==0 or self.tabs.currentIndex()==1:
            'Saves plot as .png'
            file_choices = "PNG *.png"
            
            path = unicode(QFileDialog.getSaveFileName(self, 
                            'Save file', '', 
                            file_choices))
            if path:
                self.x_axis.set_facecolor('white')
                self.y_axis.set_facecolor('white')
                self.z_axis.set_facecolor('white')
                
                self.canvas.print_figure(path, dpi=self.dpi)
                self.statusBar().showMessage('Saved to %s' % path, 2000)
                
        elif self.tabs.currentIndex()==2:
            'Saves plot as .png'
            file_choices = "PNG *.png"
            
            path = unicode(QFileDialog.getSaveFileName(self, 
                            'Save file', '', 
                            file_choices))
            if path:
                self.emcCanvas.print_figure(path, dpi=self.dpi)
                self.statusBar().showMessage('Saved to %s' % path, 2000)
        elif self.tabs.currentIndex()==3:
            'Saves plot as .png'
            file_choices = "PNG *.png"
            
            path = unicode(QFileDialog.getSaveFileName(self, 
                            'Save file', '', 
                            file_choices))
            if path:
                self.canvas3d.print_figure(path, dpi=self.dpi)
                self.statusBar().showMessage('Saved to %s' % path, 2000)
        print self.tabs.currentIndex()
    
    def on_about(self):#display program information
        #=======================================================================
        #          Name:    on_about
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function displays applicaiton version info when 
        #                    the user selects "about" from the menu
        #
        #=======================================================================
        'calls about data'
        
        msg = "NGRF Rotator\r\n"\
                + "Version: " + version + "\r\n"\
                + "Author: " + author + "\r\n"\
                + "Contact: " + email + "\r\n"\
                + "Copyright " + year + "\r\n"\
                + website
        QMessageBox.about(self, "About", msg.strip())
    
    def click_manualTarget(self, event):#sets turntable target and begins test
        #=======================================================================
        #
        #          Name:    click_manualTarget
        #
        #    Parameters:    event(mouse click event)
        #
        #        Return:    None
        #
        #   Description:    this function will begin a manual test when the user
        #                    clicks a point of the data collection plot in manual mode
        #
        #=======================================================================
        """
        Uses the button_press_event to begin manual mode test
        """
        if self.manual_mode:
            print event.xdata
            print event.ydata
            worker_data=[event.xdata*180/3.14]
            self.worker.do_work(self.worker.Functions.goto_location,worker_data)

    def click_setup(self):#activates setup dialog
        #=======================================================================
        #
        #          Name:    click_setup
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function executes the setup dialog box when the user clicks the 
        #                    "setup" pushbutton
        #
        #=======================================================================
        'initiates setup dialog'
        
        self.setup.exec_()
        #=======================================================================
        # update line edit boxes in setup dialog
        #=======================================================================
        self.setup.e_span.setText(str(self.cal.cal_span/1e6))  
        self.setup.e_cfreq.setText(str(self.cal.cal_freq/1e6))
        self.setup.e_sweep.setText(str(self.cal.cal_sc_sweepTime*1e3))
   
    def on_data_ready(self,new_data):#sends raw data to data lists and starts drawing plots
        #=======================================================================
        #
        #          Name:    on_data_ready
        #
        #    Parameters:    new_data
        #
        #        Return:    None
        #
        #   Description:    this function adds new data to data collection arrays
        #                    when new data is available from the spectrum analyzer
        #                    this functions also redraws the selected data collection plot
        #                    when new data become available
        #
        #=======================================================================
        'adds new data to dat arrays, calls to redraw plot'
        #===================================================================
        # create arrays for drawing plot
        #===================================================================
        
        self.angles.append(new_data[0])
        
        self.data.append(self.cal.calibrate_data(new_data[1]))
        self.progress.setValue(new_data[0])
        
        
        if (self.rotationAxis=='Z'):
            self.TEST_Z.appendToAngleArray(new_data[0])
            self.TEST_Z.appendToRawData(new_data[1])
            self.TEST_Z.appendToCalData(self.cal.calibrate_data(new_data[1]))
            self.TEST_Z.setHoldsData(True)
        elif(self.rotationAxis=='X'):
            self.TEST_X.appendToAngleArray(new_data[0])
            self.TEST_X.appendToRawData(new_data[1])
            self.TEST_X.appendToCalData(self.cal.calibrate_data(new_data[1]))
            self.TEST_X.setHoldsData(True)
        elif(self.rotationAxis=='Y'):
            self.TEST_Y.appendToAngleArray(new_data[0])
            self.TEST_Y.appendToRawData(new_data[1])
            self.TEST_Y.appendToCalData(self.cal.calibrate_data(new_data[1]))
            self.TEST_Y.setHoldsData(True)
        
        self.draw_dataPlots()
        
    def click_start(self):#begin test
        #=======================================================================
        #
        #          Name:    click_start
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function begins the test when the user clicks
        #                    the "start" push button
        #
        #=======================================================================
        'begin test'
        
        RUN=True
        
        if(self.cal.get_tabState()==1):#only run test if calibration test type is set to radiation pattern or no calibration
            msg = "Cannot run Radiation Pattern test while Test Type is set to \"EMC Pre-Compliance\" in Calibration tab!" 
            QMessageBox.critical(self, "Error", msg)
            RUN=False
#             self.show_errorDialog("Calibration Error!", "Cannot run Radiation Pattern test while Test Type is set to \"EMC Pre-Compliance\" in Calibration tab", "Please ensure calibration Test Type is set to \"Radiation Pattern\" or \"No Calibration\"")
        elif(self.currentTest.getHoldsData()):
            quit_msg = "The selected plot currently Holds Data.\nWould you Like to overwrite the previously collected data?"
            reply = QMessageBox.question(self, 'Message', 
                             quit_msg, QMessageBox.Yes, QMessageBox.No)
        
            if reply == QMessageBox.Yes:
                RUN=True
            else:
                RUN=False
            
        if(RUN):
            self.test_type=PATTERN
            self.worker.set_test_type(PATTERN)      #set type of test in worker
            vals=self.setup.get_values()            #get resolution from setup dialog
            self.worker.set_resolution(vals[6])     #set resolution from setup dialog
            #=======================================================================
            # enable/disabkle gui buttons during test
            #=======================================================================
            self.b_pause.setEnabled(True)
            self.b_stop.setEnabled(True)
            self.b_start.setEnabled(False)
            self.b_setup.setEnabled(False)
            self.rb_axisSelZ.setEnabled(False)
            self.rb_axisSelX.setEnabled(False)
            self.rb_axisSelY.setEnabled(False)
            self.cal.b_specan.setEnabled(False)
            self.cal.testConfigEnable(False)
            
            #=======================================================================
            # apply settings to specan for test
            #=======================================================================
            
            self.cal.apply_specanSettings()
            
            #=======================================================================
            # get name of plotfordisplay
            #=======================================================================
#             text, ok = QInputDialog.getText(self, 'Name of data', 'Enter a data name:')
#             if ok:
#                 self.legend=str(text)
#                 self.update_figureInfo()
# #                 self.fig.suptitle('DUT: '+self.cal.cal_dutLabel+ '\nSerial No.: '+self.cal.cal_dutSN, fontsize=14, fontweight='bold')
# #                 self.fig.text(0.95, .1, "Customer: "+self.cal_customer+"/nOrientation: "+self.cal.cal_orentation+"\nTester: "+self.cal.cal_tester , verticalalignment='bottom', horizontalalignment='right', fontsize=12)
#             else:
#                 self.click_stop()
#                 return
            #=======================================================================
            # clear arrays that will store axis data
            #=======================================================================
            self.data=[]
            self.angles=[]
            
        
            if (self.rotationAxis=='Z'):
#                 self.zRawData=[]
#                 self.zCalData=[]
                self.TEST_Z.clearAllData()
                    
            elif(self.rotationAxis=='X'):
#                 self.xRawData=[]
#                 self.xCalData=[]
                self.TEST_X.clearAllData()
                
            elif(self.rotationAxis=='Y'):
#                 self.yRawData=[]
#                 self.yCalData=[]
                self.TEST_Y.clearAllData()
                
            self.worker.do_work(self.worker.Functions.rotate)    
            
    def update_figureInfo(self):
        today = datetime.date.today()
        
        self.fig.clf()
        self.fig.suptitle('Radiation Pattern Testing', fontsize=14, fontweight='bold')
        self.fig.text(.06, .005, "Customer: \nTester: \nDate: ", verticalalignment='bottom', horizontalalignment='right', fontsize=9)
        self.fig.text(.06, .005, self.cal.cal_customer+"\n"+self.cal.cal_tester+"\n"+today.strftime('%d, %b %Y') , verticalalignment='bottom', horizontalalignment='left', fontsize=9)
        
        
        
        ##Z
        test=self.TEST_Z
        test.setSubplot(self.fig.add_subplot(131,polar=True)) 
            
        while(1):  
            plt=test.getSubplot()
            plt.set_title(test.getTitle(),y=1.08, fontsize=10,fontweight=200) 
            plt.text(0.5,-0.1,"Testing Distance: \nRotation Axis: RBW: ", horizontalalignment='right', verticalalignment='top',transform=plt.transAxes)
            plt.text(0.5,-0.1,str(test.getDistance())+"\n"+str(self.curAxis)+"\n"+str(self.cal.cal_sc_rbw), horizontalalignment='left', verticalalignment='top',transform=plt.transAxes)
            
            if self.currentTest==test:
                plt.set_facecolor('white')
                self.axes=plt
            else:
                plt.set_facecolor('grey')
                
            if test==self.TEST_Z:
                test=self.TEST_X
                continue
            if test==self.TEST_X:
                test.setSubplot(self.fig.add_subplot(132,polar=True))
                test=self.TEST_Y
                continue
            if test==self.TEST_Y:
                test.setSubplot(self.fig.add_subplot(133,polar=True))
                test=self.TEST_Y
                break
        ##Draw        
        temp=self.rotationAxis
        
        #self.canvas.draw()
        self.rb_axisSelY.click()
        self.draw_dataPlots()
        self.rb_axisSelX.click()
        self.draw_dataPlots()
        self.rb_axisSelZ.click()
        self.draw_dataPlots()
        
        if (temp=='Z'):
            self.rb_axisSelZ.click()
        elif(temp=='X'):
            self.rb_axisSelX.click()
        elif(temp=='Y'):
            self.rb_axisSelY.click()    
            
    def click_stop(self):#abort current test
        #=======================================================================
        #          Name:    click_stop
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    this function manages the functionality when the 
        #                    user clicks the stop buttons, it disable/enable the appropriate
        #                    pushbuttons and change the state of the worker object.
        #                    the test cannot be started again from the same point
        #
        #=======================================================================
        'abort current test'
        self.b_run_test.setEnabled(True)
        self.b_abort.setEnabled(False)
        self.b_emcPause.setEnabled(False)
        self.b_pause.setEnabled(False)
        self.b_stop.setEnabled(False)
        self.b_setup.setEnabled(True)
        self.b_start.setEnabled(True)
        self.cal.b_specan.setEnabled(True)
        self.b_manual.setEnabled(True)
        self.rb_axisSelZ.setEnabled(True)
        self.rb_axisSelX.setEnabled(True)
        self.rb_axisSelY.setEnabled(True)
        self.worker.cancel_work=True
        self.emcTestRunning=False
        self.cal.testConfigEnable(True)
        
    def click_pause(self):#pause mid-test without reseting data
        #=======================================================================
        #          Name:    click_pause
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This function manages the functionality when the user
        #                    clicks the pause push button, it will stop the turn-table mid test
        #                    and change the state of the worker object
        #
        #=======================================================================
        'Pause mid-test'
        self.b_stop.setEnabled(not self.b_pause.isChecked())            
        self.worker.pause_work(self.b_pause.isChecked())
    
    def click_manual(self):#activates manual mode
        #=======================================================================
        #          Name:    click_manual
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function puts the application in and out of manual mode
        #                    and set disables/enables the appropriate push buttons
        #
        #=======================================================================
        'switch to manual mode'
        self.manual_mode=self.b_manual.isChecked()
        if self.manual_mode:
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_start.setEnabled(False)
#             self.rb_axisSelZ.setEnabled(False)
#             self.rb_axisSelX.setEnabled(False)
#             self.rb_axisSelY.setEnabled(False)
        else:
            self.b_pause.setEnabled(False)
            self.b_stop.setEnabled(False)
            self.b_start.setEnabled(True)
            self.cal.b_applytestConfig.setEnabled(True)
#             self.rb_axisSelZ.setEnabled(True)
#             self.rb_axisSelX.setEnabled(True)
#             self.rb_axisSelY.setEnabled(True)
            self.b_manual.setEnabled(True)
    
    def click_clear(self):#clears data from active axis list and plot
        #=======================================================================
        #          Name:    click_clear
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function clears the data from the selected axis plot
        #                    when the user clicks the clear button
        #
        #=======================================================================
        'clears data arrays and resest plot data'
        
        if self.currentTest.getHoldsData():
            open_msg = "Are you sure you want to delete this data?"
            reply = QMessageBox.warning(self, 'Data Overwrite Warning', open_msg, QMessageBox.Yes, QMessageBox.No)
            if reply == QMessageBox.No:
                return
        self.data=[]    #reset raw data list
        self.angles=[]  #reset angles list
        
        self.currentTest.clearAllData()
            
        self.update_figureInfo()
            
    def update_plot_settings(self):
        #=======================================================================
        #          Name:    update_plot_settings
        #
        #    Parameters:    None
        #
        #        Return:    None    
        #
        #   Description:    this function redraws the main data collection plot 
        #                    with or without the grid lines depending on user selection
        #
        #=======================================================================
        'update grid setting for Plot'
        self.axes.grid(self.grid_cb.isChecked())
        self.canvas.draw()
    
    def click_axisSelect(self):#select rotation axis for test
        #=======================================================================
        #          Name:    click_axisSelect
        #    
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function manages the functionality when the user
        #                    clicks one of the axis select radio buttons
        #
        #=======================================================================
        'select rotation axis for test'
        
        self.pltColor='000000'
        
        if (self.rb_axisSelX.isChecked()):
            self.rotationAxis='X'
            self.axes=self.x_axis
            self.axes=self.TEST_X.getSubplot()
            
            self.TEST_X.getSubplot().set_facecolor('white')
            self.TEST_Y.getSubplot().set_facecolor('grey')
            self.TEST_Z.getSubplot().set_facecolor('grey')
            self.currentTest=self.TEST_X
            self.pltColor=YCOLOR
            
        elif(self.rb_axisSelY.isChecked()):
            self.rotationAxis='Y'
            self.axes=self.y_axis
            self.axes=self.TEST_Y.getSubplot()
            self.TEST_X.getSubplot().set_facecolor('grey')
            self.TEST_Y.getSubplot().set_facecolor('white')
            self.TEST_Z.getSubplot().set_facecolor('grey')
            self.currentTest=self.TEST_Y
            self.pltColor=YCOLOR
        else:
            self.rotationAxis='Z'
            self.axes=self.z_axis
            self.axes=self.TEST_Z.getSubplot()
            self.TEST_X.getSubplot().set_facecolor('grey')
            self.TEST_Y.getSubplot().set_facecolor('grey')
            self.TEST_Z.getSubplot().set_facecolor('white')
            self.currentTest=self.TEST_Z
            self.pltColor=YCOLOR
         
        self.draw_dataPlots()
            
        #change curAxis label and format text
        self.curAxis.setText('<span style=" font-size:14pt; font-weight:600; color:'+self.pltColor+';">Current Rotation Axis: ' + str(self.rotationAxis)+'</span>')
        
        print "Current Rotation Axis: " + self.rotationAxis
        
        
    def default_dataPlot(self,title,title_color):#draw plots   
        #=======================================================================
        #
        #          Name:    default_dataPlots
        #
        #    Parameters:    (string)title, (string)color
        #
        #        Return:    None
        #
        #   Description:    resets plot to default settings
        #
        #=======================================================================   
        self.axes.clear()  
        self.axes.grid(self.grid_cb.isChecked())
        self.axes.set_title(title,fontsize=14,fontweight=300,color=title_color)
        
    def draw_dataPlots(self):#draw plots
        #=======================================================================
        #          Name:    draw_dataPlots
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This functions draws the main data collection plots on 
        #                    the data collection tab
        #
        #=======================================================================
        """ Redraws the figure
        """
        # clear the axes and redraw the plot anew
        self.axes.clear()  
        plt=self.axes
        test=self.currentTest     
        self.axes.grid(self.grid_cb.isChecked())
#         self.axes.set_title(self.legend,fontsize=14,fontweight=300)
        
        r = np.array(test.dataArrayCal)
        theta = np.array(test.angleArray) * np.pi / 180
        
        
        plt.plot(theta,r,lw=1,color='r')
        plt.set_title(test.getTitle(),y=1.08, fontsize=10,fontweight=200) 
        plt.text(0.5,-0.1,"Testing Distance: \nRotation Axis: \nRBW: ", horizontalalignment='right', verticalalignment='top',transform=plt.transAxes)
        plt.text(0.5,-0.1,str(test.getDistance())+"\n"+self.rotationAxis+"\n"+str(float(self.cal.cal_sc_rbw/1e3))+"KHz", horizontalalignment='left', verticalalignment='top',transform=plt.transAxes)
            #set up grid for plot
            
            
        if len(r)>0:
            if self.customScaling:
                gridmin=self.gridmin
            else:
                gridmin=10*round(np.amin(r)/10)
            if gridmin>np.amin(r):
                gridmin = gridmin-10
            if self.customScaling:
                gridmax=self.gridmax
            else:
                gridmax=10*round(np.amax(r)/10)
            if gridmax < np.amax(r):
                gridmax=gridmax+10
 
                      
            if abs(gridmax)>0 and abs(gridmin)>0:
                self.axes.set_ylim(gridmin,gridmax)
                self.axes.set_yticks(np.arange(gridmin,gridmax,(gridmax-gridmin)/5))
        
        self.canvas.draw()

    def create_tabs(self):#create tab architecture for application
        #=======================================================================
        #          Name:    create_tabs
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function creates theQTabWidget that holds the 
        #                    tabs of the application, it also populates with the application's tabs 
        #
        #=======================================================================

        #create tab widget to hold tabs
        #self.t_bar=QTabBar(shape=QTabBar.TriangularEast)
        self.tabs=QTabWidget()
        
        #create data collection tab
        
        self.tab_dataCollection= QWidget()
        self.tabs.addTab(self.tab_dataCollection,"Data Collection")
        
        #create calibration tab
        self.tab_calibration=QWidget()
        self.tabs.addTab(self.tab_calibration,"Calibration")
        
        #create emc testing tab
        self.tab_emc=QWidget()
        self.tabs.addTab(self.tab_emc,"EMC")
        
        #create 3d imaging tab
        self.tab_3D=QWidget()
        self.tabs.addTab(self.tab_3D,"3D Rendering")
        
        self.setCentralWidget(self.tabs)
        

    def click_applyScaling(self):
        self.customScaling = True
        
        min=int(self.e_scaleMin.text())
        max=int(self.e_scaleMax.text())
        if min>max:
            self.show_errorDialog("invalid input", "Minimum value is larger than maximum value", "")
            self.e_scaleMax.setText("1")
            self.e_scaleMin.setText("0")
            return
        
        
        self.gridmin=min
        self.gridmax=max
        
#         self.draw_dataPlots()
        self.update_figureInfo()
        
    def create_3dTab(self):#create 3d rendering tab
        #=======================================================================
        #          Name:    create_3dTab
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    this function create the form and user interface of the 
        #                   3d plotting tab
        #
        #=======================================================================
        
        #==========================================================================
        #create Label for current axis V2.0
        #===========================================================================
        
        self.l_3d=QLabel()
        self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:blue;">Render 3D Interpolation of Radiation Pattern</span>')
        self.l_3d.setAlignment(Qt.AlignLeft)
        
        #=======================================================================
        # create figure and canvas for 3d rendering
        #=======================================================================
        self.fig3d = Figure(figsize=(6.0, 6.0), dpi=self.dpi)
        self.canvas3d = FigureCanvas(self.fig3d)
        self.fig3d.set_facecolor('#8E8E8E')
        #=======================================================================
        # self.canvas.setParent(self.main_frame)
        #=======================================================================
        self.canvas3d.setParent(self.tab_3D)
    
        #=======================================================================
        # create "Render" button
        #=======================================================================
        self.b_render= QPushButton("&Render")
        self.b_render.setEnabled(True)
        self.connect(self.b_render, SIGNAL('clicked()'), self.draw_3dPlot)
        self.b_render.setToolTip("render collected data in 3D, NOTE: will only work when all 3 axes have data")
        
        #=======================================================================
        # create subplots
        #=======================================================================
        self.plt3dx = self.fig3d.add_subplot(143,projection='3d', aspect='equal')
        self.plt3dy = self.fig3d.add_subplot(144,projection='3d', aspect='equal')
        self.plt3dz = self.fig3d.add_subplot(142,projection='3d', aspect='equal')
        self.plt3dISO = self.fig3d.add_subplot(141,projection='3d', aspect='equal')
        #===================================================================
        # create colobars to show field strength
        #===================================================================
        m = matplotlib.cm.ScalarMappable(cmap=matplotlib.cm.jet)
        m.set_array(self.data)
        
        self.cbarISO = self.fig3d.colorbar(m,ax=self.plt3dISO, orientation="horizontal")
        self.cbarISO.set_label('Power (dBm)',)
        
        self.cbarx = self.fig3d.colorbar(m,ax=self.plt3dx, orientation="horizontal")
        self.cbarx.set_label('Power (dBm)',)
        
        self.cbary = self.fig3d.colorbar(m,ax=self.plt3dy, orientation="horizontal")
        self.cbary.set_label('Power (dBm)',)
        
        self.cbarz = self.fig3d.colorbar(m,ax=self.plt3dz, orientation="horizontal")
        self.cbarz.set_label('Power (dBm)',)
        
        self.canvas3d.draw()
        
        hbox3d = QHBoxLayout()
        
        #=======================================================================
        # create button bar
        #=======================================================================
        hbox3d.addStretch()
        for w in [self.b_render]:
            hbox3d.addWidget(w)
            hbox3d.setAlignment(w, Qt.AlignVCenter)
        hbox3d.addStretch()
            
        vbox3d = QVBoxLayout()              #create layout 
        vbox3d.addWidget(self.l_3d)     
        vbox3d.addWidget(self.canvas3d,10)  #add graph area to display
        vbox3d.addWidget(self.mpl_toolbar)  #add matplotlib toolbar to display
        vbox3d.addLayout(hbox3d)            #add control buttons to display
        
        self.tab_3D.setLayout(vbox3d)
        
    def draw_3dPlot(self):#Draws 3D interpolation of data
        #=======================================================================
        #          Name:    draw_3dPlot
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    This function draws the 3d plot of the collected data
        #
        #=======================================================================
        'Draws 3D interpolation of data'
        
        self.b_render.setEnabled(False)             #disable button while rendering
        
        self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; black;">Rendering..........</span>')
        
        #initialize variables for interpolation and drawing
        theta, phi = np.linspace(0, 2 * np.pi, 100), np.linspace(0, np.pi, 50)

        THETA, PHI = np.meshgrid(theta, phi)
        
        R = np.ones((50,100)) 
        
        #=======================================================================
        #check if data arrays have any test data in them
        #if they are empty the app will not render them
        #=======================================================================
        
        zhasData=False
        xhasData=False
        yhasData=False
        
        for i in self.TEST_Z.dataArrayCal:
            if(i!=0):                   #set draw to true if array has data
                zhasData=True
        for i in self.TEST_X.dataArrayCal:
            if(i!=0):                   #set draw to true if array has data
                xhasData=True
        for i in self.TEST_Y.dataArrayCal:
            if(i!=0):                   #set draw to true if array has data
                yhasData=True  
         
    #======================================================================
    # create 3Dimage from 3 axes projections
    #======================================================================
        if(zhasData and xhasData and yhasData):#only draw if all axes have data
            for th in range(100):
                for rh in range(50):
                    if th<25:
                        if rh<25:
                            
                            xzInterp = self.interp(rh,0,self.TEST_X.dataArrayCal[75+rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            yzInterp = self.interp(rh,0,self.TEST_Y.dataArrayCal[75+rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
                            
                        else:
                            xzInterp = self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_X.dataArrayCal[rh-25])
                            
                            yzInterp = self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_Y.dataArrayCal[rh-25])
                            
                            xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
                     
                    elif th<50:
                        if rh<25:
                             
                            xzInterp=self.interp(rh,0,self.TEST_X.dataArrayCal[75-rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            yzInterp=self.interp(rh,0,self.TEST_Y.dataArrayCal[75+rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
                        else:
                             
                            xzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_X.dataArrayCal[75-rh])
                             
                            yzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_Y.dataArrayCal[rh-25])
                             
                            xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
                      
                    elif th<75:
                        if rh<25:
                             
                            xzInterp=self.interp(rh,0,self.TEST_X.dataArrayCal[75-rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            yzInterp=self.interp(rh,0,self.TEST_Y.dataArrayCal[75-rh],24,self.TEST_Z.dataArrayCal[100-th])
                             
                            xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
                             
                        else:
                            xzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_X.dataArrayCal[75-rh])
                             
                            yzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_Y.dataArrayCal[75-rh])
                             
                            xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
                       
                    else:
                        if rh<25:
                            xzInterp=self.interp(rh,0,self.TEST_X.dataArrayCal[75+rh],24,self.TEST_Z.dataArrayCal[100-th])
                            
                            yzInterp=self.interp(rh,0,self.TEST_Y.dataArrayCal[75-rh],24,self.TEST_Z.dataArrayCal[100-th])
                            
                            xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
                            
                        else:
                            
                            xzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_X.dataArrayCal[rh-25])
                            
                            yzInterp=self.interp(rh,25,self.TEST_Z.dataArrayCal[100-th],49,self.TEST_Y.dataArrayCal[75-rh])
                            
                            xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
#                         
#                     if th<25:
#                         if rh<25:
#                             
#                             xzInterp = self.interp(rh,0,self.xCalData[75+rh],24,self.zCalData[100-th])
#                              
#                             yzInterp = self.interp(rh,0,self.yCalData[75+rh],24,self.zCalData[100-th])
#                              
#                             xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
#                             
#                         else:
#                             xzInterp = self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[rh-25])
#                             
#                             yzInterp = self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[rh-25])
#                             
#                             xyInterp = self.interp(th,0,xzInterp,24,yzInterp)
#                      
#                     elif th<50:
#                         if rh<25:
#                              
#                             xzInterp=self.interp(rh,0,self.xCalData[75-rh],24,self.zCalData[100-th])
#                              
#                             yzInterp=self.interp(rh,0,self.yCalData[75+rh],24,self.zCalData[100-th])
#                              
#                             xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
#                         else:
#                              
#                             xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[75-rh])
#                              
#                             yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[rh-25])
#                              
#                             xyInterp=self.interp(th,25,yzInterp,49,xzInterp)
#                       
#                     elif th<75:
#                         if rh<25:
#                              
#                             xzInterp=self.interp(rh,0,self.xCalData[75-rh],24,self.zCalData[100-th])
#                              
#                             yzInterp=self.interp(rh,0,self.yCalData[75-rh],24,self.zCalData[100-th])
#                              
#                             xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
#                              
#                         else:
#                             xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[75-rh])
#                              
#                             yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[75-rh])
#                              
#                             xyInterp=self.interp(th,50,xzInterp,74,yzInterp)
#                        
#                     else:
#                         if rh<25:
#                             xzInterp=self.interp(rh,0,self.xCalData[75+rh],24,self.zCalData[100-th])
#                             
#                             yzInterp=self.interp(rh,0,self.yCalData[75-rh],24,self.zCalData[100-th])
#                             
#                             xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
#                             
#                         else:
#                             
#                             xzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.xCalData[rh-25])
#                             
#                             yzInterp=self.interp(rh,25,self.zCalData[100-th],49,self.yCalData[75-rh])
#                             
#                             xyInterp=self.interp(th,75,yzInterp,99,xzInterp)
                    #Radius is equal to this interpolation         
                    R[rh,th]=xyInterp

            #create unchanged copy of Radius array for colorbars
            cbR=R.copy()
            
            #add lowest value to all Radius array elements so all values will be >=0
            R+=abs(np.amin(R))
            
            #change array values from spherical to cartesian and multiply by radius
            X = R * np.sin(PHI) * np.cos(THETA)
            Y = R * np.sin(PHI) * np.sin(THETA)
            Z = R * np.cos(PHI)
            
            #set color of plot to change based on value of radius
            strengthColor = matplotlib.cm.jet((R)/np.amax(R))
            
            #===================================================================
            # draw 3D plots
            #===================================================================
            self.plt3dISO.cla()
            self.plt3dx.cla()
            self.plt3dy.cla()
            self.plt3dz.cla()
            
            # ISOMetric
            self.plt3dISO.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dISO.view_init(elev=32, azim=-45) 
            
            # X Axis
            self.plt3dx.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dx.view_init(elev=0, azim=270)    

            # Y axis
            self.plt3dy.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dy.view_init(elev=0, azim=0)   
               
            # Z Axis
            self.plt3dz.plot_surface(X, Y, Z, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=1,facecolors = strengthColor)
            self.plt3dz.view_init(elev=90, azim=270)     
            
            #Ensure 3D plots are square so images is not distorted
            self.axisEqual3D(self.plt3dx)
            self.axisEqual3D(self.plt3dy)
            self.axisEqual3D(self.plt3dz)
            self.axisEqual3D(self.plt3dISO)
            
            
            self.plt3dx.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dISO.w_xaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            self.plt3dx.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dISO.w_yaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            self.plt3dx.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dy.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dz.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            self.plt3dISO.w_zaxis.set_pane_color((1.0, 1.0, 1.0, 1.0))
            
            #set 3d plot titles
            self.plt3dx.set_title('Y-Z Plane (Rotation on X-Axis)',y=1, size=12)    
            self.plt3dy.set_title('X-Z Plane (Rotation on Y-Axis)',y=1, size=12)
            self.plt3dz.set_title('X-Y Plane (Rotation on Z-Axis)',y=1, size=12)
            self.plt3dISO.set_title('Isometric',y=1, size=12)
            
            #set plot axis labels
            self.plt3dISO.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dISO.set_ylabel("X")
            self.plt3dISO.set_zlabel("Z")
            
            self.plt3dx.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dx.set_ylabel("X")
            self.plt3dx.set_zlabel("Z")
            
            self.plt3dy.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dy.set_ylabel("X")
            self.plt3dy.set_zlabel("Z")
            
            self.plt3dz.set_xlabel("Y")     #X and Y Axis Flipped but still accuarate
            self.plt3dz.set_ylabel("X")
            self.plt3dz.set_zlabel("Z")
            
            #remove ticks
            self.plt3dISO.set_xticks([], [])
            self.plt3dISO.set_yticks([], [])
            self.plt3dISO.set_zticks([], [])
            
            self.plt3dx.set_xticks([], [])
            self.plt3dx.set_yticks([], [])
            self.plt3dx.set_zticks([], [])
            
            self.plt3dy.set_xticks([], [])
            self.plt3dy.set_yticks([], [])
            self.plt3dy.set_zticks([], [])
            
            self.plt3dz.set_xticks([], [])
            self.plt3dz.set_yticks([], [])
            self.plt3dz.set_zticks([], [])
            
            #===================================================================
            # update colobars
            #===================================================================
            
            #set color values for colorbar
            m = cm.ScalarMappable(cmap=matplotlib.cm.jet)
            m.set_array(cbR)
            
            #x colorbar
            self.cbarISO.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbarISO.draw_all()
            
            #x colorbar
            self.cbarx.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbarx.draw_all()

            #y colorbar
            self.cbary.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbary.draw_all()
            
            #z colorbar
            self.cbarz.set_clim(vmin=np.amin(cbR),vmax=np.amax(cbR)) 
            self.cbarz.draw_all()
            
            #set tight layout and draw plots
            self.fig3d.tight_layout()
            self.canvas3d.draw()
            
            #===================================================================
            # warning/ information label text
            #===================================================================
            self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:blue;">3D Interpolation of Radiation Pattern</span>')
        else:
            self.l_3d.setText('<span style=" font-size:14pt; font-weight:600; color:red;">Unable to plot! One or more data arrays are empty!</span>')
            
        
        self.b_render.setEnabled(True)      #enable button after rendering
        
        self.render3D=True                  #set render 3d to true so 3D plot will be added to test report
        
        #Mayavi rendering
#         mlab.figure(size=(1000, 800))
#         mlab.mesh(X, Y, Z,scalars=cbR-.2)
#         mlab.show()
        
    def interp(self,x,x1,y1,x3,y3): 
        #=======================================================================
        #
        #          Name:    interp
        #
        #    Parameters:    (float) x, x1, y1, x3, y3
        #
        #        Return:    (float) retval
        #
        #   Description:    returns the linear interpolation at point x between points (x1,y1) and (x3,y3)
        #
        #=======================================================================
        retval=(((x-x1)*(y3-y1))/(x3-x1))+y1
        return retval
    
    def axisEqual3D(self,ax):
        #=======================================================================
        #
        #          Name:    axisEqual3D
        #
        #    Parameters:    (subplot of matplotlib figure) ax
        #
        #        Return:    None
        #
        #   Description:    THis function sets the axes equal in a 3D matplotlib plot
        #                    so the image wont be distorted
        #
        #=======================================================================
        extents = np.array([getattr(ax, 'get_{}lim'.format(dim))() for dim in 'xyz'])
        sz = extents[:,1] - extents[:,0]
        centers = np.mean(extents, axis=1)
        maxsize = max(abs(sz))
        r = maxsize/2
        for ctr, dim in zip(centers, 'xyz'):
            getattr(ax, 'set_{}lim'.format(dim))(ctr - r, ctr + r)
            
    def click_emcPause(self):#pause mid-test without reseting data
        #=======================================================================
        #          Name:    click_pause
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This function manages the functionality when the user
        #                    clicks the pause push button, it will stop the turn-table mid test
        #                    and change the state of the worker object
        #
        #=======================================================================
        'Pause mid-test'
        self.b_abort.setEnabled(not self.b_emcPause.isChecked())            
        self.worker.pause_work(self.b_emcPause.isChecked())       
         
    def create_emcTab(self,tab):#create EMC testing tab
        #=======================================================================
        #          Name:    create_emcTab
        #
        #    Parameters:    (pointer to tab object)tab
        #
        #        Return:    None
        #
        #   Description:    this function creates the form and user interface of the EMC testing Tab
        #
        #=======================================================================
        'create EMC pre compliance testing tab'
        
        #set tab style
        tab.setStyleSheet(self.cal.create_styleSheet('dataTab'))
        
        #=======================================================================
        # create test result plot
        #=======================================================================

#         self.fill_dataArray()#populate data arrays
            
        self.figEmc = Figure(figsize = (10,6))
        self.emcCanvas = FigureCanvas(self.figEmc)
        self.emcCanvas.setParent(self.tab_emc)
        self.emcPlot=self.figEmc.add_subplot(111)
        self.emcPlot.grid(True)
        self.emcPlot.set_title('EMC Compliance Testing\nField Strength vs. Frequency',fontsize=14,fontweight=200)
        self.figEmc.tight_layout()
        
        #=======================================================================
        # create warning label groupbox
        #=======================================================================
        
        #create distance warning message area
        self.emc_distWarning=QLabel('<span style="  color:lime; font-size:11pt; font-weight:400;">--Distance Good-- Ready to run test</span>')
        
        #=======================================================================
        # Create "Run Test" pushbutton
        #=======================================================================
        self.b_run_test= QPushButton("&Run Test")
        self.b_run_test.setEnabled(False)
        self.connect(self.b_run_test, SIGNAL('clicked()'), self.run_emcTest)
        self.b_run_test.setToolTip("Run EMC pre-compliance test on collected data")
        
        self.b_abort= QPushButton("&Abort")
        self.b_abort.setEnabled(False)
        self.connect(self.b_abort, SIGNAL('clicked()'), self.click_stop)
        self.b_abort.setToolTip("stop")
        
        self.b_emcPause= QPushButton("&Pause")
        self.b_emcPause.setEnabled(False)
        self.connect(self.b_emcPause, SIGNAL('clicked()'), self.click_emcPause)
        self.b_emcPause.setToolTip("pause")
        #=======================================================================
        # create Regulation selection area
        #=======================================================================
        
        #select regulations (FCC/CISPR)
        self.regs=QGroupBox("Compliance Testing Regulations")
        regVbox=QVBoxLayout()
        regVbox.setAlignment(Qt.AlignCenter)
        rHbox=QHBoxLayout()
        
        self.r_fcc=QRadioButton('FCC')
        self.r_cispr=QRadioButton('CISPR')
        
        rHbox.addWidget(self.r_fcc)
        rHbox.addWidget(self.r_cispr)
        self.connect(self.r_fcc, SIGNAL('clicked()'), self.set_emcRegulations)
        self.connect(self.r_cispr, SIGNAL('clicked()'), self.set_emcRegulations)
        self.r_fcc.setToolTip("Select radiation regulations")
        
        self.r_cispr.setToolTip("Select radiation regulations")
        regVbox.addLayout(rHbox)
        self.regs.setLayout(regVbox)
        
        #select class (A/B)
        classbox=QGroupBox("DUT Class")
        cVbox=QVBoxLayout()
        cVbox.setAlignment(Qt.AlignCenter)
        cHbox=QHBoxLayout()
        
        self.r_classA=QRadioButton('Class A')
        self.r_classB=QRadioButton('Class B')
        
        cHbox.addWidget(self.r_classA)
        cHbox.addWidget(self.r_classB)
        
        self.connect(self.r_classA, SIGNAL('clicked()'), self.set_emcRegulations)
        self.connect(self.r_classB, SIGNAL('clicked()'), self.set_emcRegulations)
        
        self.r_classA.setToolTip("Select radiation regulations")
        self.r_classB.setToolTip("Select radiation regulations")
        cVbox.addLayout(cHbox)
        classbox.setLayout(cVbox)
        
        #=======================================================================
        # create polarity selection box
        #=======================================================================
        #select regulations (FCC/CISPR)
        polBox=QGroupBox("Select Antenna Polarity")
        polBoxLayout=QHBoxLayout()
        
        self.r_Vertical=QRadioButton('Vertical')
        self.r_Horizontal=QRadioButton('Horizontal')
        
        polBoxLayout.addWidget(self.r_Vertical)
        polBoxLayout.addWidget(self.r_Horizontal)
        self.connect(self.r_Vertical, SIGNAL('clicked()'), self.set_emcAntPolarity)
        self.connect(self.r_Horizontal, SIGNAL('clicked()'), self.set_emcAntPolarity)
        self.r_Vertical.setToolTip("Set Antenna polarity to Vertical")
        self.r_Horizontal.setToolTip("Set Antenna polarity to Horizontal")

        self.r_Vertical.click()
        polBox.setLayout(polBoxLayout)
        
        #=======================================================================
        # create Test settings groupbox
        #=======================================================================
        lfbox=QFormLayout()
        lfbox.setAlignment(Qt.AlignCenter)
        #Left form
        
        #=======================================================================
        # create test information groupbox
        #=======================================================================
        infoBox=QGroupBox("Test Information")
        infoBoxLayout=QFormLayout()
        infoBox.setLayout(infoBoxLayout)
        self.emc_gui_dist=QLabel("")
        infoBoxLayout.addRow(QLabel("Test Distance (m):"),self.emc_gui_dist)
        
        lfbox.addRow(infoBox)
        
        #=======================================================================
        # create error margin groupbox
        #=======================================================================
        errorMarginBox=QGroupBox("Error Margin")
        errorMarginBoxLayout=QFormLayout()
        errorMarginBox.setLayout(errorMarginBoxLayout)
        self.e_emc_margin  =QLineEdit('2')
        errorMarginBoxLayout.addRow(QLabel("Error Margin (dBuV/m)"),self.e_emc_margin)
        lfbox.addRow(errorMarginBox);
        
        #=======================================================================
        # create far field groupbox
        #=======================================================================
        farFieldBox=QGroupBox("Test Resolution")
        farFieldBoxLayout=QFormLayout()
        farFieldBox.setLayout(farFieldBoxLayout)
        
        self.e_emc_resolution  =QLineEdit('1')
        farFieldBoxLayout.addRow(QLabel("Number of Data Points"),self.e_emc_resolution)
#         self.e_emc_resolution.returnPressed.connect(self.set_emcRegulations)
        
#         b_checkFarfield = QPushButton("Check")
#         b_checkFarfield.clicked.connect(self.set_emcRegulations)
#         b_checkFarfield.setToolTip("check far field to ensure testing is valid")
#         farFieldBoxLayout.addRow(QLabel("Check Farfield"),b_checkFarfield)
        
        lfbox.addRow(farFieldBox)
        #add regulations to form layout
        lfbox.addRow(self.regs)
        #add class selection to form layout
        lfbox.addRow(classbox)
        
        lfbox.addRow(polBox)
        #=======================================================================
        # create test results Group box
        #=======================================================================
        #format style of test results
        self.emc_testResults=QLabel('<span style="  color:yellow; font-size:14pt; font-weight:400;">-----No Test Data-----</span>')
        self.emc_testResults.setAlignment(Qt.AlignLeft)
        
        self.emc_VtestFailNum=QLabel('<span style="  color:lime; font-size:14pt; font-weight:400;">-----No Vertical Failures Identified-----</span>')
        self.emc_VtestFailNum.setAlignment(Qt.AlignLeft)
        
        self.emc_HtestFailNum=QLabel('<span style="  color:lime; font-size:14pt; font-weight:400;">-----No Horizontal Failures Identified-----</span>')
        self.emc_HtestFailNum.setAlignment(Qt.AlignLeft)
        
        #setup result layout
        resultsBox=QGroupBox("Test Results")
        resultsBox.setStyleSheet(self.cal.create_styleSheet('EMC2'))#apply styling
        resultsBoxLayout=QVBoxLayout()
        resultsBox.setLayout(resultsBoxLayout)
        resultsBoxLayout.addWidget(self.emc_testResults)
        resultsBoxLayout.addWidget(self.emc_VtestFailNum)
        resultsBoxLayout.addWidget(self.emc_HtestFailNum)
        #=======================================================================
        # set some defaults for EMC tab 
        #=======================================================================
        #set default radio buttons select to default
        self.r_classB.click()       #set button to default
        self.r_fcc.click()          #set button to default
        
        #set regulations to default so warnings display correct message
        self.set_emcRegulations()

        #=======================================================================
        # Create Layout for EMC Testing
        #=======================================================================
        
        vbox=QVBoxLayout()
        vbox.setAlignment(Qt.AlignCenter)
        vbox.setAlignment(Qt.AlignTop)
        
        #create warning groupbox
        warningBox=QGroupBox("Warnings")
        warningBox.setStyleSheet(self.cal.create_styleSheet('EMC2'))#apply styling
        warningBoxLayout=QVBoxLayout()
        warningBox.setLayout(warningBoxLayout)
        warningBoxLayout.addWidget(self.emc_distWarning)
        vbox.addWidget(warningBox)

        #create horizontal box to hold setup gui and plot
        hbox=QHBoxLayout()
        
        #create settings group box
        settingBox=QGroupBox("Test Settings")
        settingBox.setLayout(lfbox)
        settingBox.setStyleSheet(self.cal.create_styleSheet('setup'))#apply styling
        
        #create vertical box to hold settings
        settingsVBox=QVBoxLayout();
        
        #populate settings vertical box
        settingsVBox.addWidget(settingBox)
        settingsVBox.addWidget(self.b_run_test)
        settingsVBox.addWidget(self.b_abort)
        settingsVBox.addWidget(self.b_emcPause)
        
        #populate horizontal box
        hbox.addLayout(settingsVBox)
        hbox.addStretch()
        hbox.addWidget(self.emcCanvas)
        
        #add horizontal to tab layout
        vbox.addLayout(hbox)
        
        #add test results to tab layout
        vbox.addWidget(resultsBox)
        
        #set tab layout
        self.tab_emc.setLayout(vbox)
        
    def get_emcTestLimit(self,target):#return the max field strength in uV/m type='fcc' or 'CISPR' target = target frequency in Hz
        
        #=======================================================================
        #          Name:    get_emcTestLimit
        #
        #    Parameters:    Target(frequency in Hz)
        #
        #        Return:    (float) value of FCC or CISPR limit for target frequency
        #
        #   Description:    given the test frequency and regulation type
        #                   this function returns the maximum legal EMC in dBuV/m
        #
        #=======================================================================
        
        retval=0
        
        if (self.emc_regs=='FCC'):#return FCC values
            if self.emc_class=='A':
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @3 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 49.5
                    elif target<216e6:      #88-216MHz
                        retval = 54
                    elif target<906e6:      #216-906MHz
                        retval = 56.5
                    else:                   #906MHz-40GHz
                        retval = 60
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @10 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 40
                    elif target<216e6:      #88-216MHz
                        retval = 43.5
                    elif target<906e6:      #216-906MHz
                        retval = 46
                    else:                   #906MHz-40GHz
                        retval = 54
                if(self.cal.cal_dist==30):
                    #===============================================================
                    # FCC Class A Max Values in dBuV/m @30 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 29.5
                    elif target<216e6:      #88-216MHz
                        retval = 34
                    elif target<906e6:      #216-906MHz
                        retval = 36.5
                    else:                   #906MHz-40GHz
                        retval = 40
            else:
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # FCC Class B Values in dBuV/m @3 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 40
                    elif target<216e6:      #88-216MHz
                        retval = 43.5
                    elif target<906e6:      #88-906MHz
                        retval = 46
                    else:                   #906MHz-40GHz
                        retval = 54
                        
                if(self.cal.cal_dist==10):
                    #===============================================================
                    # FCC Class B Values in dBuV/m @10 meters
                    #===============================================================
                    if target<88e6:         #30-88MHz
                        retval = 29.5
                    elif target<216e6:      #88-216MHz
                        retval = 33
                    elif target<906e6:      #216-906MHz
                        retval = 35.5
                    else:                   #906MHz-40GHz
                        retval = 43.5
        else:#return CISPR values
            if self.emc_class=='A':
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # CISPR Class A Max Values in dBuV/m @ 3 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 50.46


                    else:                   #>230MHz
                        retval = 57.46


                if(self.cal.cal_dist==10):
                    #===============================================================
                    # CISPR Class A Max Values in dBuV/m @ 10 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 40
                    else:                   #>230MHz
                        retval = 47
                if(self.cal.cal_dist==30):
                    #===============================================================
                    # CISPR Class A Max Values in dBuV/m @ 30 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 30
                    else:                   #>230MHz
                        retval = 37
            else:
                if(self.cal.cal_dist==3):
                    #===============================================================
                    # CISPR Class B Max Values in dBuV/m @ 3 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 40.46

                    else:                   #>230MHz
                        retval = 47.46

                if(self.cal.cal_dist==10):
                    #===============================================================
                    # CISPR Class B Max Values in dBuV/m @ 10 meters
                    #===============================================================
                    if target<230e6:        #<230MHz
                        retval = 30
                    else:                   #>230MHz
                        retval = 37
            
        #print 'retval '+ str(retval)
        
        return retval
    def create_cal_arrays(self):
        #=======================================================================
        #
        #          Name:    create_cal_arrays
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    this function builds a 10000 element long array of 
        #                    calibration amounts based on frequency it will display a 
        #                    dialog box showing completion percentage when running
        #
        #=======================================================================
        
        progress=QProgressDialog(labelText="Creating Calibration Array",minimum=0,maximum=9999)
        progress.setCancelButton(None)
        progress.show()
        
        if (self.emc_regs=='FCC'):
            SPAN=5.969999e9
        elif (self.emc_regs=='CISPR'):
            SPAN=969e6
        
        freqResolution=int(SPAN/10000)
        
        self.calArray=[]
        self.calFreqArray=[]
        
        for i in range(10000):
            self.calFreqArray.append(int(30e6+i*freqResolution))
            self.calArray.append(self.cal.get_calNum(30e6+freqResolution*i))
            
            progress.setValue(i)
        progress.close()
    def condense_EMCResults(self):
        for i in self.EMC_Vertical_Results:
            pass
        
    def run_emcTest(self,tableAngle=0):#run EMC Test
        #=======================================================================
        #    Name:            run_emcTest
        #
        #    Parameters:      None
        #
        #    Return:          None  
        #
        #    Description:     tests all data arrays and shows results in graph and 
        #                     in results label
        #
        #=======================================================================
        
        if(self.cal.get_tabState()!=1):#only run test if calibration test type is set to radiation pattern or no calibration
            msg = "Cannot run EMC Pre-Compliance test while Test Type is set to \"Radiation Pattern\" or \"No Calibration\" in Calibration tab!" 
            QMessageBox.critical(self, "Error", msg)
#             self.show_errorDialog("Calibration Error!", "Cannot run EMC Pre-Compliance test while Test Type is set to \"Radiation Pattern\" or \"No Calibration\" in Calibration tab", "Please ensure calibration Test Type is set to \"EMC Pre-Compliance\"")
        else:
            #=======================================================================
            # Setup testing Bandwidth
            #=======================================================================
            checkingFailSpan = False
            
            if (self.emc_regs=='FCC'):
                #fcc is broken into values from 30Mhz-1Ghz and 1GHz-6GHz
                
                #1 GHz ~ 6GHz
                SPANhigh=5e9
                CENTERhigh=2500e6+1e9
                
                #30MHz ~ 1 GHz
                SPANlow=970e6
                CENTERlow=485e6+30e6
                
                #create start and stop frequencies of high bandwidth
                STARThigh=int(CENTERhigh-int(SPANhigh/2))
                ENDhigh=int(CENTERhigh+int(SPANhigh/2)) 
                
            elif (self.emc_regs=='CISPR'):
                #1 GHz ~ 6GHz
                SPANlow=969e6
                CENTERlow=485e6+30e6
               
            #create start and stop frequencies of low bandwidth 
            STARTlow=int(CENTERlow-int(SPANlow/2))
            ENDlow=int(CENTERlow+int(SPANlow/2))   
            
            #set error margin from user input
            margin = float(self.e_emc_margin.text())
            
            #=======================================================================
            # initialize test: will only run at begining of test
            #=======================================================================
            if (self.emcTestRunning==False):        #prevent this script from running once test has started
                self.b_run_test.setEnabled(False)   #disable run test button while testing
                self.b_abort.setEnabled(True)       #enable abort test button while testing
                self.b_emcPause.setEnabled(True)    #enable pause test button while testing
                
                print 'Running EMC TEST\n\t---Begin---\nMargin: '+str(float(self.e_emc_margin.text()))
                
                self.worker.set_resolution(self.e_emc_resolution.text())    #set number of data points to take
                print "creating calibration array"
                self.create_cal_arrays()                                    #create array of calibrated values to add to test data later
                
    #             self.emc_reportRow=2                                      #set the
                
                #set test type for worker and self
                self.test_type=EMC
                
                self.worker.set_test_type(EMC)  #set type of test so worker perform the correct test
                
                #clear result arrays for selected polarity
                if(self.emc_polarity==POL_V):
                    self.EMC_Vertical_Results=[]
                    self.EMC_VerticalFailureNum=0
                else:
                    self.EMC_Horizontal_Results=[]
                    self.EMC_HorizontalFailureNum=0
                
                #reset failure limit reacehed to false    
                self.emc_failureLimitReached=False
                
                #reset resolution count
                self.emc_resuolutionCount=0
                
                #start worker
                self.worker.do_work(self.worker.Functions.rotate)
                
                #set test running to true
                self.emcTestRunning=True
            
                #configure specan attenuation and gain
                self.worker.specan.sh.configureLevel(0, 0)
                self.worker.specan.sh.configureGain('auto')
                
               
            #===================================================
    
            #=======================================================================
            # get data from specan
            #=======================================================================
            else:
    
                print "getting data"
                #data units
                
                #setup 30-1000Mhz specan settings
                if (self.emc_regs=='FCC'):
                    self.worker.specan.sh.configureAcquisition("min-max","log-scale" )
                    self.worker.specan.sh.configureSweepCoupling(120e3,120e3,0.01,"native","no-spur-reject")
                else:
                    self.worker.specan.sh.configureAcquisition("min-max","log-scale")
                    self.worker.specan.sh.configureSweepCoupling(120e3,120e3,0.01,"native","no-spur-reject")
                    
                #set specan center and span    
                self.worker.specan.set_frequency(CENTERlow,SPANlow)
                
                #get data for 30 ~ 1000 MHz
                dataReturn=self.worker.specan.get_full_sweep()
                
                #get bin size in order to calculate frequencies
                traceInfo=self.worker.specan.sh.queryTraceInfo()
                binsize=traceInfo["arr-bin-size"]
                
                #calculate frequencies from trace info
                print "getting frequencies"
    
                dataiter=0
                freqArray=[]
                
                for i in dataReturn:
                    freqArray.append(int(STARTlow+(dataiter*binsize)))
                    dataiter+=1
                
                #if FCC selected get  1 ~ 6 GHz data
                if (self.emc_regs=='FCC'):
                    
                    #setup FCC settings for specan
                    self.worker.specan.sh.configureAcquisition("average","log-scale" )
                    self.worker.specan.sh.configureSweepCoupling(1.262e6,1.262e6,0.01,"native","no-spur-reject")
                    
                    #cet center and span for 1 ~ 6 GHz
                    self.worker.specan.set_frequency(CENTERhigh,SPANhigh)
                    
                    #get data from specan
                    highDataReturn=self.worker.specan.get_full_sweep()
                    
                    #add new data to low bandwitdth data
                    dataReturn=np.concatenate((dataReturn,highDataReturn), axis = 1)
    
                    #calculate frequencies from trace info
                    traceInfo=self.worker.specan.sh.queryTraceInfo()
                    binsize=traceInfo["arr-bin-size"]
                    
                    dataiter=0
                    for i in highDataReturn:
                        freqArray.append(int(STARThigh+(dataiter*binsize)))
                        dataiter+=1
                
                #=======================================================================
                # setup plot data
                #=======================================================================
                
                #add cal to full data array and interpolate between nonconsecutive calibration points
                
                full_CalArray=[]
                print "Create full calibration array"
                dataiter=0
                calIter=0
                for element in dataReturn:
                    if calIter<len(self.calArray)-2:
                        val=(self.interp(freqArray[dataiter], self.calFreqArray[calIter], self.calArray[calIter], self.calFreqArray[calIter+1], self.calArray[calIter]))
                    else:
                        val=(self.calArray[calIter]) 
                    full_CalArray.append(val)
                    
                    if freqArray[dataiter]<=self.calFreqArray[calIter]:
                        dataiter+=1
                    else:
                        calIter+=1
                        if calIter>len(self.calArray)-1:
                            calIter-=1
                            
                full_CalArray=np.array(full_CalArray)
                
                #add calibration data to data array
                full_dataPlusCal=[]
                print "adding calibration to data"
                dataiter=0
                for element in dataReturn:
                    full_dataPlusCal.append(dataReturn[dataiter]+full_CalArray[dataiter])
                    dataiter+=1        
                full_dataPlusCal=np.array(full_dataPlusCal)
                 
                
                #FULL convert to dBuVm
                full_dataTodBuVm=[]
                print "Converting to dBuVm"
                dataiter=0
                for i in full_dataPlusCal:
                    full_dataTodBuVm.append(self.get_fieldStrength_dBuVm(full_dataPlusCal[dataiter]))
                    dataiter+=1
                full_dataTodBuVm=np.array(full_dataTodBuVm)  
                
                
                #get full limit array for plot    
                fullLimitArray = []
                print "creating limit array"
                dataiter=0
                for i in freqArray:
                    fullLimitArray.append(self.get_emcTestLimit(i))
                    dataiter+=1  
                fullLimitArray=np.array(fullLimitArray) 
                
                #=======================================================================
                # plot results
                #=======================================================================
                print "plotting"
                
                #clear plot for new data
                self.emcPlot.clear()
                
                #set x axis scaling to logrythmic
    #             self.emcPlot.set_xscale("log")
                
                #add grid to plot
                self.emcPlot.grid(True,which='both',ls='-')
                
                #plot calibrated data converted to dbuV/m    
                self.emcPlot.plot(freqArray,full_dataTodBuVm,ls='-',lw=.5,label='Field Strength')
                
                #plot limit
                self.emcPlot.plot(freqArray,fullLimitArray,lw=2,color='r',ls='-',label=self.emc_regs + ' Class ' + self.emc_class + " Limit  at "+str(self.cal.cal_dist)+"m")
                
                #if user selected a error margin, plot it
                if(margin!=0):
                    marginArray = fullLimitArray-margin
                    self.emcPlot.plot(freqArray,marginArray,lw=2,color='#ff6666',label="Error Margin", ls='--')    
                
                #===================================================================
                # configure result plot
                #===================================================================
                
                #set plot labels
                self.emcPlot.set_xlabel("Frequency Hz")
                self.emcPlot.set_ylabel("Field Strength (dBuV/m)")
                
                #format set axis values to scientific notation
                self.emcPlot.xaxis.set_major_formatter(ticker.FormatStrFormatter('%0.0e'))
                
                #set plot range
                ymin,ymax=self.emcPlot.get_ylim()        
                self.emcPlot.set_ylim([ymin-10,ymax+10])
                
                # adjust spacing of plots elements
                self.figEmc.subplots_adjust(wspace=.1,bottom=.2)
                
                #set title
                if(self.emc_polarity==POL_V):
                    self.emcPlot.set_title('EMC Pre-Compliance Test (Vertical Antenna Polarity)\nField Strength vs. Frequency',fontsize=14,fontweight=200)
                else:                
                    self.emcPlot.set_title('EMC Pre-Compliance Test (Horizontal Antenna Polarity)\nField Strength vs. Frequency',fontsize=14,fontweight=200)
                    
                #set x axis limis based on test configuration  
                if (self.emc_regs=='FCC'):
                    self.emcPlot.set_xlim(0,6e9)
                else:
                    self.emcPlot.set_xlim(0,1e9)
                
                #setup legend
#                 self.emcPlot.legend(fontsize=8,loc="best")
                
                #set tight layout
                self.figEmc.tight_layout()
                
                #draw plot
                self.emcCanvas.draw()
                self.emc_testComplete=True #set testComplete to true because test is run
                
                
                #===================================================================
                # Run test
                #===================================================================
                print 'running test'
                
                #setup iterators
                dataiter=0
                FailNum=0
                
                row=self.emc_reportRow+1
                angle=tableAngle
            
                #initialize passfail to  Passso test will change if a failure occurs
                passfail="Pass"
                
                #perform test on calibrated data
                for i in full_dataTodBuVm:
                    testval=i
                    testValLim=fullLimitArray[dataiter]
                    
                    #if failures exist find highest consecutive value and add them to their respective polarity results list
                    if (testval+margin) > testValLim:
                        if (checkingFailSpan):
                            
                            if((testval)-testValLim)>(self.EMC_DATASET[4]-self.EMC_DATASET[5]):
#                             if testval>self.EMC_DATASET[4]:
                                self.EMC_DATASET=[]
                                self.EMC_DATASET.append(angle)
                                self.EMC_DATASET.append(freqArray[dataiter])
                                self.EMC_DATASET.append(dataReturn[dataiter])
                                self.EMC_DATASET.append(full_CalArray[dataiter])
                                self.EMC_DATASET.append(testval)
                                self.EMC_DATASET.append(fullLimitArray[dataiter])      
                        else:
                            checkingFailSpan=True
                            self.EMC_DATASET=[]
                            self.EMC_DATASET.append(angle)
                            self.EMC_DATASET.append(freqArray[dataiter])
                            self.EMC_DATASET.append(dataReturn[dataiter])
                            self.EMC_DATASET.append(full_CalArray[dataiter])
                            self.EMC_DATASET.append(testval)
                            self.EMC_DATASET.append(fullLimitArray[dataiter])
                        row+=1
                        passfail="Fail"
                        FailNum+=1
                    dataiter+=1    
#                     else:
#                         if (checkingFailSpan):
#                             checkingFailSpan=False
                if(FailNum>0):
                    if(self.emc_polarity==POL_V):
                        self.EMC_Vertical_Results.append(self.EMC_DATASET)
                        self.emcPlot.plot(self.EMC_DATASET[1], self.EMC_DATASET[4], lw=0, marker='v', markersize=5, color='black', label='Peak Failure')
                        print "dataset[1]", self.EMC_DATASET[1]
                        print "dataset[2]", self.EMC_DATASET[4]
                        self.EMC_VerticalFailureNum+=1
                    else:                
                        self.EMC_Horizontal_Results.append(self.EMC_DATASET)
                        self.emcPlot.plot(self.EMC_DATASET[1], self.EMC_DATASET[4], lw=0, marker='v', markersize=5, color='black',label='Peak Failure')
                        print "dataset[1]", self.EMC_DATASET[1]
                        print "dataset[2]", self.EMC_DATASET[4]
                        self.EMC_HorizontalFailureNum+=1
                    self.emc_reportRow+=1
                        
                self.emcPlot.legend(fontsize=8,loc="best")    
                    #if there are more than 2000 failures stop test    
                    
#                     if FailNum>2000:
#                         print "Failure list is greater than 2000 data points\n\t---stopping test---"
#                         self.emc_failureLimitReached=True
#                         EMC_DATASET=[]
#                         EMC_DATASET.append(angle)
#                         EMC_DATASET.append("More than 2000 Failures Identified")
#                         EMC_DATASET.append('')
#                         EMC_DATASET.append('')
#                         EMC_DATASET.append('')
#                         EMC_DATASET.append('')
#                         if(self.emc_polarity==POL_V):
#                             self.EMC_Vertical_Results.append(EMC_DATASET)
#                             self.EMC_VerticalFailureNum+=1
#                         else:                
#                             self.EMC_Horizontal_Results.append(EMC_DATASET)
#                             self.EMC_HorizontalFailureNum+=1
#                         self.emc_reportRow+=1
#                         break
                
                #update Horizontal polarity results label
                if(self.EMC_HorizontalFailureNum>0):
                    if(self.emc_failureLimitReached):
                        self.emc_HtestFailNum.setText('<span style="  color:red; font-size:14pt; font-weight:400;">Horizontal Failures: Greater than '+str(self.EMC_HorizontalFailureNum)+' (Failure Limit Reached)</span>')
                    else:
                        self.emc_HtestFailNum.setText('<span style="  color:red; font-size:14pt; font-weight:400;">Horizontal Failures:<t/> '+str(self.EMC_HorizontalFailureNum)+'</span>')
                else:
                    self.emc_HtestFailNum.setText('<span style="  color:lime; font-size:14pt; font-weight:400;">-----No Horizontal Failures Identified-----</span>')
                    
                #update Vertical Polarity results label    
                if(self.EMC_VerticalFailureNum>0):
                    if(self.emc_failureLimitReached):
                        self.emc_VtestFailNum.setText('<span style="  color:red; font-size:14pt; font-weight:400;">Vertical Failures:<t/> Greater than '+str(self.EMC_VerticalFailureNum)+' (Failure Limit Reached)</span>')
                    else:
                        self.emc_VtestFailNum.setText('<span style="  color:red; font-size:14pt; font-weight:400;">Vertical Failures:<t/> '+str(self.EMC_VerticalFailureNum)+'</span>')
                else:
                    self.emc_VtestFailNum.setText('<span style="  color:lime; font-size:14pt; font-weight:400;">-----No Vertical Failures Identified-----</span>')
                
                #display pass or fail
                if passfail=='Pass': 
                    print 'EMC Test complete--PASS'
                    self.emc_testResults.setText('<span style="  color:lime; font-size:14pt; font-weight:400;">-----Test At '+str(angle)+' Degrees Passed-----</span>')
                    
                elif passfail=='Fail' : 
                    self.emc_testResults.setText('<span style="  color:red; font-size:14pt; font-weight:400;">-----Test At '+str(angle)+' Degrees Failed-----</span>')
                    
                else:
                    print 'EMC Test Will Not Run!--INSUFFICIENT DATA'
                    self.emc_testResults.setText('<span style="  color:yellow; font-size:14pt; font-weight:400;">-----No Test Data-----</span>')
                    passfail="Fail"
                
                #save figures for report    
                if(self.emc_polarity==POL_V):
                    self.emcCanvas.print_figure("temp_emcVFig.png", dpi=100, facecolor=self.figEmc.get_facecolor())       
                else:      
                    self.emcCanvas.print_figure("temp_emcHFig.png", dpi=100, facecolor=self.figEmc.get_facecolor()) 
                    
                self.emc_testComplete=True              #used to allow app to save EMC testing data
                self.emc_resuolutionCount+=1            #count resolutions 
                #enable run test button after test      #enable run test button after test
                self.cal.update_calibration()           #update cal ibration incase any value wer changes during test
                self.emc_testStatus=passfail            #set status of test    
                
                
                return passfail                         #set return value to test status
  
    def get_farField(self):
        #=======================================================================
        #
        #          Name:    get_farField
        #
        #    Parameters:    None
        #
        #        Return:    (float)farField
        #
        #   Description:    this function returns the far filed of the antenna given the test frequency
        #
        #=======================================================================
        DUTDiameter=float(self.e_emc_resolution.text())
        waveLength = 2.997e8/self.cal.cal_freq
        
        farField=(2*DUTDiameter**2)/waveLength

#         print "far-field: ", farField, " m"
        
        return farField
  
    def get_fieldStrength_dBuVm(self,value):             #Take Recieved power and distance, and convert to Field strength
        'Takes Recieved power and distance, and convert to Field strength'
        #=======================================================================
        #          Name:    get_fieldStrength_dBuVm
        #
        #    Parameters:    (float)value 
        #
        #        Return:    (float)fieldStrength
        #
        #   Description:    this function takes the collected data in dBm and returns the 
        #                   electrical field strength in dBuV/m
        #
        #=======================================================================
        
        EiRP=(10**((float(value)-30)/10))                       #convert dBm to W (EiRP)
        fieldStrength=(math.sqrt(30*EiRP))/self.cal.cal_dist    #calculate field strength from distance and EiRP
        fieldStrength=20*math.log10(fieldStrength/1e-6)         #convert V/m to dBuV/m
        
        return fieldStrength
    
    def get_fieldStrength_uVm(self,value):#Take Recieved power and distance, and convert to Field strength
        'Takes Recieved power and distance, and convert to Field strength'
        #=======================================================================
        #          Name:    get_fieldStrength_uVm
        #
        #    Parameters:    (float)value 
        #
        #        Return:    (float)fieldStrength
        #
        #   Description:    this function takes the collected data in dBm and returns the 
        #                   electrical field strength in uV/m
        #
        #=======================================================================
        
        EiRP=(10**((float(value)-30)/10))                       #convert dBm to W (EiRP)
        fieldStrength=(math.sqrt(30*EiRP))/self.cal.cal_dist    #calculate field strength at distance and EiRP    
        fieldStrength=fieldStrength*1e6                         #convert from V/m to uV/m
        
        return fieldStrength
    
    def set_emcAntPolarity(self):
        #=======================================================================
        #
        #          Name:    set_emcAntPolarity
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    this function set the polarity of the calibrated RX antenna 
        #                    based on user input
        #
        #=======================================================================
        if self.r_Vertical.isChecked():
            self.emc_polarity=POL_V
        else:
            self.emc_polarity=POL_H
    
    def set_emcRegulations(self):#sets which set of EMC Regulations should be tested
        #=======================================================================
        #          Name:    set_emcRegulations
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    This Function sets the EMC regulations for testign
        #                   Based on user input
        #
        #                   This function also sets the warnings based on the user input and testing 
        #                   parameters
        #
        #=======================================================================
        'sets which set of EMC Regulations should be tested'
        
        dist=[3,10,30]#holds required testing distance
        
        #initialize min and max testing frequencies
        minFreq=30e6
        maxFreq=3e9
        
        if self.r_fcc.isChecked():
        #=======================================================================
        # set fcc testing settings
        #=======================================================================
            self.emc_regs='FCC'
            
            #set min and max testing frequencies
            minFreq=30e6                    #30 MHz
            maxFreq=40e9                    #40 GHz
            
            #set desting distance
            if self.r_classA.isChecked():
                self.emc_class='A'
                dist=[3,10,30]                     #emc testing distance should be 10 meters
            else:
                self.emc_class='B'
                dist=[3,10,10]                     #emc testing distance should be 3 meters
                
        elif self.r_cispr.isChecked():
        #=======================================================================
        # set cispr testing settings
        #=======================================================================
            self.emc_regs='CISPR'
            
            #set min and max testing frequencies
            minFreq=30e6
            maxFreq=1e9
            
            #set desting distance
            if self.r_classA.isChecked():
                self.emc_class='A'
                dist=[3,10,30]                     #emc testing distance should be 30 meters
            else:
                self.emc_class='B'
                dist=[3,10,10]                     #emc testing distance should be 10 meters
        
        #=======================================================================
        # set distance warning label text
        #=======================================================================
        if (self.cal.cal_dist==dist[0] or self.cal.cal_dist==dist[1] or self.cal.cal_dist==dist[2]):
            self.emc_distWarning.setText('<span style="  color:lime; font-size:11pt; font-weight:400;">--Testing Distance ('+ str(self.cal.cal_dist) +' m) Good-- Ready to run test</span>')
        else:
            self.emc_distWarning.setText('<span style="  color:Red; font-size:11pt; font-weight:400;">--WARNING--</br>Testing distance set to '+str(self.cal.cal_dist)+' m, ' + self.emc_regs+ ' Class '+ self.emc_class+' Testing Requires '+str(dist[0])+', '+str(dist[1])+' or '+str(dist[2])+' m. </span>')
        
    def create_status_bar(self):#create status bar at bottom of aplication
        #=======================================================================
        #          Name:    create_status_bar
        #
        #    Parameters:    None
        #
        #        Return:    None
        #
        #   Description:    creates status bar at bottom of application
        #
        #=======================================================================
        self.status_text = QLabel("Click Setup to find instruments.")
        self.statusBar().addWidget(self.status_text, 1)
        
    def create_menu(self):#create menu at top of application 
        #=======================================================================
        #          Name:    create_menu
        #
        #    Parameters:    None    
        #
        #        Return:    None
        #
        #   Description:    This function creates the menu at the top of the application
        #
        #=======================================================================    
        self.file_menu = self.menuBar().addMenu("&File")
        
        open_file_action = self.create_action("&Open Report",
            shortcut="Ctrl+O", slot=self.click_open, 
            tip="Load a CSV file, first row is Title, first column is deg, subsequent columns mag")
        
        save_csv_action = self.create_action("&Save Report",
            shortcut="Ctrl+S", slot=self.save_report, 
            tip="Save a CSV file, first row is Title, first column is deg, subsequent columns mag")
        
        save_file_action = self.create_action("Save plot",
            slot=self.save_plot, 
            tip="Save the plot")
        quit_action = self.create_action("&Quit", slot=self.close, 
            shortcut="Ctrl+Q", tip="Close the application")
        
        self.add_actions(self.file_menu, 
            (open_file_action,save_csv_action,save_file_action, None, quit_action))
        
        self.help_menu = self.menuBar().addMenu("&Help")
        about_action = self.create_action("&About", 
            shortcut='F1', slot=self.on_about, 
            tip='About')
        
        self.add_actions(self.help_menu, (about_action,))

#     def fill_dataArray(self):#fill data arrays so they are all the same size
#         #=======================================================================
#         #          Name:    fill_dataArray
#         #    
#         #    Parameters:    None
#         #
#         #        Return:    None
#         #
#         #   Description:    this functions fills the data arrays with zeros which is needed for 
#         #                    saving reports and plotting
#         #
#         #=======================================================================
#        
#         #find longest array
#         longest=len(self.zRawData)
#         if(longest<len(self.xRawData)):
#             longest=len(self.xRawData)
#             
#         if(longest<len(self.yRawData)):
#             longest=len(self.yRawData)
#             
#         if(longest<len(self.zCalData)):
#             longest=len(self.zCalData)
#             
#         if(longest<len(self.xCalData)):
#             longest=len(self.xCalData)
#             
#         if(longest<len(self.yCalData)):
#             longest=len(self.yCalData)
#             
#         if(longest<len(self.angles)):
#             longest=len(self.angles)
#             
#         #set length of all arrays to be the same as the longest
#         short=longest-(len(self.zRawData))
#         if short>0:
#             for i in range(0,short):
#                 self.zRawData.append(0)
#                  
#         short=longest-(len(self.xRawData))
#         if short>0:
#             for i in range(0,short):
#                 self.xRawData.append(0)
#                  
#         short=longest-(len(self.yRawData))
#         if short>0:
#             for i in range(0,short):
#                 self.yRawData.append(0)
#                  
#         short=longest-(len(self.zCalData))
#         if short>0:
#             for i in range(0,short):
#                 self.zCalData.append(0)
#                  
#         short=longest-(len(self.xCalData))
#         if short>0:
#             for i in range(0,short):
#                     self.xCalData.append(0)
#                  
#         short=longest-(len(self.yCalData))
#         if short>0:
#             for i in range(0,short):
#                 self.yCalData.append(0)
#          
#         short=longest-(len(self.angles))
#         if short>0:
#             for i in range(0,short):
#                 self.angles.append((i*3.6))
        
    def add_actions(self, target, actions):#do something..apparently 
        #=======================================================================
        #          Name:    add_actions
        #
        #    Parameters:    target , actions
        #
        #        Return:    None
        #
        #   Description:    'Unknown'
        #
        #=======================================================================
        for action in actions:
            if action is None:
                target.addSeparator()
            else:
                target.addAction(action)

    def create_action(  self, text, slot=None, shortcut=None, 
                        icon=None, tip=None, checkable=False, 
                        signal="triggered()"):
        #=======================================================================
        #          Name:    create_action
        #
        #    Parameters:    text , slot , shortcut, icon, tip , checkable, signal
        #
        #        Return:    action
        #
        #   Description:    'Unknown'
        #
        #=======================================================================
        action = QAction(text, self)
        if icon is not None:
            action.setIcon(QIcon(":/%s.png" % icon))
        if shortcut is not None:
            action.setShortcut(shortcut)
        if tip is not None:
            action.setToolTip(tip)
            action.setStatusTip(tip)
        if slot is not None:
            self.connect(action, SIGNAL(signal), slot)
        if checkable:
            action.setCheckable(True)
        return action

def main():
    #=======================================================================
    #          Name:    main
    #
    #    Parameters:    None
    #
    #        Return:    None
    #
    #   Description:    this function creates and executes the main application
    #
    #=======================================================================
    app = QApplication(sys.argv)                        #create Qapplication (pyqt)
    app.setStyle(QStyleFactory.create("plastique"))     #change style for readability
    form = AppForm()                                    #create Qmainwindow subclass(pyqt)
    form.resize(800,600)
    form.move(100,10)                                    #move app to upper left corner of display
    #form.showMaximized()
    form.show()                                         #Make application visible (PYQT)
    app.exec_()                                         #enter main loop of QApplication class (pyqt)


if __name__ == "__main__":
    main()
